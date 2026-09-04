"""
student_directory.py
====================
Authoritative Students Directory for AR School Management System.

- Student list (Treeview) + search + filters
- THIS MONTH fee status: Paid / Pending / Partial / Overdue / No Cycle
- TOTAL BALANCE (all months outstanding)
- Export Excel, Edit, Remove, Profile, ID Card
"""

from __future__ import annotations

import os
import shutil
from datetime import datetime

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import db
import rbac
import theme
import reports
import accounting
import student_lifecycle


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_float(raw_text, field_label, default=None, parent=None):
    text = (raw_text or "").strip()
    if not text:
        if default is not None:
            return default, True
        messagebox.showerror("Invalid Input", f"{field_label} is required.", parent=parent)
        return None, False
    try:
        return float(text), True
    except ValueError:
        messagebox.showerror(
            "Invalid Input",
            f"{field_label} must be a valid number (e.g. 1500 or 1500.50).",
            parent=parent,
        )
        return None, False


def _log_activity(username, action):
    try:
        db.run(
            "INSERT INTO audit_logs (username, action, timestamp) VALUES (?, ?, ?)",
            (username, action, datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            commit=True,
        )
    except Exception as e:
        print(f"Audit Log Error: {e}")


def _current_year_month():
    now = datetime.now()
    return now.year, now.month


def _fee_maps_for_directory():
    """
    total_outstanding  : student_id -> sum(amount_due - amount_paid) ALL cycles
    current_month_info : student_id -> {amount_due, amount_paid, balance, status}
                         status: PAID | PENDING | PARTIAL | OVERDUE | NO_CYCLE
    """
    year, month = _current_year_month()
    total_outstanding = {}
    current_month_info = {}
    ym = f"{year}-{month:02d}"
    month_str = f"{month:02d}"

    # ---- Total outstanding (all cycles) ----
    try:
        rows = db.run(
            "SELECT student_id, COALESCE(SUM(amount_due - amount_paid), 0) "
            "FROM fee_cycles GROUP BY student_id",
            fetchall=True,
        ) or []
        for sid, bal in rows:
            total_outstanding[sid] = float(bal or 0)
    except Exception:
        pass

    def _process_cycle_rows(rows):
        for row in rows or []:
            try:
                sid = row[0]
                due = float(row[1] or 0)
                paid = float(row[2] or 0)
                st = (row[3] if len(row) > 3 else "") or ""
            except (IndexError, TypeError, ValueError):
                continue
            bal = due - paid
            st_up = st.strip().upper()
            if st_up in ("PAID", "OVERDUE", "PARTIAL", "PENDING"):
                status = st_up
            elif bal <= 0 and due > 0:
                status = "PAID"
            elif paid > 0 and bal > 0:
                status = "PARTIAL"
            elif bal > 0:
                status = "PENDING"
            else:
                status = "NO_CYCLE"
            current_month_info[sid] = {
                "amount_due": due,
                "amount_paid": paid,
                "balance": bal,
                "status": status,
            }

    # ---- Current month: try many common column layouts ----
    queries = [
        (
            "SELECT student_id, COALESCE(amount_due,0), COALESCE(amount_paid,0), "
            "COALESCE(status,'') FROM fee_cycles WHERE year=? AND month=?",
            (year, month),
        ),
        (
            "SELECT student_id, COALESCE(amount_due,0), COALESCE(amount_paid,0), "
            "COALESCE(status,'') FROM fee_cycles WHERE period_year=? AND period_month=?",
            (year, month),
        ),
        (
            "SELECT student_id, COALESCE(amount_due,0), COALESCE(amount_paid,0), "
            "COALESCE(status,'') FROM fee_cycles WHERE fee_year=? AND fee_month=?",
            (year, month),
        ),
        (
            "SELECT student_id, COALESCE(amount_due,0), COALESCE(amount_paid,0), "
            "COALESCE(status,'') FROM fee_cycles WHERE cycle_year=? AND cycle_month=?",
            (year, month),
        ),
        (
            "SELECT student_id, COALESCE(amount_due,0), COALESCE(amount_paid,0), "
            "COALESCE(status,'') FROM fee_cycles WHERE cycle_month=?",
            (ym,),
        ),
        (
            "SELECT student_id, COALESCE(amount_due,0), COALESCE(amount_paid,0), "
            "COALESCE(status,'') FROM fee_cycles WHERE fee_month=?",
            (ym,),
        ),
        (
            "SELECT student_id, COALESCE(amount_due,0), COALESCE(amount_paid,0), "
            "COALESCE(status,'') FROM fee_cycles WHERE month_label=?",
            (ym,),
        ),
        (
            "SELECT student_id, COALESCE(amount_due,0), COALESCE(amount_paid,0), "
            "COALESCE(status,'') FROM fee_cycles WHERE period=?",
            (ym,),
        ),
        (
            "SELECT student_id, COALESCE(amount_due,0), COALESCE(amount_paid,0), "
            "COALESCE(status,'') FROM fee_cycles WHERE month=?",
            (ym,),
        ),
        (
            "SELECT student_id, COALESCE(amount_due,0), COALESCE(amount_paid,0), "
            "COALESCE(status,'') FROM fee_cycles "
            "WHERE strftime('%Y', due_date)=? AND strftime('%m', due_date)=?",
            (str(year), month_str),
        ),
        (
            "SELECT student_id, COALESCE(amount_due,0), COALESCE(amount_paid,0), "
            "COALESCE(status,'') FROM fee_cycles "
            "WHERE strftime('%Y', cycle_date)=? AND strftime('%m', cycle_date)=?",
            (str(year), month_str),
        ),
        (
            "SELECT student_id, COALESCE(amount_due,0), COALESCE(amount_paid,0), "
            "COALESCE(status,'') FROM fee_cycles "
            "WHERE strftime('%Y', created_at)=? AND strftime('%m', created_at)=?",
            (str(year), month_str),
        ),
        (
            "SELECT student_id, COALESCE(amount_due,0), COALESCE(amount_paid,0), "
            "COALESCE(status,'') FROM fee_cycles "
            "WHERE strftime('%Y', fee_date)=? AND strftime('%m', fee_date)=?",
            (str(year), month_str),
        ),
        (
            "SELECT student_id, COALESCE(amount_due,0), COALESCE(amount_paid,0), "
            "COALESCE(status,'') FROM fee_cycles "
            "WHERE due_date LIKE ? OR created_at LIKE ? OR cycle_date LIKE ? "
            "OR fee_date LIKE ? OR period LIKE ?",
            (f"{ym}%", f"{ym}%", f"{ym}%", f"{ym}%", f"{ym}%"),
        ),
    ]

    for sql, params in queries:
        if current_month_info:
            break
        try:
            rows = db.run(sql, params, fetchall=True)
            if rows:
                _process_cycle_rows(rows)
        except Exception:
            continue

    return total_outstanding, current_month_info


def _status_display(code: str) -> str:
    mapping = {
        "PAID": "Paid",
        "PENDING": "Pending",
        "PARTIAL": "Partial",
        "OVERDUE": "Overdue",
        "NO_CYCLE": "No Cycle",
    }
    return mapping.get((code or "").upper(), code or "—")


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_student_directory_into(parent, user_role, current_user, app_callbacks=None):
    """
    Build full Students Directory into `parent`.

    app_callbacks optional:
      open_admission()
      open_student_profile(student_id)
      on_students_changed()
    """
    controller = _StudentDirectoryController(
        parent, user_role, current_user, app_callbacks or {}
    )
    controller.build()
    return controller


# ---------------------------------------------------------------------------
# Controller
# ---------------------------------------------------------------------------

class _StudentDirectoryController:
    def __init__(self, parent, user_role, current_user, callbacks):
        self.parent = parent
        self.user_role = user_role
        self.current_user = current_user
        self.callbacks = callbacks
        self.root = parent.winfo_toplevel()

        self.tree = None
        self.ent_search = None
        self.cmb_fee_filter = None
        self.cmb_class_filter = None
        self.cmb_status_filter = None
        self.lbl_student_count = None
        self.show_archived_var = None
        # Stat card value labels (reference-style dashboard cards)
        self._stat_total = None
        self._stat_active = None
        self._stat_pending = None
        self._stat_classes = None

    def _make_stat_card(self, parent, icon_text, icon_bg, title, subtitle, value_attr):
        """Reference-style white metric card with coloured icon badge."""
        card = tk.Frame(
            parent, bg=theme.WHITE,
            highlightbackground="#e2e8f0", highlightthickness=1,
            padx=14, pady=12,
        )
        card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))

        inner = tk.Frame(card, bg=theme.WHITE)
        inner.pack(fill=tk.X)

        icon_box = tk.Label(
            inner, text=icon_text, font=("Segoe UI", 14),
            bg=icon_bg, fg="white", width=3, height=1, padx=6, pady=4,
        )
        icon_box.pack(side=tk.LEFT, padx=(0, 12))

        text_col = tk.Frame(inner, bg=theme.WHITE)
        text_col.pack(side=tk.LEFT, fill=tk.X, expand=True)

        tk.Label(
            text_col, text=title, font=("Segoe UI", 9),
            bg=theme.WHITE, fg="#64748b",
        ).pack(anchor="w")
        val_lbl = tk.Label(
            text_col, text="—", font=("Segoe UI", 18, "bold"),
            bg=theme.WHITE, fg="#0f172a",
        )
        val_lbl.pack(anchor="w")
        tk.Label(
            text_col, text=subtitle, font=("Segoe UI", 8),
            bg=theme.WHITE, fg="#94a3b8",
        ).pack(anchor="w")
        setattr(self, value_attr, val_lbl)
        return card

    def build(self):
        can_add = rbac.can(self.user_role, "student.add")
        can_delete = rbac.can(self.user_role, "student.delete")
        can_edit = rbac.can(self.user_role, "student.edit")
        can_view = rbac.can(self.user_role, "student.view")
        can_fee = rbac.can(self.user_role, "student.fee.view")

        # Ensure parent background matches reference light canvas
        try:
            self.parent.configure(bg="#f1f5f9")
        except Exception:
            pass

        # ---- Page header (title + subtitle + Add Student) — matches reference ----
        header = tk.Frame(self.parent, bg="#f1f5f9", padx=4, pady=4)
        header.pack(fill=tk.X, padx=8, pady=(8, 4))

        title_col = tk.Frame(header, bg="#f1f5f9")
        title_col.pack(side=tk.LEFT, fill=tk.Y)
        tk.Label(
            title_col, text="Students", font=("Segoe UI", 20, "bold"),
            bg="#f1f5f9", fg="#0f172a",
        ).pack(anchor="w")
        tk.Label(
            title_col, text="Manage students, admissions and academic records",
            font=("Segoe UI", 9), bg="#f1f5f9", fg="#64748b",
        ).pack(anchor="w")

        if can_add:
            add_btn = tk.Button(
                header, text="+  Add Student", command=self._open_admission,
                bg="#2563eb", fg="white", font=("Segoe UI", 10, "bold"),
                bd=0, padx=16, pady=8, cursor="hand2", activebackground="#1d4ed8",
                activeforeground="white",
            )
            add_btn.pack(side=tk.RIGHT, pady=4)

        # ---- Stat cards row (Total / Active / Pending Fees / Classes) ----
        stats_row = tk.Frame(self.parent, bg="#f1f5f9")
        stats_row.pack(fill=tk.X, padx=8, pady=(8, 10))

        self._make_stat_card(
            stats_row, "👥", "#3b82f6", "Total Students", "All students", "_stat_total",
        )
        self._make_stat_card(
            stats_row, "🛡", "#22c55e", "Active Students", "Active this month", "_stat_active",
        )
        self._make_stat_card(
            stats_row, "⚠", "#f59e0b", "Pending Fees", "Students have pending dues", "_stat_pending",
        )
        last = self._make_stat_card(
            stats_row, "📋", "#a855f7", "Total Classes", "All classes", "_stat_classes",
        )
        last.pack_configure(padx=(0, 0))

        # ---- Main white content card (filters + actions + table) ----
        main_card = tk.Frame(
            self.parent, bg=theme.WHITE,
            highlightbackground="#e2e8f0", highlightthickness=1,
        )
        main_card.pack(fill=tk.BOTH, expand=True, padx=8, pady=(0, 10))

        body = tk.Frame(main_card, bg=theme.WHITE, padx=14, pady=12)
        body.pack(fill=tk.BOTH, expand=True)

        # Filter / search bar — reference layout
        filter_bar = tk.Frame(body, bg=theme.WHITE)
        filter_bar.pack(fill=tk.X, pady=(0, 10))

        search_wrap = tk.Frame(
            filter_bar, bg=theme.WHITE,
            highlightbackground="#e2e8f0", highlightthickness=1,
        )
        search_wrap.pack(side=tk.LEFT, padx=(0, 12))
        self.ent_search = tk.Entry(
            search_wrap, font=("Segoe UI", 10), bd=0, width=28,
            bg=theme.WHITE, fg="#0f172a", insertbackground="#0f172a",
        )
        self.ent_search.pack(side=tk.LEFT, ipady=7, padx=(10, 4))
        self.ent_search.bind("<KeyRelease>", self.load_table)
        tk.Label(
            search_wrap, text="🔍", font=("Segoe UI", 10),
            bg=theme.WHITE, fg="#94a3b8",
        ).pack(side=tk.LEFT, padx=(0, 8))

        fee_col = tk.Frame(filter_bar, bg=theme.WHITE)
        fee_col.pack(side=tk.LEFT, padx=(0, 10))
        tk.Label(
            fee_col, text="Fee Status", font=("Segoe UI", 8),
            bg=theme.WHITE, fg="#64748b",
        ).pack(anchor="w")
        self.cmb_fee_filter = ttk.Combobox(
            fee_col,
            values=[
                "All Fees",
                "This Month Paid",
                "This Month Pending",
                "This Month Partial",
                "This Month Overdue",
                "Has Total Balance",
                "No Total Balance",
            ],
            state="readonly",
            width=16,
            font=("Segoe UI", 9),
        )
        self.cmb_fee_filter.set("All Fees")
        self.cmb_fee_filter.pack()
        self.cmb_fee_filter.bind("<<ComboboxSelected>>", self.load_table)

        class_col = tk.Frame(filter_bar, bg=theme.WHITE)
        class_col.pack(side=tk.LEFT, padx=(0, 10))
        tk.Label(
            class_col, text="Class", font=("Segoe UI", 8),
            bg=theme.WHITE, fg="#64748b",
        ).pack(anchor="w")
        self.cmb_class_filter = ttk.Combobox(
            class_col, values=["All Classes"], state="readonly",
            width=12, font=("Segoe UI", 9),
        )
        self.cmb_class_filter.set("All Classes")
        self.cmb_class_filter.pack()
        self.cmb_class_filter.bind("<<ComboboxSelected>>", self.load_table)

        status_col = tk.Frame(filter_bar, bg=theme.WHITE)
        status_col.pack(side=tk.LEFT, padx=(0, 10))
        tk.Label(
            status_col, text="Status", font=("Segoe UI", 8),
            bg=theme.WHITE, fg="#64748b",
        ).pack(anchor="w")
        self.cmb_status_filter = ttk.Combobox(
            status_col,
            values=["Active Only", "Archived Only", "All Status"],
            state="readonly",
            width=12,
            font=("Segoe UI", 9),
        )
        self.cmb_status_filter.set("Active Only")
        self.cmb_status_filter.pack()
        self.cmb_status_filter.bind("<<ComboboxSelected>>", self.load_table)

        self.show_archived_var = tk.BooleanVar(value=False)

        clear_btn = tk.Button(
            filter_bar, text="↻  Clear Filters", command=self._clear_filters,
            bg=theme.WHITE, fg="#64748b", font=("Segoe UI", 9),
            bd=0, padx=10, pady=6, cursor="hand2",
            activebackground="#f1f5f9", activeforeground="#0f172a",
        )
        clear_btn.pack(side=tk.RIGHT, pady=(10, 0))

        # ---- Action toolbar ----
        action_bar = tk.Frame(body, bg=theme.WHITE)
        action_bar.pack(fill=tk.X, pady=(0, 10))

        def _toolbar_btn(parent, text, cmd, bg, fg="white", side=tk.LEFT):
            btn = tk.Button(
                parent, text=text, command=cmd, bg=bg, fg=fg,
                font=("Segoe UI", 9, "bold"), bd=0, padx=12, pady=7,
                cursor="hand2", activeforeground=fg,
            )
            btn.pack(side=side, padx=(0, 6))
            return btn

        if can_add:
            _toolbar_btn(action_bar, "+  Add Student", self._open_admission, "#2563eb")
        if can_edit:
            _toolbar_btn(action_bar, "✎  Edit", self._edit_selected, "#64748b")
        if can_delete:
            _toolbar_btn(action_bar, "⏸  Deactivate", self._remove_selected, "#f59e0b")
        if can_view:
            _toolbar_btn(action_bar, "👤  View Profile", self._open_profile, "#3b82f6")
            _toolbar_btn(action_bar, "🪪  ID Card", self._reprint_id_card, "#64748b")
        _toolbar_btn(
            action_bar, "📊  Export Excel", self._export_excel, "#16a34a", side=tk.RIGHT,
        )

        # ---- Table (id first so get_selected_student_id keeps working) ----
        if can_fee:
            cols = (
                "id", "name", "fname", "class", "phone",
                "month_status", "total_balance", "status",
            )
        else:
            cols = ("id", "name", "fname", "class", "phone", "status")

        table_frame = tk.Frame(body, bg=theme.WHITE)
        table_frame.pack(fill=tk.BOTH, expand=True)

        style = ttk.Style()
        try:
            style.configure(
                "Students.Treeview",
                font=("Segoe UI", 10),
                rowheight=32,
                background=theme.WHITE,
                fieldbackground=theme.WHITE,
                foreground="#0f172a",
            )
            style.configure(
                "Students.Treeview.Heading",
                font=("Segoe UI", 9, "bold"),
                background="#f8fafc",
                foreground="#64748b",
            )
            style.map("Students.Treeview", background=[("selected", "#dbeafe")])
        except Exception:
            pass

        self.tree = ttk.Treeview(
            table_frame, columns=cols, show="headings", style="Students.Treeview",
        )

        headings = {
            "id": "Student ID",
            "name": "Student",
            "fname": "Father Name",
            "class": "Class",
            "phone": "Phone",
            "month_status": "Fee Status",
            "total_balance": "Balance",
            "status": "Status",
        }
        col_widths = {
            "id": 110, "name": 150, "fname": 120, "class": 70, "phone": 115,
            "month_status": 130, "total_balance": 100, "status": 80,
        }
        for col in cols:
            self.tree.heading(col, text=headings.get(col, col.upper()))
            anchor = "w" if col in ("name", "fname") else "center"
            self.tree.column(col, width=col_widths.get(col, 90), anchor=anchor, minwidth=60)

        self.tree.tag_configure("fee_paid", foreground="#16a34a")
        self.tree.tag_configure("fee_pending", foreground="#ea580c")
        self.tree.tag_configure("fee_partial", foreground="#ca8a04")
        self.tree.tag_configure("fee_overdue", foreground="#dc2626")
        self.tree.tag_configure("bal_zero", foreground="#16a34a")
        self.tree.tag_configure("bal_due", foreground="#dc2626")
        self.tree.tag_configure("status_active", foreground="#16a34a")
        self.tree.tag_configure("status_archived", foreground="#64748b")

        scroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scroll.set)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.pack(fill=tk.BOTH, expand=True)

        footer = tk.Frame(body, bg=theme.WHITE)
        footer.pack(fill=tk.X, pady=(8, 0))
        self.lbl_student_count = tk.Label(
            footer, text="", font=("Segoe UI", 9),
            bg=theme.WHITE, fg="#64748b",
        )
        self.lbl_student_count.pack(side=tk.LEFT)

        self.load_table()

    def _refresh_class_filter_options(self):
        if not self.cmb_class_filter:
            return
        current = self.cmb_class_filter.get() or "All Classes"
        rows = db.run(
            "SELECT DISTINCT class_sec FROM students "
            "WHERE class_sec IS NOT NULL AND TRIM(class_sec) <> '' "
            "ORDER BY class_sec",
            fetchall=True,
        ) or []
        classes = ["All Classes"] + [r[0] for r in rows if r[0]]
        self.cmb_class_filter["values"] = classes
        self.cmb_class_filter.set(current if current in classes else "All Classes")

    def _update_stat_cards(self, all_rows, total_outstanding, current_month_info):
        """Refresh the four reference-style metric cards from live data."""
        total = len(all_rows)
        active = 0
        pending_fees = 0
        classes = set()
        for s_id, name, fname, phone, cls, total_f, paid_f, status in all_rows:
            status = status or "Active"
            if status == "Active":
                active += 1
            if cls:
                classes.add(str(cls).strip())
            bal = total_outstanding.get(s_id)
            if bal is None:
                bal = float(total_f or 0) - float(paid_f or 0)
            cm = current_month_info.get(s_id)
            month_unpaid = False
            if cm and cm.get("status") in ("PENDING", "PARTIAL", "OVERDUE", "NO_CYCLE"):
                month_unpaid = float(cm.get("balance") or 0) > 0 or cm.get("status") != "PAID"
            if ((bal and bal > 0) or month_unpaid) and status == "Active":
                pending_fees += 1

        if self._stat_total is not None:
            self._stat_total.config(text=str(total))
        if self._stat_active is not None:
            self._stat_active.config(text=str(active))
        if self._stat_pending is not None:
            self._stat_pending.config(text=str(pending_fees))
        if self._stat_classes is not None:
            self._stat_classes.config(text=str(len(classes)))

    def load_table(self, ev=None):
        if not self.tree:
            return
        search = self.ent_search.get().strip() if self.ent_search else ""
        self.tree.delete(*self.tree.get_children())
        self._refresh_class_filter_options()

        fee_filter = self.cmb_fee_filter.get() if self.cmb_fee_filter else "All Fees"
        class_filter = self.cmb_class_filter.get() if self.cmb_class_filter else "All Classes"
        status_filter = self.cmb_status_filter.get() if self.cmb_status_filter else "Active Only"

        if self.show_archived_var is not None:
            self.show_archived_var.set(status_filter in ("Archived Only", "All Status"))

        rows = db.run(
            "SELECT student_id, name, father_name, phone, class_sec, "
            "COALESCE(total_fee, 0), COALESCE(paid_fee, 0), COALESCE(status, 'Active') "
            "FROM students",
            fetchall=True,
        ) or []

        total_outstanding, current_month_info = _fee_maps_for_directory()
        can_fee = rbac.can(self.user_role, "student.fee.view")

        try:
            self._update_stat_cards(rows, total_outstanding, current_month_info)
        except Exception:
            pass

        shown = 0

        for s_id, name, fname, phone, cls, total_f, paid_f, status in rows:
            status = status or "Active"
            total_f = float(total_f or 0)
            paid_f = float(paid_f or 0)

            if status_filter == "Active Only" and status != "Active":
                continue
            if status_filter == "Archived Only" and status != "Archived":
                continue
            if class_filter and class_filter != "All Classes":
                if (cls or "") != class_filter:
                    continue

            if search:
                q = search.lower()
                hay = " ".join([
                    str(s_id or ""), str(name or ""), str(fname or ""),
                    str(phone or ""), str(cls or ""),
                ]).lower()
                if q not in hay:
                    continue

            if s_id in total_outstanding:
                bal = total_outstanding[s_id]
            else:
                bal = total_f - paid_f

            cm = current_month_info.get(s_id)
            if cm:
                month_status_code = cm["status"]
                month_bal = float(cm["balance"])
                month_status_label = _status_display(month_status_code)
                if month_status_code == "PAID":
                    display_month = "Paid"
                elif month_bal > 0:
                    display_month = month_status_label
                else:
                    display_month = month_status_label
            else:
                default_balance = max(0.0, total_f - paid_f)
                if default_balance > 0:
                    month_status_code = "PENDING"
                    month_bal = default_balance
                    display_month = "Pending"
                else:
                    month_status_code = "PAID"
                    month_bal = 0.0
                    display_month = "Paid"

            if fee_filter == "This Month Paid" and month_status_code != "PAID":
                continue
            if fee_filter == "This Month Pending" and month_status_code not in ("PENDING", "NO_CYCLE"):
                continue
            if fee_filter == "This Month Partial" and month_status_code != "PARTIAL":
                continue
            if fee_filter == "This Month Overdue" and month_status_code != "OVERDUE":
                continue
            if fee_filter == "Has Total Balance" and bal <= 0:
                continue
            if fee_filter == "No Total Balance" and bal > 0:
                continue

            if bal <= 0:
                bal_display = "Rs. 0"
            else:
                bal_display = f"Rs. {bal:,.0f}"

            tags = []
            code = (month_status_code or "").upper()
            if code == "PAID":
                tags.append("fee_paid")
            elif code == "PENDING":
                tags.append("fee_pending")
            elif code == "PARTIAL":
                tags.append("fee_partial")
            elif code == "OVERDUE":
                tags.append("fee_overdue")
            if bal <= 0:
                tags.append("bal_zero")
            else:
                tags.append("bal_due")
            if status == "Active":
                tags.append("status_active")
            else:
                tags.append("status_archived")

            if can_fee:
                self.tree.insert(
                    "", tk.END,
                    values=(
                        s_id, name, fname or "—", cls or "", phone or "",
                        display_month, bal_display, status,
                    ),
                    tags=tuple(tags),
                )
            else:
                self.tree.insert(
                    "", tk.END,
                    values=(s_id, name, fname or "—", cls or "", phone or "", status),
                    tags=tuple(tags),
                )
            shown += 1

        if self.lbl_student_count:
            if shown == 0:
                self.lbl_student_count.config(text="Showing 0 students")
            else:
                self.lbl_student_count.config(
                    text=f"Showing 1 to {shown} of {shown} students"
                )

    def _clear_filters(self):
        if self.ent_search:
            self.ent_search.delete(0, tk.END)
        if self.cmb_fee_filter:
            self.cmb_fee_filter.set("All Fees")
        if self.cmb_class_filter:
            self.cmb_class_filter.set("All Classes")
        if self.cmb_status_filter:
            self.cmb_status_filter.set("Active Only")
        if self.show_archived_var is not None:
            self.show_archived_var.set(False)
        self.load_table()

    def get_selected_student_id(self):
        selected = self.tree.focus() if self.tree else None
        values = self.tree.item(selected, "values") if selected else ()
        if not values:
            return None, None
        return values[0], values[1] if len(values) > 1 else None

    # ----- actions -----
    def _open_admission(self):
        if not rbac.can(self.user_role, "student.add"):
            messagebox.showerror(
                "Permission Denied", "You are not allowed to add students.", parent=self.root
            )
            return
        cb = self.callbacks.get("open_admission")
        if cb:
            cb()
            return
        try:
            from student_admission import launch_admission_window
            launch_admission_window(self.root, self.user_role, self.current_user)
        except Exception as e:
            messagebox.showerror("Error", f"Could not open admission:\n{e}", parent=self.root)

    def _open_profile(self):
        if not rbac.can(self.user_role, "student.view"):
            messagebox.showerror(
                "Permission Denied", "You are not allowed to view student profiles.", parent=self.root
            )
            return
        s_id, _ = self.get_selected_student_id()
        if not s_id:
            messagebox.showinfo(
                "Select Student", "Please select a student from the directory first.", parent=self.root
            )
            return
        cb = self.callbacks.get("open_student_profile")
        if cb:
            cb(s_id)
            return
        try:
            from student_profile import launch_student_profile_window
            win = launch_student_profile_window(self.root, self.user_role, self.current_user)
            if win and hasattr(win, "ent_search"):
                win.ent_search.insert(0, s_id)
                win.search_student()
        except Exception as e:
            messagebox.showerror("Error", f"Could not open profile:\n{e}", parent=self.root)

    def _reprint_id_card(self):
        if not rbac.can(self.user_role, "student.view"):
            messagebox.showerror(
                "Permission Denied", "You are not allowed to view student profiles.", parent=self.root
            )
            return
        s_id, _ = self.get_selected_student_id()
        if not s_id:
            messagebox.showinfo(
                "Select Student", "Please select a student from the directory first.", parent=self.root
            )
            return
        row = db.run(
            "SELECT name, father_name, class_sec, phone, photo_path FROM students WHERE student_id=?",
            (s_id,), fetchone=True,
        )
        if not row:
            messagebox.showerror("Not Found", f"Student '{s_id}' not found.", parent=self.root)
            return
        name, father_name, cls, phone, photo_path = row
        emer = ""
        try:
            er = db.run(
                "SELECT emergency_contact_phone FROM student_admission_extra WHERE student_id=?",
                (s_id,), fetchone=True,
            )
            if er and er[0]:
                emer = er[0]
        except Exception:
            pass
        out_path = os.path.join(os.getcwd(), f"ID_Card_{s_id}.pdf")
        try:
            reports.generate_id_card(
                s_id, name, cls, out_path,
                father_name=father_name or "", phone=phone or "",
                photo_path=photo_path, emergency_phone=emer,
            )
        except Exception as e:
            messagebox.showerror("ID Card Error", f"Could not generate ID card:\n{e}", parent=self.root)
            return
        try:
            if os.name == "nt":
                os.startfile(out_path)
            elif shutil.which("xdg-open"):
                os.system(f'xdg-open "{out_path}"')
            elif shutil.which("open"):
                os.system(f'open "{out_path}"')
        except Exception:
            pass
        _log_activity(self.current_user, f"Regenerated ID card for student {s_id}")
        messagebox.showinfo("ID Card Ready", f"ID Card generated:\n{out_path}", parent=self.root)

    def _remove_selected(self):
        """Deactivate (archive) a student — does NOT permanently delete data.

        Sets students.status = 'Archived' so the student disappears from the
        default Active-Only list but all attendance, fees, marks stay intact.
        If the selected student is already Archived, offers Reactivate.
        """
        if not (
            rbac.can(self.user_role, "student.delete")
            or rbac.can(self.user_role, "student.edit")
        ):
            messagebox.showerror(
                "Permission Denied",
                "You are not allowed to deactivate students.",
                parent=self.root,
            )
            return
        selected = self.tree.selection() if self.tree else ()
        if not selected:
            messagebox.showwarning(
                "Select Student",
                "Pehle Students Directory se student select karein.",
                parent=self.root,
            )
            return
        values = self.tree.item(selected[0], "values")
        if not values:
            messagebox.showwarning(
                "Select Student",
                "Selected student ki information nahi mili.",
                parent=self.root,
            )
            return

        student_id = str(values[0]).strip()
        student_name = str(values[1]).strip() if len(values) > 1 else student_id

        row = db.run(
            "SELECT student_id, name, COALESCE(status, 'Active') FROM students WHERE student_id=?",
            (student_id,),
            fetchone=True,
        )
        if not row:
            messagebox.showerror(
                "Student Not Found",
                f"Student ID '{student_id}' database mein nahi mila.",
                parent=self.root,
            )
            self.load_table()
            return

        current_status = (row[2] or "Active").strip()
        is_archived = current_status.lower() in ("archived", "inactive", "deactivated")

        if is_archived:
            # Reactivate path
            if not messagebox.askyesno(
                "Reactivate Student",
                f"Is student ko wapas Active karna hai?\n\n"
                f"Student: {row[1]}\nStudent ID: {row[0]}\n"
                f"Current status: {current_status}",
                parent=self.root,
            ):
                return
            try:
                db.run(
                    "UPDATE students SET status=? WHERE student_id=?",
                    ("Active", student_id),
                    commit=True,
                )
                _log_activity(
                    self.current_user,
                    f"Reactivated student {student_id} ({row[1]})",
                )
                self.load_table()
                cb = self.callbacks.get("on_students_changed")
                if cb:
                    try:
                        cb()
                    except Exception:
                        pass
                messagebox.showinfo(
                    "Student Reactivated",
                    f"Student '{student_name}' ({student_id}) ab Active hai.",
                    parent=self.root,
                )
            except Exception as exc:
                messagebox.showerror(
                    "Reactivate Failed",
                    f"Student reactivate nahi ho saka.\n\nError:\n{exc}",
                    parent=self.root,
                )
            return

        # Deactivate path
        if not messagebox.askyesno(
            "Confirm Deactivate Student",
            f"Student ko Deactivate (Archive) karna hai?\n\n"
            f"Student: {row[1]}\nStudent ID: {row[0]}\n\n"
            "• Student Active list se hat jayega\n"
            "• Attendance, fees, marks ka data safe rahega\n"
            "• Baad mein Status filter se Reactivate kiya ja sakta hai",
            icon="warning",
            parent=self.root,
        ):
            return

        try:
            db.run(
                "UPDATE students SET status=? WHERE student_id=?",
                ("Archived", student_id),
                commit=True,
            )
            _log_activity(
                self.current_user,
                f"Deactivated (archived) student {student_id} ({row[1]})",
            )
            self.load_table()
            cb = self.callbacks.get("on_students_changed")
            if cb:
                try:
                    cb()
                except Exception:
                    pass
            messagebox.showinfo(
                "Student Deactivated",
                f"Student '{student_name}' ({student_id}) deactivate ho gaya.\n"
                "Data delete nahi hua — Status filter se 'Archived Only' mein dikhega.",
                parent=self.root,
            )
        except Exception as exc:
            messagebox.showerror(
                "Deactivate Failed",
                f"Student deactivate nahi ho saka.\n\nError:\n{exc}",
                parent=self.root,
            )

    def _edit_selected(self):
        if not rbac.can(self.user_role, "student.edit"):
            messagebox.showerror("Permission Denied", "You are not allowed to edit students.", parent=self.root)
            return
        s_id, _ = self.get_selected_student_id()
        if not s_id:
            messagebox.showinfo("Select Student", "Please select a student from the directory first.", parent=self.root)
            return

        row = db.run(
            """SELECT student_id, name, father_name, dob, phone, address, class_sec,
                      photo_path, prev_education, total_fee, paid_fee, status
               FROM students WHERE student_id=?""",
            (s_id,), fetchone=True,
        )
        if not row:
            messagebox.showerror("Error", "Student not found.", parent=self.root)
            return

        extra = db.run(
            """SELECT gender, blood_group, nationality, religion, mother_name, guardian_name,
                      guardian_cnic, occupation, alt_phone, email, current_address, city, area,
                      admission_date, academic_year, admission_type, emergency_contact_name,
                      emergency_contact_phone, emergency_relationship, emergency_notes
               FROM student_admission_extra WHERE student_id=?""",
            (s_id,), fetchone=True,
        )

        base_keys = [
            "student_id", "name", "father_name", "dob", "phone", "address",
            "class_sec", "photo_path", "prev_education", "total_fee", "paid_fee", "status",
        ]
        base = dict(zip(base_keys, row))
        extra_keys = [
            "gender", "blood_group", "nationality", "religion", "mother_name",
            "guardian_name", "guardian_cnic", "occupation", "alt_phone", "email",
            "current_address", "city", "area", "admission_date", "academic_year",
            "admission_type", "emergency_contact_name", "emergency_contact_phone",
            "emergency_relationship", "emergency_notes",
        ]
        extra_data = dict(zip(extra_keys, extra)) if extra else {k: "" for k in extra_keys}
        can_fee = rbac.can(self.user_role, "student.fee.edit")
        previous_paid = float(base.get("paid_fee") or 0)

        win = tk.Toplevel(self.root)
        win.title(f"Edit Complete Student Profile — {s_id}")
        win.geometry("900x760")
        win.minsize(760, 620)
        win.config(bg=theme.SILVER)
        win.transient(self.root)
        win.grab_set()

        header = tk.Frame(win, bg=theme.NAVY, padx=18, pady=12)
        header.pack(fill=tk.X)
        tk.Label(header, text=f"✏️ EDIT STUDENT — {s_id}", font=theme.FONT_H1, bg=theme.NAVY, fg="white").pack(side=tk.LEFT)
        tk.Label(header, text="Complete Student Information", font=theme.FONT_SMALL, bg=theme.NAVY, fg=theme.BRAND_BLUE_LIGHT).pack(side=tk.RIGHT)

        outer = tk.Frame(win, bg=theme.SILVER)
        outer.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)
        canvas = tk.Canvas(outer, bg=theme.WHITE, highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer, orient=tk.VERTICAL, command=canvas.yview)
        form = tk.Frame(canvas, bg=theme.WHITE, padx=18, pady=14)
        form.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=form, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        fields = {}

        def add_section(title):
            tk.Label(form, text=title, font=theme.FONT_H2, bg=theme.WHITE, fg=theme.NAVY).pack(
                fill=tk.X, anchor="w", pady=(8, 5)
            )

        def add_field(key, label, value="", state="normal"):
            row_frame = tk.Frame(form, bg=theme.WHITE)
            row_frame.pack(fill=tk.X, pady=4)
            tk.Label(
                row_frame, text=label, width=24, anchor="w",
                font=theme.FONT_SMALL, bg=theme.WHITE, fg=theme.TEXT_MUTED,
            ).pack(side=tk.LEFT)
            ent = tk.Entry(row_frame, font=theme.FONT_BODY)
            ent.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=3)
            ent.insert(0, "" if value is None else str(value))
            if state != "normal":
                ent.config(state=state)
            fields[key] = ent
            return ent

        add_section("Basic Student Information")
        add_field("student_id", "Student ID", base["student_id"], "disabled")
        add_field("name", "Full Name *", base["name"])
        add_field("father_name", "Father's Name", base["father_name"])
        add_field("dob", "Date of Birth (YYYY-MM-DD)", base["dob"])
        add_field("phone", "Phone", base["phone"])
        add_field("address", "Permanent Address", base["address"])
        add_field("class_sec", "Class / Section *", base["class_sec"])
        add_field("prev_education", "Previous Education", base["prev_education"])

        photo_row = tk.Frame(form, bg=theme.WHITE)
        photo_row.pack(fill=tk.X, pady=4)
        tk.Label(photo_row, text="Photo", width=24, anchor="w", font=theme.FONT_SMALL, bg=theme.WHITE, fg=theme.TEXT_MUTED).pack(side=tk.LEFT)
        photo_var = tk.StringVar(value=base.get("photo_path") or "")
        tk.Entry(photo_row, textvariable=photo_var, font=theme.FONT_BODY).pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=3)

        photo_preview_box = tk.Frame(form, bg="#cbd5e1", width=90, height=100,
                                     highlightbackground=theme.SILVER_BORDER, highlightthickness=1)
        photo_preview_box.pack(anchor="w", padx=(160, 0), pady=(0, 6))
        photo_preview_box.pack_propagate(False)
        lbl_photo_preview = tk.Label(photo_preview_box, text="No\nPhoto", bg="#cbd5e1",
                                     fg=theme.TEXT_MUTED, font=theme.FONT_SMALL)
        lbl_photo_preview.pack(expand=True)

        def _refresh_photo_preview(path=None):
            try:
                from student_photos_util import apply_photo_to_label
                apply_photo_to_label(
                    lbl_photo_preview,
                    path if path is not None else photo_var.get(),
                    size=(90, 100),
                    student_id=s_id,
                    placeholder_text="No\nPhoto",
                )
            except Exception:
                try:
                    lbl_photo_preview.configure(image="", text="No\nPhoto")
                    lbl_photo_preview.image = None
                except Exception:
                    pass

        def choose_photo():
            path = filedialog.askopenfilename(
                filetypes=[("Image Files", "*.jpg *.png *.jpeg *.gif *.bmp")],
                parent=win,
            )
            if path:
                photo_var.set(path)
                _refresh_photo_preview(path)

        tk.Button(photo_row, text="Browse", command=choose_photo, bg=theme.SLATE, fg="white",
                  bd=0, padx=10, cursor="hand2").pack(side=tk.LEFT, padx=(8, 0))
        _refresh_photo_preview()

        add_section("Fee Information")
        add_field("total_fee", "Total Fee", base["total_fee"], "normal" if can_fee else "disabled")
        add_field("paid_fee", "Paid Fee", base["paid_fee"], "normal" if can_fee else "disabled")

        add_section("Admission / Personal Information")
        for key, label in [
            ("gender", "Gender"), ("blood_group", "Blood Group"),
            ("nationality", "Nationality"), ("religion", "Religion"),
            ("mother_name", "Mother's Name"), ("guardian_name", "Guardian Name"),
            ("guardian_cnic", "Guardian CNIC"), ("occupation", "Guardian Occupation"),
            ("alt_phone", "Alternate Phone"), ("email", "Email"),
            ("current_address", "Current Address"), ("city", "City"), ("area", "Area"),
            ("admission_date", "Admission Date"), ("academic_year", "Academic Year"),
            ("admission_type", "Admission Type"),
            ("emergency_contact_name", "Emergency Contact Name"),
            ("emergency_contact_phone", "Emergency Contact Phone"),
            ("emergency_relationship", "Emergency Relationship"),
            ("emergency_notes", "Emergency Notes"),
        ]:
            add_field(key, label, extra_data.get(key, ""))

        add_field("status", "Student Status", base.get("status") or "Active")

        def do_save():
            name = fields["name"].get().strip()
            cls = fields["class_sec"].get().strip()
            if not name or not cls:
                messagebox.showerror("Required Fields", "Student Name and Class / Section are required.", parent=win)
                return
            if can_fee:
                total_f, ok1 = _safe_float(fields["total_fee"].get(), "Total Fee", default=0.0, parent=win)
                paid_f, ok2 = _safe_float(fields["paid_fee"].get(), "Paid Fee", default=0.0, parent=win)
                if not (ok1 and ok2):
                    return
                if paid_f > total_f:
                    if not messagebox.askyesno(
                        "Overpayment Warning",
                        f"Paid Fee (Rs. {paid_f:.2f}) is more than Total Fee (Rs. {total_f:.2f}). Continue?",
                        parent=win,
                    ):
                        return
            else:
                total_f = base.get("total_fee") or 0.0
                paid_f = base.get("paid_fee") or 0.0

            # Persist photo under student_photos/{student_id}.ext when a new
            # local file is chosen; keep existing DB path otherwise.
            raw_photo = photo_var.get().strip()
            stored_photo = raw_photo
            if raw_photo and os.path.isfile(raw_photo):
                try:
                    from student_photos_util import save_student_photo, photos_dir
                    # Only re-copy when source is outside our photos folder
                    # (user picked a new file via Browse).
                    if not os.path.abspath(raw_photo).startswith(os.path.abspath(photos_dir())):
                        saved = save_student_photo(raw_photo, s_id)
                        if saved:
                            stored_photo = saved
                except Exception:
                    stored_photo = raw_photo

            db.run(
                """UPDATE students SET
                   name=?, father_name=?, dob=?, phone=?, address=?, class_sec=?,
                   photo_path=?, prev_education=?, total_fee=?, paid_fee=?, status=?
                   WHERE student_id=?""",
                (
                    name, fields["father_name"].get().strip(), fields["dob"].get().strip(),
                    fields["phone"].get().strip(), fields["address"].get().strip(), cls,
                    stored_photo, fields["prev_education"].get().strip(),
                    total_f, paid_f, fields["status"].get().strip() or "Active", s_id,
                ),
                commit=True,
            )

            extra_values = tuple(fields[key].get().strip() for key in extra_keys)
            extra_exists = db.run(
                "SELECT 1 FROM student_admission_extra WHERE student_id=?", (s_id,), fetchone=True
            )
            if extra_exists:
                db.run(
                    """UPDATE student_admission_extra SET
                       gender=?, blood_group=?, nationality=?, religion=?, mother_name=?,
                       guardian_name=?, guardian_cnic=?, occupation=?, alt_phone=?, email=?,
                       current_address=?, city=?, area=?, admission_date=?, academic_year=?,
                       admission_type=?, emergency_contact_name=?, emergency_contact_phone=?,
                       emergency_relationship=?, emergency_notes=?
                       WHERE student_id=?""",
                    extra_values + (s_id,), commit=True,
                )
            else:
                db.run(
                    """INSERT INTO student_admission_extra (
                       student_id, gender, blood_group, nationality, religion, mother_name,
                       guardian_name, guardian_cnic, occupation, alt_phone, email,
                       current_address, city, area, admission_date, academic_year,
                       admission_type, emergency_contact_name, emergency_contact_phone,
                       emergency_relationship, emergency_notes
                       ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (s_id,) + extra_values, commit=True,
                )

            delta = paid_f - previous_paid
            if delta > 0 and can_fee:
                try:
                    accounting.record_fee_revenue(self.user_role, s_id, delta, self.current_user)
                except Exception:
                    pass

            _log_activity(self.current_user, f"Updated complete student profile for {s_id}")
            self.load_table()
            cb = self.callbacks.get("on_students_changed")
            if cb:
                try:
                    cb()
                except Exception:
                    pass
            messagebox.showinfo(
                "Updated",
                f"Complete student information for {s_id} has been updated successfully.",
                parent=win,
            )
            win.destroy()

        button_row = tk.Frame(win, bg=theme.SILVER, padx=12, pady=10)
        button_row.pack(fill=tk.X)
        theme.primary_button(button_row, "💾 Save All Changes", do_save, bg=theme.SUCCESS).pack(side=tk.LEFT)
        tk.Button(
            button_row, text="Cancel", command=win.destroy,
            bg=theme.WHITE, fg=theme.TEXT_MUTED, bd=0, font=theme.FONT_SMALL, cursor="hand2",
        ).pack(side=tk.LEFT, padx=12)

        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))

        def close_window():
            try:
                canvas.unbind_all("<MouseWheel>")
            except Exception:
                pass
            win.destroy()

        win.protocol("WM_DELETE_WINDOW", close_window)

    def _export_excel(self):
        if not rbac.can(self.user_role, "student.view"):
            messagebox.showerror(
                "Permission Denied", "You are not allowed to export the student list.", parent=self.root
            )
            return
        items = self.tree.get_children() if self.tree else []
        if not items:
            messagebox.showinfo(
                "Nothing to Export",
                "Current filters show no students.\nChange filters or clear them, then try again.",
                parent=self.root,
            )
            return

        can_fee = rbac.can(self.user_role, "student.fee.view")
        if can_fee:
            headers = [
                "Student ID", "Name", "Father Name", "Class", "Phone",
                "This Month Fee Status", "Total Balance", "Status",
            ]
        else:
            headers = ["Student ID", "Name", "Father Name", "Class", "Phone", "Status"]

        fee_f = self.cmb_fee_filter.get() if self.cmb_fee_filter else "All Fees"
        class_f = self.cmb_class_filter.get() if self.cmb_class_filter else "All Classes"
        status_f = self.cmb_status_filter.get() if self.cmb_status_filter else "Active Only"
        search_q = self.ent_search.get().strip() if self.ent_search else ""

        default_name = f"Student_List_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Save Student List as Excel",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Workbook", "*.xlsx"), ("All Files", "*.*")],
            parent=self.root,
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror(
                "Missing Library",
                "openpyxl is required for Excel export.\nInstall with: pip install openpyxl",
                parent=self.root,
            )
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Students"

            title = "Student Directory Export — AR School Management System"
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            cell = ws.cell(row=1, column=1, value=title)
            cell.font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F172A")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[1].height = 24

            y, m = _current_year_month()
            filter_bits = [
                f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                f"Fee month: {m:02d}/{y}",
                f"Fee filter: {fee_f}",
                f"Class: {class_f}",
                f"Status: {status_f}",
            ]
            if search_q:
                filter_bits.append(f"Search: {search_q}")
            filter_bits.append(f"Rows: {len(items)}")
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
            ws.cell(row=2, column=1, value="  |  ".join(filter_bits)).font = Font(
                name="Segoe UI", size=9, italic=True, color="64748B"
            )

            header_fill = PatternFill("solid", fgColor="0284C7")
            header_font = Font(name="Segoe UI", bold=True, color="FFFFFF")
            thin = Border(
                left=Side(style="thin", color="E2E8F0"),
                right=Side(style="thin", color="E2E8F0"),
                top=Side(style="thin", color="E2E8F0"),
                bottom=Side(style="thin", color="E2E8F0"),
            )
            for col_idx, h in enumerate(headers, start=1):
                c = ws.cell(row=4, column=col_idx, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin

            for row_idx, item_id in enumerate(items, start=5):
                values = self.tree.item(item_id, "values")
                for col_idx, val in enumerate(values, start=1):
                    c = ws.cell(row=row_idx, column=col_idx, value=val)
                    c.font = Font(name="Segoe UI", size=10)
                    c.border = thin
                    c.alignment = Alignment(horizontal="center" if col_idx != 2 else "left")

            widths = {1: 14, 2: 22, 3: 18, 4: 12, 5: 14, 6: 16, 7: 14, 8: 12}
            for col_idx in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_idx, 14)

            wb.save(path)
        except Exception as e:
            messagebox.showerror("Export Failed", f"Could not write Excel file:\n{e}", parent=self.root)
            return

        _log_activity(
            self.current_user,
            f"Exported student directory Excel ({len(items)} rows) "
            f"filters=[fee={fee_f}, class={class_f}, status={status_f}, search={search_q or '-'}]",
        )
        messagebox.showinfo(
            "Export Complete",
            f"{len(items)} student(s) exported.\n\nSaved to:\n{path}",
            parent=self.root,
        )