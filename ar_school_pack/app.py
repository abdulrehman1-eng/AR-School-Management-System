import os
import shutil
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import fee_export_bridge
import db
import rbac
import branding
import accounting
import results_engine
import reports
import theme
import ai_assistant
from fee_management_window import launch_fee_management_window
import fee_automation
from student_admission import launch_admission_window
from student_profile import launch_student_profile_window
from results_window import build_results_into, launch_results_window
from smart_attendance import launch_attendance_window
try:
    from teacher_payroll import launch_teacher_payroll_window
except ImportError:
    # Supports the current filename "teacher_payroll(2).py" as well as the
    # cleaner production filename "teacher_payroll.py".
    import importlib.util
    _teacher_payroll_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "teacher_payroll(2).py")
    _teacher_payroll_spec = importlib.util.spec_from_file_location("teacher_payroll", _teacher_payroll_path)
    if _teacher_payroll_spec is None or _teacher_payroll_spec.loader is None:
        raise ImportError(f"Could not load teacher payroll module: {_teacher_payroll_path}")
    _teacher_payroll_module = importlib.util.module_from_spec(_teacher_payroll_spec)
    _teacher_payroll_spec.loader.exec_module(_teacher_payroll_module)
    launch_teacher_payroll_window = _teacher_payroll_module.launch_teacher_payroll_window
from security import hash_password, verify_password
from settings_window import build_settings_tab as build_settings_panel

try:
    import matplotlib.pyplot as plt
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False


def log_activity(username, action):
    try:
        db.run(
            "INSERT INTO audit_logs (username, action, timestamp) VALUES (?, ?, ?)",
            (username, action, datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            commit=True,
        )
    except Exception as e:
        print(f"Audit Log Error: {e}")


def generate_next_student_id():
    row = db.run("SELECT student_id FROM students ORDER BY ROWID DESC LIMIT 1", fetchone=True)
    year = datetime.now().year
    if row and row[0].startswith(f"STU-{year}-"):
        try:
            new_num = int(row[0].split("-")[-1]) + 1
        except ValueError:
            new_num = 1
    else:
        new_num = 1
    return f"STU-{year}-{new_num:03d}"


def generate_next_teacher_id():
    row = db.run("SELECT teacher_id FROM teachers ORDER BY ROWID DESC LIMIT 1", fetchone=True)
    if row and row[0].startswith("TCH-"):
        try:
            new_num = int(row[0].split("-")[-1]) + 1
        except ValueError:
            new_num = 1
    else:
        new_num = 1
    return f"TCH-{new_num:03d}"


def safe_float(raw_text, field_label, default=None):
    """Parse a numeric form field with a friendly error instead of a raw
    traceback. Returns (value, ok). On failure shows a messagebox and
    returns (None, False) so the caller can abort the save.
    `default` is used when the field is blank (None means blank is an error)."""
    text = (raw_text or "").strip()
    if not text:
        if default is not None:
            return default, True
        messagebox.showerror("Invalid Input", f"{field_label} is required.")
        return None, False
    try:
        return float(text), True
    except ValueError:
        messagebox.showerror("Invalid Input", f"{field_label} must be a valid number (e.g. 1500 or 1500.50).")
        return None, False


# ==========================================
# LOGIN WINDOW
# ==========================================
class LoginWindow:
    def __init__(self, root):
        self.root = root
        self.root.title("AR School Management System — Login")
        self.root.geometry("460x520")
        self.root.config(bg=theme.SILVER)
        self.root.resizable(False, False)

        b = branding.get_branding()

        outer = tk.Frame(self.root, bg=theme.SILVER)
        outer.pack(fill=tk.BOTH, expand=True)

        card = tk.Frame(outer, bg=theme.NAVY, padx=30, pady=30)
        card.place(relx=0.5, rely=0.5, anchor="center", width=380, height=460)

        tk.Label(card, text="AR SOFTWARE SOLUTIONS", font=("Segoe UI", 10, "bold"),
                 bg=theme.NAVY, fg=theme.BRAND_BLUE_LIGHT).pack(pady=(4, 0))
        tk.Label(card, text="AR School Management System", font=("Segoe UI", 16, "bold"),
                 bg=theme.NAVY, fg="white", wraplength=320, justify="center").pack(pady=(2, 2))
        tk.Label(card, text="Smart Software. Simple Solutions.", font=("Segoe UI", 8, "italic"),
                 bg=theme.NAVY, fg="#94a3b8").pack(pady=(0, 6))
        tk.Label(card, text=(b["org_name"] or "").upper(), font=("Segoe UI", 9, "bold"),
                 bg=theme.NAVY, fg="#38bdf8", wraplength=320, justify="center").pack(pady=(0, 18))

        tk.Label(card, text="Username", font=("Segoe UI", 9, "bold"), bg=theme.NAVY, fg="#cbd5e1").pack(anchor="w")
        self.ent_user = tk.Entry(card, font=("Segoe UI", 11), bg=theme.NAVY_LIGHT, fg="white",
                                  insertbackground="white", relief="flat")
        self.ent_user.pack(fill=tk.X, ipady=6, pady=(4, 14))

        tk.Label(card, text="Password", font=("Segoe UI", 9, "bold"), bg=theme.NAVY, fg="#cbd5e1").pack(anchor="w")
        self.ent_pass = tk.Entry(card, show="*", font=("Segoe UI", 11), bg=theme.NAVY_LIGHT, fg="white",
                                  insertbackground="white", relief="flat")
        self.ent_pass.pack(fill=tk.X, ipady=6, pady=(4, 22))
        self.ent_pass.bind("<Return>", lambda e: self.check_login())

        tk.Button(card, text="LOGIN", command=self.check_login, bg=theme.BRAND_BLUE, fg="white",
                  font=("Segoe UI", 11, "bold"), bd=0, pady=10, cursor="hand2").pack(fill=tk.X)

        tk.Label(card, text="Default: admin/admin123 · teacher/teacher123 · reception/reception123",
                 font=("Segoe UI", 7), bg=theme.NAVY, fg="#475569", wraplength=320, justify="center").pack(side=tk.BOTTOM, pady=(20, 0))

    def check_login(self):
        user = self.ent_user.get().strip()
        pwd = self.ent_pass.get().strip()

        row = db.run("SELECT password, role, is_hashed, COALESCE(is_active,1) FROM users WHERE username=?", (user,), fetchone=True)

        if row and not row[3]:
            log_activity(user, "Login attempt on deactivated account")
            messagebox.showerror("Account Deactivated", "This account has been deactivated. Contact your administrator.")
            return

        if row and verify_password(pwd, row[0]):
            role = row[1]
            # Transparent migration of legacy plain-text passwords to hashed form.
            if not row[2]:
                db.run("UPDATE users SET password=?, is_hashed=1 WHERE username=?",
                       (hash_password(pwd), user), commit=True)
            log_activity(user, f"User logged in successfully as {role}")
            self.root.destroy()
            main_root = tk.Tk()
            theme.apply_ttk_style()
            StudentManagementApp(main_root, role, user)
            main_root.mainloop()
        else:
            log_activity(user or "(unknown)", "Failed login attempt")
            messagebox.showerror("Login Error", "Invalid Username or Password!")


# ==========================================
# MAIN APPLICATION
# ==========================================
class StudentManagementApp:
    def __init__(self, root, user_role, current_user):
        self.root = root
        self.user_role = user_role
        self.current_user = current_user
        # Prefer centralized system_settings; fall back to branding table.
        org_name = db.get_setting("school_name", "") or ""
        if not org_name:
            try:
                b = branding.get_branding()
                org_name = b.get("org_name") or ""
            except Exception:
                org_name = ""
        org_name = org_name or "SCHOOL MANAGEMENT SYSTEM"
        self.root.title(f"{org_name} — AR School Management System [{self.user_role} Panel]")
        self.root.geometry("1500x760")
        self.root.minsize(1220, 660)
        self.root.config(bg=theme.SILVER)

        # ---------------- Top header ----------------
        header = tk.Frame(self.root, bg=theme.NAVY, pady=10, padx=20)
        header.pack(fill=tk.X)
        brand_box = tk.Frame(header, bg=theme.NAVY)
        brand_box.pack(side=tk.LEFT)
        tk.Label(brand_box, text=org_name.upper(), font=theme.FONT_BRAND, bg=theme.NAVY, fg="white").pack(anchor="w")
        tk.Label(brand_box, text="AR SCHOOL MANAGEMENT SYSTEM · Smart Software. Simple Solutions.",
                 font=theme.FONT_TAGLINE, bg=theme.NAVY, fg=theme.BRAND_BLUE_LIGHT).pack(anchor="w")

        user_box = tk.Frame(header, bg=theme.NAVY)
        user_box.pack(side=tk.RIGHT)
        tk.Button(user_box, text="⏻ Logout", command=self.logout, bg=theme.DANGER, fg="white",
                  font=theme.FONT_SMALL, bd=0, padx=10, pady=4, cursor="hand2").pack(side=tk.RIGHT, padx=(10, 0))
        if rbac.can(self.user_role, "backup.run"):
            tk.Button(user_box, text="💾 USB Backup", command=self.make_backup, bg=theme.SLATE, fg="white",
                      font=theme.FONT_SMALL, bd=0, padx=10, pady=4, cursor="hand2").pack(side=tk.RIGHT, padx=(10, 0))
            tk.Button(user_box, text="♻ Restore Backup", command=self.restore_backup, bg=theme.WARNING, fg="white",
                      font=theme.FONT_SMALL, bd=0, padx=10, pady=4, cursor="hand2").pack(side=tk.RIGHT, padx=(10, 0))
        tk.Label(user_box, text=f"👤 {self.current_user}  ({self.user_role})", font=theme.FONT_BODY_BOLD,
                 bg=theme.NAVY, fg="white").pack(side=tk.RIGHT, padx=10)

        # ---------------- Body: sidebar + content ----------------
        body = tk.Frame(self.root, bg=theme.SILVER)
        body.pack(fill=tk.BOTH, expand=True)

        self.sidebar = tk.Frame(body, bg=theme.NAVY, width=220)
        self.sidebar.pack(side=tk.LEFT, fill=tk.Y)
        self.sidebar.pack_propagate(False)

        self.content = tk.Frame(body, bg=theme.SILVER)
        self.content.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=14, pady=14)

        # Every page is a frame stacked in the same spot; show_page() raises one.
        self.pages = {}
        self.nav_buttons = {}

        def add_page(key, frame):
            frame.place(x=0, y=0, relwidth=1, relheight=1)
            self.pages[key] = frame

        # Frames created up front (same ttk.Frame widgets as before — only
        # their parent changed, from a ttk.Notebook to this content stack).
        self.tab_dashboard = tk.Frame(self.content, bg=theme.SILVER)
        self.tab_admission = ttk.Frame(self.content)
        self.tab_attendance = ttk.Frame(self.content)
        self.tab_results = ttk.Frame(self.content)
        self.tab_teachers = ttk.Frame(self.content)
        self.tab_timetable = ttk.Frame(self.content)
        self.tab_accounting = ttk.Frame(self.content)
        self.tab_settings = ttk.Frame(self.content)
        self.tab_logs = ttk.Frame(self.content)

        add_page("dashboard", self.tab_dashboard)
        add_page("students", self.tab_admission)
        add_page("attendance", self.tab_attendance)

        can_results = rbac.can(self.user_role, "results.view") or rbac.can(self.user_role, "results.marks.edit")
        can_teachers = rbac.can(self.user_role, "teacher.view") or rbac.can(self.user_role, "teacher.attendance.mark")
        can_accounting = rbac.can(self.user_role, "accounting.dashboard")
        can_settings = rbac.can(self.user_role, "settings.branding") or rbac.can(self.user_role, "settings.users")
        can_logs = rbac.can(self.user_role, "audit.view")
        # Stored on self — refresh_dashboard() runs as its own method call
        # (every time the Dashboard nav item is clicked), not just inline
        # here, so these need to outlive __init__'s local scope.
        self.can_results = can_results
        self.can_teachers = can_teachers
        self.can_accounting = can_accounting
        self.can_settings = can_settings
        self.can_logs = can_logs

        if can_results:
            add_page("results", self.tab_results)
        # Teacher & Payroll is now owned by teacher_payroll.py and opens
        # as the single authoritative popup window. The legacy inline
        # teachers page is intentionally not registered as a user-facing page.
        add_page("timetable", self.tab_timetable)
        if can_accounting:
            add_page("accounting", self.tab_accounting)
        if can_settings:
            add_page("settings", self.tab_settings)
        if can_logs:
            add_page("logs", self.tab_logs)

        # ---------------- Sidebar navigation ----------------
        nav_items = [("dashboard", "Dashboard", "🏠", True)]
        nav_items.append(("students", "Students", "🎓", True))
        nav_items.append(("results", "Results & Academics", "📊", can_results))
        nav_items.append(("teachers", "Teachers & Payroll", "👨‍🏫", can_teachers))
        nav_items.append(("timetable", "Timetable", "🕐", True))
        nav_items.append(("accounting", "Finance", "💰", can_accounting))
        nav_items.append(("settings", "Settings", "⚙️", can_settings))
        nav_items.append(("logs", "Audit Logs", "📜", can_logs))

        ai_btn = theme.sidebar_button(self.sidebar, "AI Admin Assistant", "🤖", command=self.open_ai_assistant)
        ai_btn.pack(fill=tk.X, side=tk.BOTTOM, pady=(0, 10))

        nav_frame = tk.Frame(self.sidebar, bg=theme.NAVY)
        nav_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))

        # One user-facing Attendance entry only. The legacy app.py Attendance
        # page is kept internally, but users are routed to the unified
        # smart_attendance.py screen so the feature never appears twice.
        attendance_btn = theme.sidebar_button(nav_frame, "Attendance", "🗓️", command=self.open_attendance)
        attendance_btn.pack(fill=tk.X)
        for key, label, icon, allowed in nav_items:
            if not allowed:
                continue
            if key == "teachers":
                btn = theme.sidebar_button(
                    nav_frame, label, icon, command=self.open_teacher_payroll
                )
            else:
                btn = theme.sidebar_button(
                    nav_frame, label, icon, command=lambda k=key: self.show_page(k)
                )
            btn.pack(fill=tk.X)
            self.nav_buttons[key] = btn

        # Build every allowed page's content exactly as before — the
        # existing build_*_tab methods and their DB calls are untouched.
        # The legacy Attendance page is still built for compatibility, but
        # it is intentionally not exposed in the sidebar/Quick Actions.
        self.build_dashboard_tab()
        self.build_admission_tab()
        self.build_attendance_tab()
        if can_results:
            self.build_results_tab()
        # Teacher & Payroll is provided by teacher_payroll.py. Do not build
        # the old inline page here, which would create a duplicate UI.
        self.build_timetable_tab()
        if can_accounting:
            self.build_accounting_tab()
        if can_settings:
            self.build_settings_tab()
        if can_logs:
            self.build_logs_tab()

        self.show_page("dashboard")

        # Fee automation (refresh overdue statuses + generate this
        # month's cycles for students who don't have one yet) — runs
        # once, shortly after the window is visible, and is wrapped so
        # it can NEVER crash or delay the main application. Everything
        # inside run_fee_automation() already catches its own exceptions
        # and returns a result dict; this is a second, outer safety net
        # in case fee_automation.py itself is missing/broken on an older
        # install.
        self.root.after(500, self._run_fee_automation_safely)
        # If the app stays open past month-end, re-run auto cycle generation
        # once the calendar month changes (poll every 15 minutes).
        self.root.after(2000, self._start_fee_automation_worker)
        # Continuous auto-absent worker: fires after school_closing_time even if
        # the app was opened earlier in the day and stays open past closing.
        self.root.after(1500, self._start_auto_absent_worker)
        self.root.protocol("WM_DELETE_WINDOW", self._on_app_close)

    def _run_fee_automation_safely(self, force=False):
        try:
            result = fee_automation.run_fee_automation(
                self.user_role, self.current_user, force=force,
            )
            if result.get("success"):
                print(
                    f"[Fee Automation] {result['month']:02d}/{result['year']}: "
                    f"{len(result['created'])} cycle(s) created, "
                    f"{len(result['skipped'])} already existed, "
                    f"{result['overdue_changed']} cycle(s) marked overdue "
                    f"(due={result.get('due_date') or '-'})."
                )
            elif result.get("reason"):
                print(f"[Fee Automation] Skipped: {result['reason']}")
            return result
        except Exception as exc:
            # Never let a fee-automation problem take down the ERP.
            print(f"[Fee Automation] Error (non-fatal, app continues normally): {exc}")
            return {"success": False, "reason": str(exc)}

    def _start_fee_automation_worker(self):
        """Background poll: if calendar month rolled over while the app is
        still open, generate the new month's cycles automatically."""
        def _tick():
            try:
                if fee_automation.needs_month_run() and rbac.can(
                    self.user_role, "fee.cycle.generate"
                ):
                    self._run_fee_automation_safely()
            except Exception as exc:
                print(f"[Fee Automation] Worker tick error (non-fatal): {exc}")
            try:
                self.root.after(15 * 60 * 1000, _tick)  # every 15 minutes
            except Exception:
                pass
        _tick()

    def _start_auto_absent_worker(self):
        """Start the background auto-absent engine.

        Design:
        - Closing time comes from system_settings.school_closing_time
        - Toggle system_settings.auto_absent_enabled must be on
        - Any Active student with NO attendance row for today
          (not Present, not Late, not Leave, not already Absent)
          is inserted as status=Absent, method='Auto System'
        - Worker polls every 60s so if the app stays open past closing,
          Absents are applied automatically without restart
        - Runs once per calendar day (attendance_auto_absent_log marker)
        - Does NOT require attendance.mark permission (system process)
        """
        try:
            from smart_attendance import start_auto_absent_worker, try_auto_absent_now
            # Immediate attempt if we are already past closing (e.g. late login)
            try_auto_absent_now(force=False)
            start_auto_absent_worker(self.root, poll_ms=60000, first_delay_ms=3000)
            print(
                f"[Auto-Absent] Worker started "
                f"(closing={db.get_setting('school_closing_time', '14:00')}, "
                f"enabled={db.get_setting('auto_absent_enabled', '1')})."
            )
        except Exception as exc:
            print(f"[Auto-Absent] Worker failed to start (non-fatal): {exc}")

    def _maybe_auto_backup_on_exit(self):
        """If auto_backup_on_exit is enabled, write a consistent backup."""
        try:
            if str(db.get_setting("auto_backup_on_exit", "1")) not in ("1", "true", "True", "yes"):
                return
            folder = (db.get_setting("backup_folder_path", "") or "").strip()
            if folder and os.path.isdir(folder):
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                dest = os.path.join(folder, f"AutoBackup_{timestamp}.db")
                db.close_conn()
                try:
                    shutil.copy(db.DB_PATH, dest)
                    print(f"[Auto-Backup] Saved to {dest}")
                finally:
                    db.get_conn()
            else:
                path = db.backup_database(reason="auto_exit")
                print(f"[Auto-Backup] Saved to {path}")
        except Exception as exc:
            print(f"[Auto-Backup] Error (non-fatal): {exc}")

    def _on_app_close(self):
        try:
            self._maybe_auto_backup_on_exit()
        finally:
            self.root.destroy()

    def open_admission(self):
        """Open the single professional New Student Admission window."""
        if not rbac.can(self.user_role, "student.add"):
            messagebox.showerror("Permission Denied", "You are not allowed to add students.")
            return
        return launch_admission_window(self.root, self.user_role, self.current_user)

    def open_student_profile(self):
        """Open the professional Student Profile for the selected directory row."""
        if not rbac.can(self.user_role, "student.view"):
            messagebox.showerror("Permission Denied", "You are not allowed to view student profiles.")
            return
        selected = self.tree_student.focus()
        values = self.tree_student.item(selected, "values") if selected else ()
        if not values:
            messagebox.showinfo("Select Student", "Please select a student from the directory first.")
            return
        student_id = values[0]
        win = launch_student_profile_window(self.root, self.user_role, self.current_user)
        if win and hasattr(win, "ent_search"):
            win.ent_search.insert(0, student_id)
            win.search_student()

    def reprint_selected_id_card(self):
        """Regenerate ID card for the student selected in the directory."""
        if not rbac.can(self.user_role, "student.view"):
            messagebox.showerror("Permission Denied", "You are not allowed to view student profiles.")
            return
        selected = self.tree_student.focus()
        values = self.tree_student.item(selected, "values") if selected else ()
        if not values:
            messagebox.showinfo("Select Student", "Please select a student from the directory first.")
            return
        s_id = values[0]
        row = db.run(
            "SELECT name, father_name, class_sec, phone, photo_path FROM students WHERE student_id=?",
            (s_id,), fetchone=True,
        )
        if not row:
            messagebox.showerror("Not Found", f"Student '{s_id}' not found.")
            return
        name, father_name, cls, phone, photo_path = row
        emer = ""
        try:
            er = db.run(
                "SELECT emergency_contact_phone FROM student_admission_extra WHERE student_id=?",
                (s_id,), fetchone=True,
            )
            if er and er[0]:
                emer = er[0]
        except Exception:
            pass
        out_path = os.path.join(os.getcwd(), f"ID_Card_{s_id}.pdf")
        try:
            reports.generate_id_card(
                s_id, name, cls, out_path, father_name=father_name or "", phone=phone or "",
                photo_path=photo_path, emergency_phone=emer,
            )
        except Exception as e:
            messagebox.showerror("ID Card Error", f"Could not generate ID card:\n{e}")
            return
        try:
            if os.name == "nt":
                os.startfile(out_path)
            elif shutil.which("xdg-open"):
                os.system(f'xdg-open "{out_path}"')
            elif shutil.which("open"):
                os.system(f'open "{out_path}"')
        except Exception:
            pass
        log_activity(self.current_user, f"Regenerated ID card for student {s_id}")
        messagebox.showinfo("ID Card Ready", f"ID Card generated:\n{out_path}")


    def open_fee_management(self):
        """Open the single, authoritative Fee Management window (merges
        the former Collect Fee / Fee Cycles screens into one).

        fee_management_window.py owns this workflow; this method only
        routes the user to it. student_fee_collection.py and
        fee_cycle_window.py still exist and still work if imported
        directly elsewhere, but the main app no longer opens either of
        them on its own — Fee Management is the one user-facing entry
        point now."""
        launch_fee_management_window(self.root, self.user_role, self.current_user)

    def open_whatsapp_fee_reminders(self):
        """Open the Remaining Fees Excel export window.

        Collects students with outstanding balances and writes a clean
        Excel file. No external WhatsApp / Node project is contacted —
        you take the file and use it manually in any other tool.
        """
        if not rbac.can(self.user_role, "fee.reports.view"):
            messagebox.showerror("Permission Denied", "You are not allowed to export fee reports.", parent=self.root)
            return

        win = tk.Toplevel(self.root)
        win.title("Export Remaining Fees")
        win.geometry("650x470")
        win.minsize(580, 400)
        win.config(bg=theme.SILVER)
        win.transient(self.root)

        header = tk.Frame(win, bg=theme.NAVY, padx=16, pady=12)
        header.pack(fill=tk.X)
        tk.Label(header, text="📁 EXPORT REMAINING FEES", font=theme.FONT_H1, bg=theme.NAVY, fg="white").pack(anchor="w")
        tk.Label(header, text="Active students with outstanding balance → Excel file", font=theme.FONT_SMALL, bg=theme.NAVY, fg="#94a3b8").pack(anchor="w")

        body = tk.Frame(win, bg=theme.WHITE, padx=16, pady=14)
        body.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)
        form = tk.Frame(body, bg=theme.WHITE)
        form.pack(fill=tk.X)
        tk.Label(form, text="Month:", bg=theme.WHITE, font=theme.FONT_BODY_BOLD).pack(side=tk.LEFT)
        ent_month = tk.Entry(form, width=5, font=theme.FONT_BODY)
        ent_month.insert(0, str(datetime.now().month))
        ent_month.pack(side=tk.LEFT, padx=(6, 16))
        tk.Label(form, text="Year:", bg=theme.WHITE, font=theme.FONT_BODY_BOLD).pack(side=tk.LEFT)
        ent_year = tk.Entry(form, width=7, font=theme.FONT_BODY)
        ent_year.insert(0, str(datetime.now().year))
        ent_year.pack(side=tk.LEFT, padx=(6, 16))

        status_lbl = tk.Label(body, text="Ready. Choose month/year then click Export to Excel.", bg=theme.WHITE, fg=theme.TEXT_MUTED, font=theme.FONT_SMALL, anchor="w")
        status_lbl.pack(fill=tk.X, pady=(12, 8))
        result_box = tk.Text(body, height=14, wrap="word", font=theme.FONT_SMALL, bg="#f8fafc", fg=theme.TEXT_DARK, relief="flat")
        result_box.pack(fill=tk.BOTH, expand=True)

        def values():
            try:
                month = int(ent_month.get().strip())
                year = int(ent_year.get().strip())
                if not 1 <= month <= 12 or not 2000 <= year <= 9999:
                    raise ValueError
                return month, year
            except ValueError:
                messagebox.showerror("Invalid Date", "Enter a valid month (1-12) and year.", parent=win)
                return None

        def export_excel():
            selected = values()
            if not selected:
                return
            month, year = selected
            try:
                result = fee_export_bridge.export_remaining_fees_to_excel(
                    self.user_role, month, year, actor=self.current_user
                )
            except Exception as exc:
                messagebox.showerror("Export Error", str(exc), parent=win)
                status_lbl.config(text="❌ Export failed — see error above.", fg=theme.DANGER)
                return

            result_box.delete("1.0", tk.END)
            if not result.get("success"):
                msg = result.get("message") or "No remaining fee data found."
                result_box.insert(tk.END, msg + "\n")
                status_lbl.config(text="ℹ️ No data to export.", fg=theme.TEXT_MUTED)
                return

            count = result.get("count", 0)
            total = result.get("total_amount", 0)
            path = result.get("path", "")
            result_box.insert(tk.END, f"✅ Excel file ready.\n\n")
            result_box.insert(tk.END, f"Students with remaining fee: {count}\n")
            result_box.insert(tk.END, f"Total remaining amount: Rs. {total:,.2f}\n\n")
            result_box.insert(tk.END, f"File saved at:\n{path}\n\n")
            result_box.insert(
                tk.END,
                "You can now open this Excel file and use the data manually "
                "in any other project (copy/paste or upload).\n"
                "This app no longer connects to any external WhatsApp / Node project."
            )
            status_lbl.config(
                text=f"✅ Exported {count} student(s). File ready for manual use.",
                fg=theme.SUCCESS,
            )
            log_activity(self.current_user, f"Exported remaining fees Excel ({count} students) for {month:02d}/{year}")

            # Best-effort: open the folder so the user can find the file quickly
            try:
                if os.name == "nt":
                    os.startfile(os.path.dirname(path) or path)
                elif shutil.which("xdg-open"):
                    os.system(f'xdg-open "{os.path.dirname(path)}"')
                elif shutil.which("open"):
                    os.system(f'open "{os.path.dirname(path)}"')
            except Exception:
                pass

        buttons = tk.Frame(body, bg=theme.WHITE)
        buttons.pack(fill=tk.X, pady=(10, 0))
        theme.primary_button(buttons, "📁 Export to Excel", export_excel, bg=theme.SUCCESS).pack(side=tk.LEFT)

    def open_attendance(self):
        """Open the single, authoritative user-facing Attendance screen.

        The legacy Attendance page remains in app.py for compatibility and
        to preserve its existing functionality/code, but it is not exposed
        as a second user-facing screen. All Attendance entry points route to
        smart_attendance.py, which is the unified Attendance experience.
        """
        return launch_attendance_window(self.root, self.user_role, self.current_user)

    def open_teacher_payroll(self):
        """Open the single, authoritative Teacher & Payroll module."""
        return launch_teacher_payroll_window(
            self.root, self.user_role, self.current_user
        )

    def show_page(self, key):
        if key == "teachers":
            return self.open_teacher_payroll()
        frame = self.pages.get(key)
        if not frame:
            return
        frame.tkraise()
        for k, btn in self.nav_buttons.items():
            theme.set_sidebar_active(btn, k == key)
        if key == "dashboard":
            self.refresh_dashboard()

    def open_ai_assistant(self):
        win = tk.Toplevel(self.root)
        win.title("AI Admin Assistant — AR School Management System")
        win.geometry("520x600")
        win.config(bg=theme.SILVER)

        header = tk.Frame(win, bg=theme.NAVY, pady=10, padx=14)
        header.pack(fill=tk.X)
        tk.Label(header, text="🤖 AI Admin Assistant", font=theme.FONT_H1, bg=theme.NAVY, fg="white").pack(anchor="w")
        tk.Label(header, text="Answers are read-only and grounded in your current database — never invented.",
                 font=theme.FONT_SMALL, bg=theme.NAVY, fg="#94a3b8").pack(anchor="w")

        chat_frame = tk.Frame(win, bg=theme.SILVER)
        chat_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        chat_canvas = tk.Canvas(chat_frame, bg=theme.SILVER, highlightthickness=0)
        chat_scroll = ttk.Scrollbar(chat_frame, orient=tk.VERTICAL, command=chat_canvas.yview)
        chat_body = tk.Frame(chat_canvas, bg=theme.SILVER)
        chat_body.bind("<Configure>", lambda e: chat_canvas.configure(scrollregion=chat_canvas.bbox("all")))
        chat_canvas.create_window((0, 0), window=chat_body, anchor="nw", width=470)
        chat_canvas.configure(yscrollcommand=chat_scroll.set)
        chat_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        chat_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        def add_bubble(text, is_user):
            row = tk.Frame(chat_body, bg=theme.SILVER)
            row.pack(fill=tk.X, pady=4, anchor="e" if is_user else "w")
            bubble = tk.Label(row, text=text, font=theme.FONT_BODY, bg=(theme.BRAND_BLUE if is_user else theme.WHITE),
                               fg=("white" if is_user else theme.TEXT_DARK), justify="left", anchor="w",
                               wraplength=360, padx=10, pady=8)
            bubble.pack(side=tk.RIGHT if is_user else tk.LEFT)
            win.after(10, lambda: chat_canvas.yview_moveto(1.0))

        add_bubble("Assalam-o-Alaikum! Ask me about students, fees, attendance, results, teachers, or finance — "
                   "I'll only answer from what's actually in the database.", False)

        suggestions = tk.Frame(win, bg=theme.SILVER)
        suggestions.pack(fill=tk.X, padx=10)
        for sq in ai_assistant.SUGGESTED_QUESTIONS[:3]:
            tk.Button(suggestions, text=sq, font=theme.FONT_SMALL, bg=theme.SILVER_BORDER, fg=theme.TEXT_DARK,
                      bd=0, padx=8, pady=4, cursor="hand2",
                      command=lambda t=sq: (entry.delete(0, tk.END), entry.insert(0, t), send())).pack(side=tk.LEFT, padx=(0, 6), pady=6)

        input_row = tk.Frame(win, bg=theme.SILVER, padx=10, pady=10)
        input_row.pack(fill=tk.X)
        entry = tk.Entry(input_row, font=theme.FONT_BODY)
        entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=6)

        def send():
            q = entry.get().strip()
            if not q:
                return
            add_bubble(q, True)
            entry.delete(0, tk.END)
            answer = ai_assistant.answer_question(self.user_role, q, self.current_user)
            add_bubble(answer, False)

        entry.bind("<Return>", lambda e: send())
        tk.Button(input_row, text="Send", command=send, bg=theme.BRAND_BLUE, fg="white",
                  font=theme.FONT_BODY_BOLD, bd=0, padx=16, cursor="hand2").pack(side=tk.LEFT, padx=(8, 0))

    def logout(self):
        if not messagebox.askyesno("Logout", "Are you sure you want to logout?"):
            return
        log_activity(self.current_user, "User logged out")
        try:
            self._maybe_auto_backup_on_exit()
        except Exception:
            pass
        self.root.destroy()
        login_root = tk.Tk()
        theme.apply_ttk_style()
        LoginWindow(login_root)
        login_root.mainloop()

    # ------------------------------------------
    # DASHBOARD
    # ------------------------------------------
    def build_dashboard_tab(self):
        self.dash_scroll_canvas = tk.Canvas(self.tab_dashboard, bg=theme.SILVER, highlightthickness=0)
        vscroll = ttk.Scrollbar(self.tab_dashboard, orient=tk.VERTICAL, command=self.dash_scroll_canvas.yview)
        self.dash_body = tk.Frame(self.dash_scroll_canvas, bg=theme.SILVER)
        self.dash_body.bind("<Configure>", lambda e: self.dash_scroll_canvas.configure(scrollregion=self.dash_scroll_canvas.bbox("all")))
        self.dash_scroll_canvas.create_window((0, 0), window=self.dash_body, anchor="nw")
        self.dash_scroll_canvas.configure(yscrollcommand=vscroll.set)
        self.dash_scroll_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vscroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.refresh_dashboard()

    def refresh_dashboard(self):
        for w in self.dash_body.winfo_children():
            w.destroy()

        tk.Label(self.dash_body, text=f"Welcome back, {self.current_user}", font=theme.FONT_H1,
                 bg=theme.SILVER, fg=theme.TEXT_DARK).pack(anchor="w", pady=(0, 2))
        tk.Label(self.dash_body, text=datetime.now().strftime("%A, %d %B %Y"), font=theme.FONT_SMALL,
                 bg=theme.SILVER, fg=theme.TEXT_MUTED).pack(anchor="w", pady=(0, 14))

        # ----- real stats from the database -----
        total_students = db.run("SELECT COUNT(*) FROM students WHERE COALESCE(status,'Active')='Active'", fetchone=True)[0]
        archived_students = db.run("SELECT COUNT(*) FROM students WHERE status='Archived'", fetchone=True)[0]
        total_teachers = db.run("SELECT COUNT(*) FROM teachers", fetchone=True)[0]
        today = datetime.now().strftime("%Y-%m-%d")
        present_today = db.run("SELECT COUNT(*) FROM attendance WHERE date=? AND status='Present'", (today,), fetchone=True)[0]
        absent_today = db.run("SELECT COUNT(*) FROM attendance WHERE date=? AND status='Absent'", (today,), fetchone=True)[0]
        pending_fees = db.run("SELECT COALESCE(SUM(total_fee-paid_fee),0) FROM students WHERE COALESCE(status,'Active')='Active' AND total_fee > paid_fee", fetchone=True)[0]

        cards_row = tk.Frame(self.dash_body, bg=theme.SILVER)
        cards_row.pack(fill=tk.X, pady=(0, 12))
        theme.stat_card(cards_row, "Active Students", total_students, theme.BRAND_BLUE,
                         subtitle=f"{archived_students} archived").pack(side=tk.LEFT, padx=(0, 12), fill=tk.BOTH, expand=True)
        theme.stat_card(cards_row, "Teachers", total_teachers, theme.SLATE).pack(side=tk.LEFT, padx=12, fill=tk.BOTH, expand=True)
        theme.stat_card(cards_row, "Present Today", present_today, theme.SUCCESS).pack(side=tk.LEFT, padx=12, fill=tk.BOTH, expand=True)
        theme.stat_card(cards_row, "Absent Today", absent_today, theme.DANGER).pack(side=tk.LEFT, padx=(12, 0), fill=tk.BOTH, expand=True)

        if rbac.can(self.user_role, "student.fee.view"):
            cards_row2 = tk.Frame(self.dash_body, bg=theme.SILVER)
            cards_row2.pack(fill=tk.X, pady=(0, 12))
            theme.stat_card(cards_row2, "Pending Fees (Rs.)", f"{pending_fees:,.0f}", theme.WARNING).pack(side=tk.LEFT, padx=(0, 12), fill=tk.BOTH, expand=True)
            if rbac.can(self.user_role, "accounting.dashboard"):
                totals = accounting.dashboard_totals(self.user_role)
                theme.stat_card(cards_row2, "This Month Revenue (Rs.)", f"{totals['month_revenue']:,.0f}", theme.SUCCESS).pack(side=tk.LEFT, padx=12, fill=tk.BOTH, expand=True)
                theme.stat_card(cards_row2, "This Month Expenses (Rs.)", f"{totals['month_expense']:,.0f}", theme.DANGER).pack(side=tk.LEFT, padx=12, fill=tk.BOTH, expand=True)
                theme.stat_card(cards_row2, "Net Income (Rs.)", f"{totals['net_income']:,.0f}",
                                 theme.SUCCESS if totals['net_income'] >= 0 else theme.DANGER).pack(side=tk.LEFT, padx=(12, 0), fill=tk.BOTH, expand=True)

        # ----- quick actions -----
        actions_card, actions_body = theme.section_card(self.dash_body, "Quick Actions")
        actions_card.pack(fill=tk.X, pady=(0, 12))
        qa = [("➕ Add Student", "__admission__", rbac.can(self.user_role, "student.add")),
              ("🗓️ Attendance", "__attendance__", True),
              ("💰 Fee Management", "__fee_management__", rbac.can(self.user_role, "student.fee.view")),
              ("📁 Export Remaining Fees", "__whatsapp_fee_reminders__", rbac.can(self.user_role, "fee.reports.view")),
              ("👨‍🏫 Add Teacher", "teachers", self.can_teachers and rbac.can(self.user_role, "teacher.add")),
              ("📝 Enter Marks", "results", self.can_results and rbac.can(self.user_role, "results.marks.edit")),
              ("💰 View Finance", "accounting", rbac.can(self.user_role, "accounting.dashboard"))]
        for label, target, allowed in qa:
            if not allowed:
                continue
            if target == "__fee_management__":
                command = self.open_fee_management
            elif target == "__attendance__":
                command = self.open_attendance
            elif target == "__admission__":
                command = self.open_admission
            elif target == "__whatsapp_fee_reminders__":
                command = self.open_whatsapp_fee_reminders
            else:
                command = lambda t=target: self.show_page(t)
            theme.primary_button(actions_body, label, command).pack(side=tk.LEFT, padx=(0, 10), pady=6)

        # ----- recent activity, side by side -----
        lists_row = tk.Frame(self.dash_body, bg=theme.SILVER)
        lists_row.pack(fill=tk.BOTH, expand=True)

        admissions_card, admissions_body = theme.section_card(lists_row, "Recent Admissions")
        admissions_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 12))
        recent_students = db.run(
            "SELECT student_id, name, class_sec FROM students WHERE COALESCE(status,'Active')='Active' ORDER BY ROWID DESC LIMIT 5",
            fetchall=True)
        if recent_students:
            for s_id, name, cls in recent_students:
                row = tk.Frame(admissions_body, bg=theme.WHITE)
                row.pack(fill=tk.X, pady=2)
                tk.Label(row, text=f"{name}", font=theme.FONT_BODY, bg=theme.WHITE, fg=theme.TEXT_DARK).pack(side=tk.LEFT)
                tk.Label(row, text=f"{s_id} · {cls}", font=theme.FONT_SMALL, bg=theme.WHITE, fg=theme.TEXT_MUTED).pack(side=tk.RIGHT)
        else:
            tk.Label(admissions_body, text="No students admitted yet.", font=theme.FONT_SMALL, bg=theme.WHITE, fg=theme.TEXT_MUTED).pack(anchor="w", pady=8)

        payments_card, payments_body = theme.section_card(lists_row, "Recent Payments")
        payments_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(12, 0))
        if rbac.can(self.user_role, "accounting.revenue.view"):
            recent_payments = db.run(
                "SELECT student_id, amount, date FROM accounting_revenue ORDER BY id DESC LIMIT 5", fetchall=True)
            if recent_payments:
                for s_id, amount, date in recent_payments:
                    row = tk.Frame(payments_body, bg=theme.WHITE)
                    row.pack(fill=tk.X, pady=2)
                    tk.Label(row, text=f"{s_id or 'General'}", font=theme.FONT_BODY, bg=theme.WHITE, fg=theme.TEXT_DARK).pack(side=tk.LEFT)
                    tk.Label(row, text=f"Rs. {amount:,.0f} · {date}", font=theme.FONT_SMALL, bg=theme.WHITE, fg=theme.TEXT_MUTED).pack(side=tk.RIGHT)
            else:
                tk.Label(payments_body, text="No payments recorded yet.", font=theme.FONT_SMALL, bg=theme.WHITE, fg=theme.TEXT_MUTED).pack(anchor="w", pady=8)
        else:
            tk.Label(payments_body, text="You don't have permission to view finance records.", font=theme.FONT_SMALL, bg=theme.WHITE, fg=theme.TEXT_MUTED).pack(anchor="w", pady=8)

        if HAS_MATPLOTLIB and rbac.can(self.user_role, "accounting.dashboard"):
            theme.primary_button(self.dash_body, "📊 View Revenue vs Expense Chart", self.show_finance_chart, bg=theme.SLATE).pack(anchor="w", pady=(12, 0))

    # ------------------------------------------
    # TAB 1: ADMISSION
    # ------------------------------------------
    def build_admission_tab(self):
        """Students page: directory + one professional New Admission entry point.

        The old inline admission form is intentionally no longer rendered.
        New admissions are handled exclusively by student_admission.py so the
        user never sees two different admission forms.
        """
        can_add = rbac.can(self.user_role, "student.add")
        can_delete = rbac.can(self.user_role, "student.delete")

        header = tk.Frame(self.tab_admission, bg=theme.NAVY, padx=16, pady=12)
        header.pack(fill=tk.X, padx=10, pady=(10, 8))
        tk.Label(header, text="🎓 STUDENT MANAGEMENT", font=theme.FONT_H1,
                 bg=theme.NAVY, fg="white").pack(side=tk.LEFT)
        if can_add:
            theme.primary_button(header, "➕ New Student Admission", self.open_admission,
                                 bg=theme.SUCCESS).pack(side=tk.RIGHT)

        directory = theme.section_card(self.tab_admission, "Student Directory")
        directory_card, body = directory
        directory_card.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        search_bar = tk.Frame(body, bg=theme.WHITE)
        search_bar.pack(fill=tk.X, pady=(0, 4))
        tk.Label(search_bar, text="Search:", font=theme.FONT_SMALL, bg=theme.WHITE).pack(side=tk.LEFT, padx=(0, 4))
        self.ent_search = tk.Entry(search_bar, font=theme.FONT_BODY)
        self.ent_search.pack(side=tk.LEFT, padx=(0, 8), ipady=3, fill=tk.X, expand=True)
        self.ent_search.bind("<KeyRelease>", self.load_student_table)

        # ---- Filters row: Fee / Class / Status ----
        filter_bar = tk.Frame(body, bg=theme.WHITE)
        filter_bar.pack(fill=tk.X, pady=(0, 8))

        tk.Label(filter_bar, text="Fee:", font=theme.FONT_SMALL, bg=theme.WHITE,
                 fg=theme.TEXT_MUTED).pack(side=tk.LEFT, padx=(0, 4))
        self.cmb_fee_filter = ttk.Combobox(
            filter_bar,
            values=[
                "All Fees",
                "Fee Pending",
                "Fee Paid",
                "Fee Overdue",
                "Partial Payment",
            ],
            state="readonly",
            width=16,
            font=theme.FONT_SMALL,
        )
        self.cmb_fee_filter.set("All Fees")
        self.cmb_fee_filter.pack(side=tk.LEFT, padx=(0, 12))
        self.cmb_fee_filter.bind("<<ComboboxSelected>>", self.load_student_table)

        tk.Label(filter_bar, text="Class:", font=theme.FONT_SMALL, bg=theme.WHITE,
                 fg=theme.TEXT_MUTED).pack(side=tk.LEFT, padx=(0, 4))
        self.cmb_class_filter = ttk.Combobox(
            filter_bar, values=["All Classes"], state="readonly", width=14, font=theme.FONT_SMALL,
        )
        self.cmb_class_filter.set("All Classes")
        self.cmb_class_filter.pack(side=tk.LEFT, padx=(0, 12))
        self.cmb_class_filter.bind("<<ComboboxSelected>>", self.load_student_table)

        tk.Label(filter_bar, text="Status:", font=theme.FONT_SMALL, bg=theme.WHITE,
                 fg=theme.TEXT_MUTED).pack(side=tk.LEFT, padx=(0, 4))
        self.cmb_status_filter = ttk.Combobox(
            filter_bar,
            values=["Active Only", "Archived Only", "All Status"],
            state="readonly",
            width=14,
            font=theme.FONT_SMALL,
        )
        self.cmb_status_filter.set("Active Only")
        self.cmb_status_filter.pack(side=tk.LEFT, padx=(0, 12))
        self.cmb_status_filter.bind("<<ComboboxSelected>>", self.load_student_table)

        # Legacy checkbox kept in sync with Status filter for older call sites
        self.show_archived_var = tk.BooleanVar(value=False)

        theme.primary_button(
            filter_bar, "↻ Clear Filters", self._clear_student_filters, bg=theme.SLATE,
        ).pack(side=tk.LEFT, padx=(4, 8))

        self.lbl_student_count = tk.Label(
            filter_bar, text="", font=theme.FONT_SMALL, bg=theme.WHITE, fg=theme.TEXT_MUTED,
        )
        self.lbl_student_count.pack(side=tk.RIGHT, padx=4)

        action_bar = tk.Frame(body, bg=theme.WHITE)
        action_bar.pack(fill=tk.X, pady=(0, 8))
        if can_add:
            theme.primary_button(action_bar, "➕ New Student Admission", self.open_admission,
                                 bg=theme.SUCCESS).pack(side=tk.LEFT, padx=(0, 8))
        can_edit = rbac.can(self.user_role, "student.edit")
        if can_edit:
            theme.primary_button(action_bar, "✏️ Edit Student", self.edit_selected_student,
                                 bg=theme.SLATE).pack(side=tk.LEFT, padx=(0, 8))
        if can_delete:
            theme.primary_button(action_bar, "🗑 Remove Student", self.remove_selected_student,
                                 bg=theme.DANGER).pack(side=tk.LEFT, padx=(0, 8))
        if rbac.can(self.user_role, "student.view"):
            theme.primary_button(action_bar, "👤 View Profile", self.open_student_profile,
                                 bg=theme.BRAND_BLUE).pack(side=tk.LEFT, padx=(8, 0))
            theme.primary_button(action_bar, "🪪 ID Card", self.reprint_selected_id_card,
                                 bg=theme.SLATE).pack(side=tk.LEFT, padx=(8, 0))
        theme.primary_button(
            action_bar, "📁 Export Excel", self.export_student_directory_excel, bg=theme.SUCCESS,
        ).pack(side=tk.RIGHT, padx=(8, 0))

        cols = ("id", "name", "fname", "class", "phone", "status")
        if rbac.can(self.user_role, "student.fee.view"):
            cols = ("id", "name", "fname", "class", "phone", "total_fee", "paid_fee", "balance", "status")

        table_frame = tk.Frame(body, bg=theme.WHITE)
        table_frame.pack(fill=tk.BOTH, expand=True)
        self.tree_student = ttk.Treeview(table_frame, columns=cols, show="headings")
        col_widths = {"id": 90, "name": 150, "fname": 130, "class": 80, "phone": 110,
                      "total_fee": 85, "paid_fee": 85, "balance": 85, "status": 85}
        for col in cols:
            self.tree_student.heading(col, text=col.upper().replace("_", " "))
            self.tree_student.column(col, width=col_widths.get(col, 90), anchor="center")

        scroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree_student.yview)
        self.tree_student.configure(yscroll=scroll.set)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree_student.pack(fill=tk.BOTH, expand=True)
        self.load_student_table()

    def browse_photo(self):
        file_path = filedialog.askopenfilename(filetypes=[("Image Files", "*.jpg *.png *.jpeg")])
        if file_path:
            self.photo_path_var.set(file_path)
            messagebox.showinfo("Photo Selected", "Image uploaded successfully!")

    def _clear_student_filters(self):
        if hasattr(self, "ent_search"):
            self.ent_search.delete(0, tk.END)
        if hasattr(self, "cmb_fee_filter"):
            self.cmb_fee_filter.set("All Fees")
        if hasattr(self, "cmb_class_filter"):
            self.cmb_class_filter.set("All Classes")
        if hasattr(self, "cmb_status_filter"):
            self.cmb_status_filter.set("Active Only")
        if hasattr(self, "show_archived_var"):
            self.show_archived_var.set(False)
        self.load_student_table()

    def _refresh_class_filter_options(self):
        """Populate Class filter from distinct class_sec values in DB."""
        if not hasattr(self, "cmb_class_filter"):
            return
        current = self.cmb_class_filter.get() or "All Classes"
        rows = db.run(
            "SELECT DISTINCT class_sec FROM students "
            "WHERE class_sec IS NOT NULL AND TRIM(class_sec) <> '' "
            "ORDER BY class_sec",
            fetchall=True,
        ) or []
        classes = ["All Classes"] + [r[0] for r in rows if r[0]]
        self.cmb_class_filter["values"] = classes
        if current in classes:
            self.cmb_class_filter.set(current)
        else:
            self.cmb_class_filter.set("All Classes")

    def _student_fee_maps(self):
        """Build per-student outstanding balance and overdue flags.

        Prefer monthly fee_cycles ledger when present; fall back to
        students.total_fee / paid_fee so older databases still filter correctly.
        """
        outstanding = {}  # student_id -> remaining amount
        has_overdue = set()
        has_partial = set()
        try:
            rows = db.run(
                "SELECT student_id, "
                "COALESCE(SUM(amount_due - amount_paid), 0), "
                "MAX(CASE WHEN status='OVERDUE' THEN 1 ELSE 0 END), "
                "MAX(CASE WHEN status='PARTIAL' THEN 1 ELSE 0 END) "
                "FROM fee_cycles GROUP BY student_id",
                fetchall=True,
            ) or []
            for sid, bal, overdue, partial in rows:
                outstanding[sid] = float(bal or 0)
                if overdue:
                    has_overdue.add(sid)
                if partial:
                    has_partial.add(sid)
        except Exception:
            pass
        return outstanding, has_overdue, has_partial

    def load_student_table(self, ev=None):
        search = self.ent_search.get().strip() if hasattr(self, "ent_search") else ""
        self.tree_student.delete(*self.tree_student.get_children())
        self._refresh_class_filter_options()

        fee_filter = getattr(self, "cmb_fee_filter", None)
        fee_filter = fee_filter.get() if fee_filter else "All Fees"
        class_filter = getattr(self, "cmb_class_filter", None)
        class_filter = class_filter.get() if class_filter else "All Classes"
        status_filter = getattr(self, "cmb_status_filter", None)
        status_filter = status_filter.get() if status_filter else "Active Only"

        # Keep legacy checkbox in sync
        if hasattr(self, "show_archived_var"):
            self.show_archived_var.set(status_filter in ("Archived Only", "All Status"))

        rows = db.run(
            "SELECT student_id, name, father_name, phone, class_sec, "
            "COALESCE(total_fee, 0), COALESCE(paid_fee, 0), COALESCE(status, 'Active') "
            "FROM students",
            fetchall=True,
        ) or []

        cycle_outstanding, has_overdue, has_partial = self._student_fee_maps()
        show_fee = rbac.can(self.user_role, "student.fee.view")
        shown = 0

        for s_id, name, fname, phone, cls, total_f, paid_f, status in rows:
            status = status or "Active"
            total_f = float(total_f or 0)
            paid_f = float(paid_f or 0)

            # Status filter
            if status_filter == "Active Only" and status != "Active":
                continue
            if status_filter == "Archived Only" and status != "Archived":
                continue

            # Class filter
            if class_filter and class_filter != "All Classes":
                if (cls or "") != class_filter:
                    continue

            # Search (ID / name / father / phone / class)
            if search:
                q = search.lower()
                hay = " ".join([
                    str(s_id or ""), str(name or ""), str(fname or ""),
                    str(phone or ""), str(cls or ""),
                ]).lower()
                if q not in hay:
                    continue

            # Fee balance: ledger first, else classic total_fee - paid_fee
            if s_id in cycle_outstanding:
                bal = cycle_outstanding[s_id]
            else:
                bal = total_f - paid_f

            # Fee filters
            if fee_filter == "Fee Pending":
                if bal <= 0:
                    continue
            elif fee_filter == "Fee Paid":
                if bal > 0:
                    continue
            elif fee_filter == "Fee Overdue":
                if s_id not in has_overdue:
                    # Fallback: treat any positive balance with no cycles as not overdue
                    continue
            elif fee_filter == "Partial Payment":
                if s_id not in has_partial:
                    # Fallback partial: some paid but still outstanding
                    if not (paid_f > 0 and bal > 0 and s_id not in cycle_outstanding):
                        continue

            if show_fee:
                self.tree_student.insert(
                    "", tk.END,
                    values=(s_id, name, fname or "", cls or "", phone or "",
                            f"{total_f:,.0f}", f"{paid_f:,.0f}", f"{bal:,.0f}", status),
                )
            else:
                self.tree_student.insert(
                    "", tk.END,
                    values=(s_id, name, fname or "", cls or "", phone or "", status),
                )
            shown += 1

        if hasattr(self, "lbl_student_count"):
            parts = [f"{shown} student(s)"]
            if fee_filter != "All Fees":
                parts.append(fee_filter)
            if class_filter != "All Classes":
                parts.append(class_filter)
            if status_filter != "Active Only":
                parts.append(status_filter)
            self.lbl_student_count.config(text=" · ".join(parts))

    def export_student_directory_excel(self):
        """Export the currently filtered Student Directory rows to an Excel file.

        Columns match what the user sees (including fee columns when permitted).
        A header row notes active filters so the file is self-describing.
        """
        if not rbac.can(self.user_role, "student.view"):
            messagebox.showerror("Permission Denied", "You are not allowed to export the student list.", parent=self.root)
            return

        items = self.tree_student.get_children()
        if not items:
            messagebox.showinfo(
                "Nothing to Export",
                "Current filters show no students.\n"
                "Change filters or clear them, then try again.",
                parent=self.root,
            )
            return

        show_fee = rbac.can(self.user_role, "student.fee.view")
        if show_fee:
            headers = [
                "Student ID", "Name", "Father Name", "Class", "Phone",
                "Total Fee", "Paid Fee", "Balance", "Status",
            ]
        else:
            headers = ["Student ID", "Name", "Father Name", "Class", "Phone", "Status"]

        fee_f = getattr(self, "cmb_fee_filter", None)
        fee_f = fee_f.get() if fee_f else "All Fees"
        class_f = getattr(self, "cmb_class_filter", None)
        class_f = class_f.get() if class_f else "All Classes"
        status_f = getattr(self, "cmb_status_filter", None)
        status_f = status_f.get() if status_f else "Active Only"
        search_q = self.ent_search.get().strip() if hasattr(self, "ent_search") else ""

        default_name = f"Student_List_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Save Student List as Excel",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Workbook", "*.xlsx"), ("All Files", "*.*")],
            parent=self.root,
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            messagebox.showerror(
                "Missing Library",
                "openpyxl is required for Excel export.\n"
                "Install with: pip install openpyxl",
                parent=self.root,
            )
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Students"

            # Title + filter summary
            title = "Student Directory Export — AR School Management System"
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            cell = ws.cell(row=1, column=1, value=title)
            cell.font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F172A")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[1].height = 24

            filter_bits = [f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}"]
            filter_bits.append(f"Fee filter: {fee_f}")
            filter_bits.append(f"Class: {class_f}")
            filter_bits.append(f"Status: {status_f}")
            if search_q:
                filter_bits.append(f"Search: {search_q}")
            filter_bits.append(f"Rows: {len(items)}")
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
            ws.cell(row=2, column=1, value="  |  ".join(filter_bits)).font = Font(
                name="Segoe UI", size=9, italic=True, color="64748B"
            )

            header_fill = PatternFill("solid", fgColor="0284C7")
            header_font = Font(name="Segoe UI", bold=True, color="FFFFFF")
            thin = Border(
                left=Side(style="thin", color="E2E8F0"),
                right=Side(style="thin", color="E2E8F0"),
                top=Side(style="thin", color="E2E8F0"),
                bottom=Side(style="thin", color="E2E8F0"),
            )
            for col_idx, h in enumerate(headers, start=1):
                c = ws.cell(row=4, column=col_idx, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin

            for row_idx, item_id in enumerate(items, start=5):
                values = self.tree_student.item(item_id, "values")
                for col_idx, val in enumerate(values, start=1):
                    c = ws.cell(row=row_idx, column=col_idx, value=val)
                    c.font = Font(name="Segoe UI", size=10)
                    c.border = thin
                    c.alignment = Alignment(horizontal="center" if col_idx != 2 else "left")

            from openpyxl.utils import get_column_letter
            widths = {1: 14, 2: 22, 3: 18, 4: 12, 5: 14, 6: 12, 7: 12, 8: 12, 9: 12}
            for col_idx in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_idx, 14)

            wb.save(path)
        except Exception as e:
            messagebox.showerror("Export Failed", f"Could not write Excel file:\n{e}", parent=self.root)
            return

        log_activity(
            self.current_user,
            f"Exported student directory Excel ({len(items)} rows) "
            f"filters=[fee={fee_f}, class={class_f}, status={status_f}, search={search_q or '-'}]",
        )
        messagebox.showinfo(
            "Export Complete",
            f"{len(items)} student(s) exported.\n\nSaved to:\n{path}",
            parent=self.root,
        )

    def save_student(self):
        if not rbac.can(self.user_role, "student.add"):
            messagebox.showerror("Permission Denied", "You are not allowed to add students.")
            return
        s_id = self.lbl_auto_id.cget("text")
        name = self.ent_name.get().strip()
        cls = self.ent_class.get().strip()
        if not name or not cls:
            messagebox.showerror("Error", "Student Name and Class are required!")
            return

        if hasattr(self, 'ent_total_fee'):
            total_f, ok1 = safe_float(self.ent_total_fee.get(), "Total Fee", default=0.0)
            paid_f, ok2 = safe_float(self.ent_paid_fee.get(), "Paid Fee", default=0.0)
            if not (ok1 and ok2):
                return
        else:
            total_f, paid_f = 0.0, 0.0

        if paid_f > total_f:
            if not messagebox.askyesno(
                "Overpayment Warning",
                f"Paid Fee (Rs. {paid_f:.2f}) is more than Total Fee (Rs. {total_f:.2f}).\n"
                "This will be recorded as an advance/overpayment. Continue?"):
                return

        db.run("""INSERT INTO students
                  (student_id, name, father_name, dob, phone, address, class_sec, photo_path,
                   prev_education, total_fee, paid_fee, status)
                  VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Active')""",
               (s_id, name, self.ent_fname.get(), self.ent_dob.get(), self.ent_phone.get(),
                self.ent_address.get(), cls, self.photo_path_var.get(), self.ent_prev_edu.get(), total_f, paid_f),
               commit=True)

        if paid_f > 0 and rbac.can(self.user_role, "student.fee.edit"):
            accounting.record_fee_revenue(self.user_role, s_id, paid_f, self.current_user, description="Initial fee payment at admission")

        log_activity(self.current_user, f"Admitted new student {name} ({s_id})")
        messagebox.showinfo("Success", f"Admission Complete!\nStudent ID: {s_id}")
        if paid_f > 0 and rbac.can(self.user_role, "student.fee.edit"):
            self._offer_fee_receipt(s_id, name, self.ent_fname.get(), cls, total_f, 0.0, paid_f)
        self.clear_admission_form()
        self.load_student_table()

    def fill_form_from_table(self, ev):
        selected = self.tree_student.focus()
        vals = self.tree_student.item(selected, "values")
        if not vals:
            return
        s_id = vals[0]
        r = db.run("SELECT * FROM students WHERE student_id=?", (s_id,), fetchone=True)
        if not r:
            return
        self.lbl_auto_id.config(text=r[0])
        self.ent_name.delete(0, tk.END); self.ent_name.insert(0, r[1])
        self.ent_fname.delete(0, tk.END); self.ent_fname.insert(0, r[2] or "")
        self.ent_dob.delete(0, tk.END); self.ent_dob.insert(0, r[3] or "")
        self.ent_phone.delete(0, tk.END); self.ent_phone.insert(0, r[4] or "")
        self.ent_address.delete(0, tk.END); self.ent_address.insert(0, r[5] or "")
        self.ent_class.delete(0, tk.END); self.ent_class.insert(0, r[6] or "")
        self.photo_path_var.set(r[7] or "")
        self.ent_prev_edu.delete(0, tk.END); self.ent_prev_edu.insert(0, r[8] or "")
        if hasattr(self, 'ent_total_fee'):
            self._prev_paid_fee = r[10]
            self.ent_total_fee.config(state="normal"); self.ent_total_fee.delete(0, tk.END); self.ent_total_fee.insert(0, str(r[9]))
            self.ent_paid_fee.config(state="normal"); self.ent_paid_fee.delete(0, tk.END); self.ent_paid_fee.insert(0, str(r[10]))
            if not rbac.can(self.user_role, "student.fee.edit"):
                self.ent_total_fee.config(state="disabled")
                self.ent_paid_fee.config(state="disabled")

    def update_student(self):
        if not rbac.can(self.user_role, "student.edit"):
            messagebox.showerror("Permission Denied", "You are not allowed to edit students.")
            return
        s_id = self.lbl_auto_id.cget("text")
        can_fee = rbac.can(self.user_role, "student.fee.edit")
        prev_paid = getattr(self, "_prev_paid_fee", 0)

        if hasattr(self, 'ent_total_fee') and can_fee:
            # Only a role with student.fee.edit may change fee figures — this
            # is enforced here (the data layer), not just by disabling the
            # widgets in the UI, so a permission_overrides change can never
            # silently be bypassed.
            total_f, ok1 = safe_float(self.ent_total_fee.get(), "Total Fee", default=0.0)
            paid_f, ok2 = safe_float(self.ent_paid_fee.get(), "Paid Fee", default=0.0)
            if not (ok1 and ok2):
                return
            if paid_f > total_f:
                if not messagebox.askyesno(
                    "Overpayment Warning",
                    f"Paid Fee (Rs. {paid_f:.2f}) is more than Total Fee (Rs. {total_f:.2f}).\n"
                    "This will be recorded as an advance/overpayment. Continue?"):
                    return
        else:
            # No permission to touch fee fields (or this build has no fee
            # widgets at all) — keep whatever is already in the database.
            row = db.run("SELECT total_fee, paid_fee FROM students WHERE student_id=?", (s_id,), fetchone=True)
            total_f, paid_f = (row if row else (0.0, 0.0))

        db.run("""UPDATE students SET name=?, father_name=?, dob=?, phone=?, address=?, class_sec=?,
                  photo_path=?, prev_education=?, total_fee=?, paid_fee=? WHERE student_id=?""",
               (self.ent_name.get(), self.ent_fname.get(), self.ent_dob.get(), self.ent_phone.get(),
                self.ent_address.get(), self.ent_class.get(), self.photo_path_var.get(),
                self.ent_prev_edu.get(), total_f, paid_f, s_id), commit=True)

        # New fee revenue only for the *increase* in paid_fee — avoids double-counting.
        delta = paid_f - prev_paid
        if delta > 0 and can_fee:
            accounting.record_fee_revenue(self.user_role, s_id, delta, self.current_user)

        log_activity(self.current_user, f"Updated student record for {s_id}")
        messagebox.showinfo("Success", "Student Details Updated!")
        if delta > 0 and can_fee:
            self._offer_fee_receipt(s_id, self.ent_name.get(), self.ent_fname.get(), self.ent_class.get(),
                                     total_f, prev_paid, delta)
        self.clear_admission_form()
        self.load_student_table()

    def _offer_fee_receipt(self, s_id, name, father_name, cls, total_fee, previous_paid, current_payment):
        """Shown once, right after a fee payment is successfully recorded.
        This only RENDERS a PDF from the numbers already computed by the
        caller — it never touches accounting_revenue itself, so opening
        this dialog (or generating the slip twice) can never create a
        duplicate accounting entry; see reports.generate_fee_receipt()."""
        balance = total_fee - (previous_paid + current_payment)
        receipt_no = f"RCPT-{s_id}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        # Optional one-time admission fee breakout for the PDF lines.
        try:
            _adm = reports.get_admission_fee_status(s_id)
        except Exception:
            _adm = None
        _adm_fee = _adm["charged"] if _adm else None
        _adm_paid = _adm["paid"] if _adm else None

        win = tk.Toplevel(self.root)
        win.title("Fee Payment Recorded")
        win.geometry("360x200")
        win.config(bg=theme.WHITE)
        tk.Label(win, text="✅ Payment Recorded", font=theme.FONT_H1, bg=theme.WHITE, fg=theme.SUCCESS).pack(pady=(16, 4))
        tk.Label(win, text=f"Rs. {current_payment:,.0f} recorded for {name} ({s_id}).",
                 font=theme.FONT_BODY, bg=theme.WHITE, wraplength=320).pack(pady=(0, 14))

        def do_view():
            out_path = os.path.join(os.getcwd(), f"Fee_Receipt_{s_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf")
            reports.generate_fee_receipt(
                receipt_no, s_id, name, father_name, cls, total_fee, previous_paid,
                current_payment, balance, datetime.now().strftime("%Y-%m-%d"),
                self.current_user, out_path,
                admission_fee=_adm_fee, admission_fee_paid=_adm_paid,
                monthly_fee=total_fee,
            )
            log_activity(self.current_user, f"Generated fee receipt for {s_id}")
            # Best-effort open in the OS default PDF viewer so the user can
            # print from there — we don't pretend to control a physical
            # printer directly, since that isn't reliably available across
            # OSes from a Tkinter desktop app without extra dependencies.
            opened = False
            try:
                if os.name == "nt":
                    os.startfile(out_path)
                    opened = True
                elif shutil.which("xdg-open"):
                    os.system(f'xdg-open "{out_path}"')
                    opened = True
                elif shutil.which("open"):
                    os.system(f'open "{out_path}"')
                    opened = True
            except Exception:
                opened = False
            if opened:
                messagebox.showinfo("Receipt Ready", f"Receipt opened for viewing/printing:\n{out_path}")
            else:
                messagebox.showinfo("Receipt Saved", f"Receipt saved (no default PDF viewer detected to auto-open):\n{out_path}")
            win.destroy()

        def do_save_as():
            path = filedialog.asksaveasfilename(defaultextension=".pdf", initialfile=f"Fee_Receipt_{s_id}.pdf",
                                                 filetypes=[("PDF Files", "*.pdf")])
            if not path:
                return
            reports.generate_fee_receipt(
                receipt_no, s_id, name, father_name, cls, total_fee, previous_paid,
                current_payment, balance, datetime.now().strftime("%Y-%m-%d"),
                self.current_user, path,
                admission_fee=_adm_fee, admission_fee_paid=_adm_paid,
                monthly_fee=total_fee,
            )
            log_activity(self.current_user, f"Saved fee receipt for {s_id} to {path}")
            messagebox.showinfo("Saved", f"Receipt saved:\n{path}")
            win.destroy()

        btn_row = tk.Frame(win, bg=theme.WHITE)
        btn_row.pack(pady=6)
        theme.primary_button(btn_row, "🖨 Print Slip", do_view).pack(side=tk.LEFT, padx=6)
        theme.primary_button(btn_row, "💾 Save Slip As...", do_save_as, bg=theme.SLATE).pack(side=tk.LEFT, padx=6)
        tk.Button(win, text="Skip", command=win.destroy, bg=theme.WHITE, fg=theme.TEXT_MUTED, bd=0,
                  font=theme.FONT_SMALL, cursor="hand2").pack(pady=(6, 0))

    def _get_selected_student_id(self):
        """Returns (student_id, name) for the currently selected directory
        row, or (None, None) if nothing is selected."""
        selected = self.tree_student.focus()
        values = self.tree_student.item(selected, "values") if selected else ()
        if not values:
            return None, None
        return values[0], values[1]

    def remove_selected_student(self):
        """Permanently remove the currently selected student."""
        if not rbac.can(self.user_role, "student.delete"):
            messagebox.showerror(
                "Permission Denied",
                "You are not allowed to remove students.",
                parent=self.root
            )
            return

        # Get the selected row directly from the Students Directory.
        selected = self.tree_student.selection()
        if not selected:
            messagebox.showwarning(
                "Select Student",
                "Pehle Students Directory se student select karein.",
                parent=self.root
            )
            return

        item_id = selected[0]
        values = self.tree_student.item(item_id, "values")
        if not values:
            messagebox.showwarning(
                "Select Student",
                "Selected student ki information nahi mili.",
                parent=self.root
            )
            return

        student_id = str(values[0]).strip()
        student_name = str(values[1]).strip() if len(values) > 1 else student_id

        # Verify the record before deleting it.
        row = db.run(
            "SELECT student_id, name FROM students WHERE student_id=?",
            (student_id,),
            fetchone=True
        )
        if not row:
            messagebox.showerror(
                "Student Not Found",
                f"Student ID '{student_id}' database mein nahi mila.",
                parent=self.root
            )
            self.load_student_table()
            return

        confirm = messagebox.askyesno(
            "Confirm Remove Student",
            f"Student ko permanently remove karna hai?\n\n"
            f"Student: {row[1]}\n"
            f"Student ID: {row[0]}\n\n"
            "⚠️ Ye action undo nahi ho sakta.",
            icon="warning",
            parent=self.root
        )
        if not confirm:
            return

        try:
            # Remove the profile/admission-extra row first.
            db.run(
                "DELETE FROM student_admission_extra WHERE student_id=?",
                (student_id,),
                commit=True
            )

            # Remove the main student record.
            db.run(
                "DELETE FROM students WHERE student_id=?",
                (student_id,),
                commit=True
            )

            # Confirm that deletion actually happened.
            still_there = db.run(
                "SELECT student_id FROM students WHERE student_id=?",
                (student_id,),
                fetchone=True
            )
            if still_there:
                raise RuntimeError(
                    "Student database se remove nahi hua."
                )

            log_activity(
                self.current_user,
                f"Permanently removed student record {student_id} ({student_name})"
            )

            # Refresh directory and clear any old selection.
            self.load_student_table()

            messagebox.showinfo(
                "Student Removed",
                f"Student '{student_name}' ({student_id}) successfully remove ho gaya.",
                parent=self.root
            )

        except Exception as exc:
            messagebox.showerror(
                "Remove Failed",
                f"Student remove nahi ho saka.\n\nError:\n{exc}",
                parent=self.root
            )

    def edit_selected_student(self):
        """Edit the complete selected student profile, including admission-extra data."""
        if not rbac.can(self.user_role, "student.edit"):
            messagebox.showerror("Permission Denied", "You are not allowed to edit students.")
            return

        s_id, _ = self._get_selected_student_id()
        if not s_id:
            messagebox.showinfo("Select Student", "Please select a student from the directory first.")
            return

        row = db.run(
            """SELECT student_id, name, father_name, dob, phone, address, class_sec,
                      photo_path, prev_education, total_fee, paid_fee, status
               FROM students WHERE student_id=?""",
            (s_id,), fetchone=True)
        if not row:
            messagebox.showerror("Error", "Student not found.")
            return

        extra = db.run(
            """SELECT gender, blood_group, nationality, religion, mother_name, guardian_name,
                      guardian_cnic, occupation, alt_phone, email, current_address, city, area,
                      admission_date, academic_year, admission_type, emergency_contact_name,
                      emergency_contact_phone, emergency_relationship, emergency_notes
               FROM student_admission_extra WHERE student_id=?""",
            (s_id,), fetchone=True)

        base_keys = [
            "student_id", "name", "father_name", "dob", "phone", "address",
            "class_sec", "photo_path", "prev_education", "total_fee", "paid_fee", "status"
        ]
        base = dict(zip(base_keys, row))

        extra_keys = [
            "gender", "blood_group", "nationality", "religion", "mother_name",
            "guardian_name", "guardian_cnic", "occupation", "alt_phone", "email",
            "current_address", "city", "area", "admission_date", "academic_year",
            "admission_type", "emergency_contact_name", "emergency_contact_phone",
            "emergency_relationship", "emergency_notes"
        ]
        extra_data = dict(zip(extra_keys, extra)) if extra else {k: "" for k in extra_keys}

        can_fee = rbac.can(self.user_role, "student.fee.edit")
        previous_paid = float(base.get("paid_fee") or 0)

        win = tk.Toplevel(self.root)
        win.title(f"Edit Complete Student Profile — {s_id}")
        win.geometry("900x760")
        win.minsize(760, 620)
        win.config(bg=theme.SILVER)
        win.transient(self.root)
        win.grab_set()

        header = tk.Frame(win, bg=theme.NAVY, padx=18, pady=12)
        header.pack(fill=tk.X)
        tk.Label(
            header, text=f"✏️ EDIT STUDENT — {s_id}",
            font=theme.FONT_H1, bg=theme.NAVY, fg="white"
        ).pack(side=tk.LEFT)
        tk.Label(
            header, text="Complete Student Information",
            font=theme.FONT_SMALL, bg=theme.NAVY, fg=theme.BRAND_BLUE_LIGHT
        ).pack(side=tk.RIGHT)

        outer = tk.Frame(win, bg=theme.SILVER)
        outer.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)

        canvas = tk.Canvas(outer, bg=theme.WHITE, highlightthickness=0)
        scrollbar = ttk.Scrollbar(outer, orient=tk.VERTICAL, command=canvas.yview)
        form = tk.Frame(canvas, bg=theme.WHITE, padx=18, pady=14)

        form.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=form, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        fields = {}

        def add_section(title):
            tk.Label(
                form, text=title, font=theme.FONT_H2,
                bg=theme.WHITE, fg=theme.NAVY
            ).pack(fill=tk.X, anchor="w", pady=(8, 5))

        def add_field(key, label, value="", state="normal"):
            row_frame = tk.Frame(form, bg=theme.WHITE)
            row_frame.pack(fill=tk.X, pady=4)
            tk.Label(
                row_frame, text=label, width=24, anchor="w",
                font=theme.FONT_SMALL, bg=theme.WHITE, fg=theme.TEXT_MUTED
            ).pack(side=tk.LEFT)
            ent = tk.Entry(row_frame, font=theme.FONT_BODY)
            ent.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=3)
            ent.insert(0, "" if value is None else str(value))
            if state != "normal":
                ent.config(state=state)
            fields[key] = ent
            return ent

        add_section("Basic Student Information")
        add_field("student_id", "Student ID", base["student_id"], "disabled")
        add_field("name", "Full Name *", base["name"])
        add_field("father_name", "Father's Name", base["father_name"])
        add_field("dob", "Date of Birth (YYYY-MM-DD)", base["dob"])
        add_field("phone", "Phone", base["phone"])
        add_field("address", "Permanent Address", base["address"])
        add_field("class_sec", "Class / Section *", base["class_sec"])
        add_field("prev_education", "Previous Education", base["prev_education"])

        photo_row = tk.Frame(form, bg=theme.WHITE)
        photo_row.pack(fill=tk.X, pady=4)
        tk.Label(
            photo_row, text="Photo", width=24, anchor="w",
            font=theme.FONT_SMALL, bg=theme.WHITE, fg=theme.TEXT_MUTED
        ).pack(side=tk.LEFT)
        photo_var = tk.StringVar(value=base.get("photo_path") or "")
        photo_entry = tk.Entry(photo_row, textvariable=photo_var, font=theme.FONT_BODY)
        photo_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=3)

        def choose_photo():
            path = filedialog.askopenfilename(
                filetypes=[("Image Files", "*.jpg *.png *.jpeg")]
            )
            if path:
                photo_var.set(path)

        tk.Button(
            photo_row, text="Browse", command=choose_photo,
            bg=theme.SLATE, fg="white", bd=0, padx=10, cursor="hand2"
        ).pack(side=tk.LEFT, padx=(8, 0))

        add_section("Fee Information")
        add_field("total_fee", "Total Fee", base["total_fee"], "normal" if can_fee else "disabled")
        add_field("paid_fee", "Paid Fee", base["paid_fee"], "normal" if can_fee else "disabled")

        add_section("Admission / Personal Information")
        for key, label in [
            ("gender", "Gender"),
            ("blood_group", "Blood Group"),
            ("nationality", "Nationality"),
            ("religion", "Religion"),
            ("mother_name", "Mother's Name"),
            ("guardian_name", "Guardian Name"),
            ("guardian_cnic", "Guardian CNIC"),
            ("occupation", "Guardian Occupation"),
            ("alt_phone", "Alternate Phone"),
            ("email", "Email"),
            ("current_address", "Current Address"),
            ("city", "City"),
            ("area", "Area"),
            ("admission_date", "Admission Date"),
            ("academic_year", "Academic Year"),
            ("admission_type", "Admission Type"),
            ("emergency_contact_name", "Emergency Contact Name"),
            ("emergency_contact_phone", "Emergency Contact Phone"),
            ("emergency_relationship", "Emergency Relationship"),
            ("emergency_notes", "Emergency Notes"),
        ]:
            add_field(key, label, extra_data.get(key, ""))

        add_field("status", "Student Status", base.get("status") or "Active")

        def do_save():
            name = fields["name"].get().strip()
            cls = fields["class_sec"].get().strip()

            if not name or not cls:
                messagebox.showerror(
                    "Required Fields",
                    "Student Name and Class / Section are required.",
                    parent=win
                )
                return

            if can_fee:
                total_f, ok1 = safe_float(
                    fields["total_fee"].get(), "Total Fee", default=0.0
                )
                paid_f, ok2 = safe_float(
                    fields["paid_fee"].get(), "Paid Fee", default=0.0
                )
                if not (ok1 and ok2):
                    return
                if paid_f > total_f:
                    if not messagebox.askyesno(
                        "Overpayment Warning",
                        f"Paid Fee (Rs. {paid_f:.2f}) is more than Total Fee "
                        f"(Rs. {total_f:.2f}). Continue?",
                        parent=win
                    ):
                        return
            else:
                total_f = base.get("total_fee") or 0.0
                paid_f = base.get("paid_fee") or 0.0

            # Update the main students table.
            db.run(
                """UPDATE students SET
                   name=?, father_name=?, dob=?, phone=?, address=?, class_sec=?,
                   photo_path=?, prev_education=?, total_fee=?, paid_fee=?, status=?
                   WHERE student_id=?""",
                (
                    name,
                    fields["father_name"].get().strip(),
                    fields["dob"].get().strip(),
                    fields["phone"].get().strip(),
                    fields["address"].get().strip(),
                    cls,
                    photo_var.get().strip(),
                    fields["prev_education"].get().strip(),
                    total_f,
                    paid_f,
                    fields["status"].get().strip() or "Active",
                    s_id
                ),
                commit=True
            )

            # Update the complete admission-extra record. If it does not exist,
            # create it so older students can also gain all profile fields.
            extra_values = tuple(
                fields[key].get().strip() for key in extra_keys
            )
            extra_exists = db.run(
                "SELECT 1 FROM student_admission_extra WHERE student_id=?",
                (s_id,), fetchone=True
            )

            if extra_exists:
                db.run(
                    """UPDATE student_admission_extra SET
                       gender=?, blood_group=?, nationality=?, religion=?, mother_name=?,
                       guardian_name=?, guardian_cnic=?, occupation=?, alt_phone=?, email=?,
                       current_address=?, city=?, area=?, admission_date=?, academic_year=?,
                       admission_type=?, emergency_contact_name=?, emergency_contact_phone=?,
                       emergency_relationship=?, emergency_notes=?
                       WHERE student_id=?""",
                    extra_values + (s_id,),
                    commit=True
                )
            else:
                db.run(
                    """INSERT INTO student_admission_extra (
                       student_id, gender, blood_group, nationality, religion, mother_name,
                       guardian_name, guardian_cnic, occupation, alt_phone, email,
                       current_address, city, area, admission_date, academic_year,
                       admission_type, emergency_contact_name, emergency_contact_phone,
                       emergency_relationship, emergency_notes
                       ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (s_id,) + extra_values,
                    commit=True
                )

            # Record only a newly increased paid amount; do not double-count old payments.
            delta = paid_f - previous_paid
            if delta > 0 and can_fee:
                accounting.record_fee_revenue(
                    self.user_role, s_id, delta, self.current_user
                )

            log_activity(
                self.current_user,
                f"Updated complete student profile for {s_id}"
            )

            self.load_student_table()
            messagebox.showinfo(
                "Updated",
                f"Complete student information for {s_id} has been updated successfully.",
                parent=win
            )

            if delta > 0 and can_fee:
                self._offer_fee_receipt(
                    s_id, name, fields["father_name"].get().strip(), cls,
                    total_f, previous_paid, delta
                )

            win.destroy()

        button_row = tk.Frame(win, bg=theme.SILVER, padx=12, pady=10)
        button_row.pack(fill=tk.X)
        theme.primary_button(
            button_row, "💾 Save All Changes", do_save, bg=theme.SUCCESS
        ).pack(side=tk.LEFT)
        tk.Button(
            button_row, text="Cancel", command=win.destroy,
            bg=theme.WHITE, fg=theme.TEXT_MUTED, bd=0,
            font=theme.FONT_SMALL, cursor="hand2"
        ).pack(side=tk.LEFT, padx=12)

        canvas.bind_all(
            "<MouseWheel>",
            lambda e: canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        )

        def close_window():
            try:
                canvas.unbind_all("<MouseWheel>")
            except Exception:
                pass
            win.destroy()

        win.protocol("WM_DELETE_WINDOW", close_window)


    def clear_admission_form(self):
        self.lbl_auto_id.config(text=generate_next_student_id())
        for attr in ['ent_name', 'ent_fname', 'ent_dob', 'ent_phone', 'ent_address', 'ent_class', 'ent_prev_edu', 'ent_total_fee', 'ent_paid_fee']:
            if hasattr(self, attr):
                w = getattr(self, attr)
                state = str(w['state'])
                w.config(state="normal")
                w.delete(0, tk.END)
                if state == "disabled":
                    w.config(state="disabled")
        self.photo_path_var.set("")
        self._prev_paid_fee = 0

    def generate_id_card(self):
        s_id = self.lbl_auto_id.cget("text")
        if not s_id:
            messagebox.showerror("Error", "Select/enter student details first!")
            return
        row = db.run("SELECT name, father_name, class_sec, phone, photo_path FROM students WHERE student_id=?",
                      (s_id,), fetchone=True)
        if not row:
            messagebox.showerror("Error", "Student not found. Save the student first.")
            return
        name, father_name, cls, phone, photo_path = row
        emer = ""
        try:
            er = db.run(
                "SELECT emergency_contact_phone FROM student_admission_extra WHERE student_id=?",
                (s_id,), fetchone=True,
            )
            if er and er[0]:
                emer = er[0]
        except Exception:
            pass
        out_path = os.path.join(os.getcwd(), f"ID_Card_{s_id}.pdf")
        reports.generate_id_card(
            s_id, name, cls, out_path, father_name=father_name, phone=phone,
            photo_path=photo_path, emergency_phone=emer,
        )
        log_activity(self.current_user, f"Generated digital ID Card for student {s_id}")
        barcode_note = "" if reports.HAS_BARCODE else "\n\nNote: install 'python-barcode' (see requirements.txt) to include a real scannable barcode — this card currently used the text-fallback for that field."
        messagebox.showinfo("Success", f"Digital ID Card (with QR + barcode) generated:\n{out_path}{barcode_note}")

    # ------------------------------------------
    # TAB 2: ATTENDANCE SCANNER
    # ------------------------------------------
    def build_attendance_tab(self):
        top = tk.LabelFrame(self.tab_attendance, text="RFID / Barcode / QR Quick Scanner Mode", font=("Segoe UI", 10, "bold"), padx=15, pady=15)
        top.pack(fill=tk.X, padx=15, pady=15)

        tk.Label(top, text="Scan ID Card / Enter ID:", font=("Segoe UI", 11, "bold")).pack(side=tk.LEFT, padx=10)
        self.ent_scan_id = tk.Entry(top, font=("Segoe UI", 12, "bold"), width=20, bg="#fef08a")
        self.ent_scan_id.pack(side=tk.LEFT, padx=10, ipady=4)
        self.ent_scan_id.focus_set()
        self.ent_scan_id.bind("<Return>", self.process_scanned_attendance)

        tk.Label(top, text="Status:").pack(side=tk.LEFT, padx=(20, 5))
        self.combo_manual_status = ttk.Combobox(top, values=["Present", "Absent", "Late", "Leave"], width=10, state="readonly")
        self.combo_manual_status.current(0)
        self.combo_manual_status.pack(side=tk.LEFT, padx=5)

        tk.Button(top, text="Mark Attendance", command=self.process_scanned_attendance, bg="#16a34a", fg="white", font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT, padx=10)
        tk.Button(top, text="📅 Monthly Attendance Report", command=self.open_monthly_attendance_report,
                  bg="#7c3aed", fg="white", font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=10)

        self.lbl_att_status = tk.Label(self.tab_attendance, text="Ready for Card / QR Scan...", font=("Segoe UI", 14, "bold"), fg="#64748b")
        self.lbl_att_status.pack(pady=10)

        self.tree_att = ttk.Treeview(self.tab_attendance, columns=("id", "student_id", "name", "date", "status", "method"), show="headings")
        for col, h in [("id", "Log ID"), ("student_id", "Student ID"), ("name", "Name"), ("date", "Date"), ("status", "Status"), ("method", "Method")]:
            self.tree_att.heading(col, text=h)
            self.tree_att.column(col, anchor="center")

        self.tree_att.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)
        self.load_attendance_table()

    def open_monthly_attendance_report(self):
        win = tk.Toplevel(self.root)
        win.title("Monthly Attendance Report")
        win.geometry("420x260")
        win.config(bg=theme.WHITE)

        tk.Label(win, text="Monthly Attendance Report", font=theme.FONT_H1, bg=theme.WHITE).pack(pady=(14, 10))

        form = tk.Frame(win, bg=theme.WHITE)
        form.pack(pady=4)
        tk.Label(form, text="Student ID:", bg=theme.WHITE).grid(row=0, column=0, sticky="e", padx=6, pady=6)
        ent_sid = tk.Entry(form, width=18)
        ent_sid.grid(row=0, column=1, pady=6)

        tk.Label(form, text="Month:", bg=theme.WHITE).grid(row=1, column=0, sticky="e", padx=6, pady=6)
        combo_month = ttk.Combobox(form, values=[f"{i:02d}" for i in range(1, 13)], width=6, state="readonly")
        combo_month.current(datetime.now().month - 1)
        combo_month.grid(row=1, column=1, sticky="w", pady=6)

        tk.Label(form, text="Year:", bg=theme.WHITE).grid(row=2, column=0, sticky="e", padx=6, pady=6)
        ent_year = tk.Entry(form, width=8)
        ent_year.insert(0, str(datetime.now().year))
        ent_year.grid(row=2, column=1, sticky="w", pady=6)

        result_lbl = tk.Label(win, text="", font=theme.FONT_SMALL, bg=theme.WHITE, justify="left")
        result_lbl.pack(pady=(6, 6))

        state = {"last": None}

        def compute():
            s_id = ent_sid.get().strip()
            month = combo_month.get()
            year = ent_year.get().strip()
            if not (s_id and month and year.isdigit() and len(year) == 4):
                messagebox.showerror("Error", "Enter a valid Student ID, Month, and 4-digit Year.")
                return None
            student = db.run("SELECT name, class_sec FROM students WHERE student_id=?", (s_id,), fetchone=True)
            if not student:
                messagebox.showerror("Error", f"Student ID '{s_id}' not found.")
                return None
            name, cls = student
            ym = f"{year}-{month}"
            month_label = ym

            # "Total working days" = distinct calendar days the SCHOOL had
            # any attendance activity that month (there's no separate
            # school-calendar/holiday table in this system) — documented
            # in the report itself as "school-wide" so it isn't mistaken
            # for an official calendar.
            total_working_days = db.run(
                "SELECT COUNT(DISTINCT date) FROM attendance WHERE date LIKE ?", (f"{ym}%",), fetchone=True)[0]

            day_rows = db.run(
                "SELECT date, status FROM attendance WHERE student_id=? AND date LIKE ? ORDER BY date",
                (s_id, f"{ym}%"), fetchall=True)
            present = sum(1 for _, s in day_rows if s == "Present")
            absent = sum(1 for _, s in day_rows if s == "Absent")
            leave = sum(1 for _, s in day_rows if s == "Leave")
            late = sum(1 for _, s in day_rows if s == "Late")
            pct = (present / total_working_days * 100) if total_working_days else 0.0

            result_lbl.config(text=(f"{name} ({s_id}), Class {cls} — {month_label}\n"
                                     f"Working days: {total_working_days}  |  Present: {present}  |  "
                                     f"Absent: {absent}  |  Leave: {leave}  |  Late: {late}\n"
                                     f"Attendance: {pct:.1f}%"))
            state["last"] = (s_id, name, cls, month_label, total_working_days, present, absent, leave, late, pct, day_rows)
            return state["last"]

        def export_pdf():
            data = state["last"] or compute()
            if not data:
                return
            s_id, name, cls, month_label, total_working_days, present, absent, leave, late, pct, day_rows = data
            out_path = os.path.join(os.getcwd(), f"Attendance_Report_{s_id}_{month_label}.pdf")
            reports.generate_attendance_report(s_id, name, cls, month_label, total_working_days, present, absent,
                                                leave, late, pct, day_rows, out_path)
            log_activity(self.current_user, f"Generated monthly attendance report for {s_id} ({month_label})")
            messagebox.showinfo("Report Generated", f"Saved to:\n{out_path}")

        btn_row = tk.Frame(win, bg=theme.WHITE)
        btn_row.pack(pady=8)
        theme.primary_button(btn_row, "Show Summary", compute).pack(side=tk.LEFT, padx=6)
        theme.primary_button(btn_row, "📄 Export PDF", export_pdf, bg=theme.SLATE).pack(side=tk.LEFT, padx=6)

    def process_scanned_attendance(self, ev=None):
        s_id = self.ent_scan_id.get().strip()
        if not s_id:
            return
        status = self.combo_manual_status.get() or "Present"
        today = datetime.now().strftime("%Y-%m-%d")
        now_time = datetime.now().strftime("%I:%M %p")

        row = db.run("SELECT name, phone FROM students WHERE student_id=?", (s_id,), fetchone=True)
        if not row:
            self.lbl_att_status.config(text=f"❌ Student ID '{s_id}' Not Found!", fg="#dc2626")
            self.root.bell()
        else:
            student_name, phone = row
            try:
                db.run("INSERT INTO attendance (student_id, date, status, method) VALUES (?, ?, ?, ?)",
                       (s_id, today, status, "Scan/Manual"), commit=True)
                self.lbl_att_status.config(text=f"✅ Attendance Marked ({status}): {student_name} ({s_id})\n📱 WhatsApp Alert Sent to Parent!", fg="#16a34a")
                print(f"[WHATSAPP ALERT SIMULATION]: Sent to {phone} -> Dear Parent, {student_name} marked '{status}' at {now_time}.")
            except Exception:
                self.lbl_att_status.config(text=f"⚠️ Already Marked Today: {student_name}", fg="#eab308")
                self.root.bell()

        self.ent_scan_id.delete(0, tk.END)
        self.load_attendance_table()

    def load_attendance_table(self):
        self.tree_att.delete(*self.tree_att.get_children())
        rows = db.run("""SELECT attendance.id, attendance.student_id, students.name, attendance.date,
                                 attendance.status, attendance.method
                          FROM attendance JOIN students ON attendance.student_id = students.student_id
                          ORDER BY attendance.id DESC LIMIT 500""", fetchall=True)
        for r in rows:
            self.tree_att.insert("", tk.END, values=r)

    # ------------------------------------------
    # TAB 3: RESULT & PERFORMANCE
    # ------------------------------------------
    def build_results_tab(self):
        """Embed the professional Results & Academics workspace from results_window.py."""
        for child in self.tab_results.winfo_children():
            child.destroy()
        self._results_workspace = build_results_into(
            self.tab_results, self.user_role, self.current_user
        )

    def open_results_window(self):
        """Optional: open Results as a dedicated popup window."""
        return launch_results_window(self.root, self.user_role, self.current_user)

    # Legacy method names retained so any external/quick-action callers
    # that still reference them do not break. They delegate to the
    # embedded Results workspace when available.

    def load_results_table(self):
        ws = getattr(self, "_results_workspace", None)
        if ws is not None:
            ws.load_history_table()

    def save_marks(self):
        ws = getattr(self, "_results_workspace", None)
        if ws is not None:
            ws.save_marks()

    def generate_marksheet_pdf(self):
        ws = getattr(self, "_results_workspace", None)
        if ws is not None:
            ws.generate_marksheet()

    def show_performance_graph(self):
        ws = getattr(self, "_results_workspace", None)
        if ws is not None:
            ws.show_performance_graph()

    def build_teachers_tab(self):
        can_add_teacher = rbac.can(self.user_role, "teacher.add")
        can_mark_tch_att = rbac.can(self.user_role, "teacher.attendance.mark")
        can_view_salary = rbac.can(self.user_role, "teacher.salary.view")
        can_pay_salary = rbac.can(self.user_role, "teacher.salary.pay")

        left = tk.LabelFrame(self.tab_teachers, text="Teacher Registration & Payroll Setup", font=("Segoe UI", 10, "bold"), padx=10, pady=10)
        left.place(x=10, y=10, width=380, height=600)

        tk.Label(left, text="Teacher ID:", font=("Segoe UI", 9, "bold")).grid(row=0, column=0, sticky="w", pady=5)
        self.lbl_tch_id = tk.Label(left, text=generate_next_teacher_id(), font=("Segoe UI", 10, "bold"), fg="#0284c7")
        self.lbl_tch_id.grid(row=0, column=1, sticky="w", pady=5)

        fields = [("Teacher Name*:", "ent_tch_name"), ("Designation:", "ent_tch_desig"),
                  ("Phone Number:", "ent_tch_phone"), ("Basic Salary (Rs):", "ent_tch_sal"),
                  ("Joining Date:", "ent_tch_join")]
        for i, (label, var_name) in enumerate(fields, start=1):
            tk.Label(left, text=label, font=("Segoe UI", 9)).grid(row=i, column=0, sticky="w", pady=5)
            ent = tk.Entry(left, font=("Segoe UI", 9))
            ent.grid(row=i, column=1, pady=5, ipady=2, sticky="we")
            setattr(self, var_name, ent)
        if not can_view_salary:
            self.ent_tch_sal.config(show="•")

        btn_frame = tk.Frame(left)
        btn_frame.place(x=5, y=220, width=350, height=120)
        btn_add = tk.Button(btn_frame, text="Add Teacher", command=self.save_teacher, bg="#16a34a", fg="white", font=("Segoe UI", 9, "bold"), bd=0)
        btn_add.place(x=5, y=5, width=160, height=35)
        btn_add.config(state="normal" if can_add_teacher else "disabled")
        tk.Button(btn_frame, text="Clear Form", command=self.clear_teacher_form, bg="#64748b", fg="white", font=("Segoe UI", 9, "bold"), bd=0).place(x=175, y=5, width=160, height=35)

        att_box = tk.LabelFrame(left, text="Daily Teacher Attendance", font=("Segoe UI", 9, "bold"), padx=5, pady=5)
        att_box.place(x=5, y=340, width=350, height=120)
        tk.Label(att_box, text="Status:").grid(row=0, column=0, padx=5)
        self.combo_tch_att = ttk.Combobox(att_box, values=["Present", "Absent", "Late", "Leave"], width=10, state="readonly")
        self.combo_tch_att.current(0)
        self.combo_tch_att.grid(row=0, column=1, padx=5)
        btn_mark = tk.Button(att_box, text="Mark Today Attendance", command=self.mark_teacher_attendance, bg="#0284c7", fg="white", font=("Segoe UI", 9, "bold"))
        btn_mark.grid(row=0, column=2, padx=5)
        btn_mark.config(state="normal" if can_mark_tch_att else "disabled")

        if can_view_salary:
            sal_box = tk.LabelFrame(left, text="Salary & Payslip Generation", font=("Segoe UI", 9, "bold"), padx=5, pady=5)
            sal_box.place(x=5, y=470, width=350, height=80)
            btn_pay = tk.Button(sal_box, text="💵 Calculate Salary & Export Payslip PDF", command=self.generate_salary_payslip, bg="#7c3aed", fg="white", font=("Segoe UI", 10, "bold"), bd=0)
            btn_pay.pack(fill=tk.X, pady=10, ipady=5)
            btn_pay.config(state="normal" if can_pay_salary else "disabled")

        right = tk.Frame(self.tab_teachers)
        right.place(x=400, y=10, width=800, height=600)
        tk.Label(right, text="TEACHERS DIRECTORY & PAYROLL", font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(0, 5))

        cols = ("id", "name", "desig", "phone", "salary", "joining") if can_view_salary else ("id", "name", "desig", "phone", "joining")
        self.tree_teacher = ttk.Treeview(right, columns=cols, show="headings")
        headers = {"id": "Teacher ID", "name": "Name", "desig": "Designation", "phone": "Phone", "salary": "Basic Salary", "joining": "Joining Date"}
        for col in cols:
            self.tree_teacher.heading(col, text=headers[col])
            self.tree_teacher.column(col, anchor="center")

        self.tree_teacher.pack(fill=tk.BOTH, expand=True)
        self.tree_teacher.bind("<ButtonRelease-1>", self.fill_teacher_form)
        self.load_teacher_table()

    def save_teacher(self):
        if not rbac.can(self.user_role, "teacher.add"):
            messagebox.showerror("Permission Denied", "Not allowed to add teachers.")
            return
        t_id = self.lbl_tch_id.cget("text")
        name = self.ent_tch_name.get().strip()
        if not name:
            messagebox.showerror("Error", "Enter Teacher Name!")
            return
        sal, ok = safe_float(self.ent_tch_sal.get(), "Basic Salary", default=0.0)
        if not ok:
            return
        db.run("INSERT INTO teachers (teacher_id, name, designation, phone, basic_salary, joining_date) VALUES (?, ?, ?, ?, ?, ?)",
               (t_id, name, self.ent_tch_desig.get(), self.ent_tch_phone.get(), sal, self.ent_tch_join.get()), commit=True)
        log_activity(self.current_user, f"Added teacher profile {name} ({t_id})")
        messagebox.showinfo("Success", f"Teacher Profile Registered: {t_id}")
        self.clear_teacher_form()
        self.load_teacher_table()

    def load_teacher_table(self):
        self.tree_teacher.delete(*self.tree_teacher.get_children())
        can_view_salary = rbac.can(self.user_role, "teacher.salary.view")
        if can_view_salary:
            rows = db.run("SELECT teacher_id, name, designation, phone, basic_salary, joining_date FROM teachers", fetchall=True)
        else:
            rows = db.run("SELECT teacher_id, name, designation, phone, joining_date FROM teachers", fetchall=True)
        for r in rows:
            self.tree_teacher.insert("", tk.END, values=r)

    def fill_teacher_form(self, ev):
        selected = self.tree_teacher.focus()
        vals = self.tree_teacher.item(selected, "values")
        if not vals:
            return
        self.lbl_tch_id.config(text=vals[0])
        self.ent_tch_name.delete(0, tk.END); self.ent_tch_name.insert(0, vals[1])
        self.ent_tch_desig.delete(0, tk.END); self.ent_tch_desig.insert(0, vals[2] or "")
        self.ent_tch_phone.delete(0, tk.END); self.ent_tch_phone.insert(0, vals[3] or "")
        if rbac.can(self.user_role, "teacher.salary.view") and len(vals) > 4:
            self.ent_tch_sal.delete(0, tk.END); self.ent_tch_sal.insert(0, str(vals[4]))
            self.ent_tch_join.delete(0, tk.END); self.ent_tch_join.insert(0, vals[5] or "")

    def clear_teacher_form(self):
        self.lbl_tch_id.config(text=generate_next_teacher_id())
        for ent in [self.ent_tch_name, self.ent_tch_desig, self.ent_tch_phone, self.ent_tch_sal, self.ent_tch_join]:
            ent.delete(0, tk.END)

    def mark_teacher_attendance(self):
        if not rbac.can(self.user_role, "teacher.attendance.mark"):
            messagebox.showerror("Permission Denied", "Not allowed to mark teacher attendance.")
            return
        t_id = self.lbl_tch_id.cget("text")
        status = self.combo_tch_att.get()
        today = datetime.now().strftime("%Y-%m-%d")
        now_time = datetime.now().strftime("%H:%M:%S")
        try:
            db.run("INSERT INTO teacher_attendance (teacher_id, date, status, in_time) VALUES (?, ?, ?, ?)",
                   (t_id, today, status, now_time), commit=True)
            messagebox.showinfo("Success", f"Teacher Attendance ({status}) Marked for {t_id}")
        except Exception:
            messagebox.showwarning("Warning", "Today's Attendance already marked for this teacher!")

    def generate_salary_payslip(self):
        if not rbac.can(self.user_role, "teacher.salary.pay"):
            messagebox.showerror("Permission Denied", "Not allowed to process salaries.")
            return
        t_id = self.lbl_tch_id.cget("text")
        name = self.ent_tch_name.get().strip()
        basic_sal, ok = safe_float(self.ent_tch_sal.get(), "Basic Salary", default=0.0)
        if not ok:
            return
        if not name or basic_sal <= 0:
            messagebox.showerror("Error", "Select a teacher and verify Basic Salary first!")
            return

        today_month = datetime.now().strftime("%Y-%m")

        # Idempotency guard: a "Salary" expense for this teacher this month
        # already means a payslip was issued — don't silently double-pay.
        already_paid = db.run(
            "SELECT id FROM accounting_expense WHERE category='Salary' AND date LIKE ? AND vendor_or_person LIKE ?",
            (f"{today_month}%", f"%({t_id})%"), fetchone=True,
        )
        if already_paid:
            if not messagebox.askyesno(
                "Salary Already Recorded",
                f"A salary payment for {name} ({t_id}) was already recorded for {today_month}.\n"
                "Generating another payslip will record a SECOND expense entry for this teacher this month.\n"
                "Continue anyway?"):
                return

        absents = db.run("SELECT COUNT(*) FROM teacher_attendance WHERE teacher_id=? AND status='Absent' AND date LIKE ?",
                          (t_id, f"{today_month}%"), fetchone=True)[0]

        per_day = basic_sal / 30.0
        deductions = per_day * absents
        net_sal = basic_sal - deductions

        out_path = os.path.join(os.getcwd(), f"Payslip_{t_id}_{today_month}.pdf")
        reports.generate_payslip(t_id, name, today_month, basic_sal, absents, deductions, net_sal, out_path)

        accounting.record_salary_expense(self.user_role, t_id, name, net_sal, today_month, self.current_user, reference=out_path)

        log_activity(self.current_user, f"Generated salary payslip for teacher {t_id}")
        messagebox.showinfo("Success", f"Payslip PDF Generated:\n{out_path}\nNet Payable: Rs. {net_sal:.2f}\n(Recorded as an accounting expense.)")

    # ------------------------------------------
    # TAB 5: CLASS TIMETABLE
    # ------------------------------------------
    def build_timetable_tab(self):
        can_manage = rbac.can(self.user_role, "timetable.manage")
        top = tk.LabelFrame(self.tab_timetable, text="Add Period Schedule", font=("Segoe UI", 10, "bold"), padx=10, pady=10)
        top.pack(fill=tk.X, padx=10, pady=10)

        tk.Label(top, text="Class:").grid(row=0, column=0, padx=5)
        self.ent_tt_cls = tk.Entry(top, width=10)
        self.ent_tt_cls.grid(row=0, column=1, padx=5)
        tk.Label(top, text="Day:").grid(row=0, column=2, padx=5)
        self.combo_tt_day = ttk.Combobox(top, values=["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"], width=10, state="readonly")
        self.combo_tt_day.current(0)
        self.combo_tt_day.grid(row=0, column=3, padx=5)
        tk.Label(top, text="Time Slot:").grid(row=0, column=4, padx=5)
        self.ent_tt_time = tk.Entry(top, width=12)
        self.ent_tt_time.insert(0, "09:00 - 09:45")
        self.ent_tt_time.grid(row=0, column=5, padx=5)
        tk.Label(top, text="Subject:").grid(row=0, column=6, padx=5)
        self.ent_tt_sub = tk.Entry(top, width=12)
        self.ent_tt_sub.grid(row=0, column=7, padx=5)
        tk.Label(top, text="Teacher:").grid(row=0, column=8, padx=5)
        self.ent_tt_tch = tk.Entry(top, width=12)
        self.ent_tt_tch.grid(row=0, column=9, padx=5)

        btn = tk.Button(top, text="Assign Period", command=self.save_timetable, bg="#16a34a", fg="white", font=("Segoe UI", 9, "bold"))
        btn.grid(row=0, column=10, padx=10)
        btn.config(state="normal" if can_manage else "disabled")

        self.tree_tt = ttk.Treeview(self.tab_timetable, columns=("id", "class", "day", "time", "subject", "teacher"), show="headings")
        for col, h in [("id", "ID"), ("class", "Class"), ("day", "Day"), ("time", "Time Slot"), ("subject", "Subject"), ("teacher", "Teacher")]:
            self.tree_tt.heading(col, text=h)
            self.tree_tt.column(col, anchor="center")
        self.tree_tt.pack(fill=tk.BOTH, expand=True, padx=10, pady=(10, 0))

        remove_row = tk.Frame(self.tab_timetable)
        remove_row.pack(fill=tk.X, padx=10, pady=10)
        btn_remove = tk.Button(remove_row, text="Remove Selected Assignment", command=self.remove_timetable_entry,
                                bg="#dc2626", fg="white", font=("Segoe UI", 9, "bold"))
        btn_remove.pack(side=tk.LEFT)
        btn_remove.config(state="normal" if can_manage else "disabled")

        self.load_timetable_table()

    def remove_timetable_entry(self):
        if not rbac.can(self.user_role, "timetable.manage"):
            messagebox.showerror("Permission Denied", "Not allowed to manage timetable.")
            return
        sel = self.tree_tt.selection()
        if not sel:
            messagebox.showerror("Error", "Select a timetable row first.")
            return
        vals = self.tree_tt.item(sel[0], "values")
        tt_id = vals[0]
        if messagebox.askyesno("Confirm", f"Remove this timetable assignment (Class {vals[1]}, {vals[2]} {vals[3]}, {vals[4]} — {vals[5]})?"):
            db.run("DELETE FROM timetable WHERE id=?", (tt_id,), commit=True)
            log_activity(self.current_user, f"Removed timetable assignment id={tt_id}")
            self.load_timetable_table()

    def save_timetable(self):
        if not rbac.can(self.user_role, "timetable.manage"):
            messagebox.showerror("Permission Denied", "Not allowed to manage timetable.")
            return
        cls = self.ent_tt_cls.get().strip()
        day = self.combo_tt_day.get()
        time_slot = self.ent_tt_time.get().strip()
        sub = self.ent_tt_sub.get().strip()
        tch = self.ent_tt_tch.get().strip()
        if not (cls and sub):
            messagebox.showerror("Error", "Enter at least a Class and a Subject.")
            return

        # Conflict checks — warn rather than silently allow, per spec.
        warnings = []
        same_class_slot = db.run(
            "SELECT subject_name, teacher_name FROM timetable WHERE class_name=? AND day_name=? AND time_slot=?",
            (cls, day, time_slot), fetchone=True)
        if same_class_slot:
            warnings.append(f"Class {cls} already has '{same_class_slot[0]}' with {same_class_slot[1] or '(no teacher set)'} "
                             f"scheduled on {day} at {time_slot}.")
        if tch:
            teacher_elsewhere = db.run(
                "SELECT class_name, subject_name FROM timetable WHERE teacher_name=? AND day_name=? AND time_slot=? AND class_name!=?",
                (tch, day, time_slot, cls), fetchone=True)
            if teacher_elsewhere:
                warnings.append(f"Teacher '{tch}' is already assigned to Class {teacher_elsewhere[0]} "
                                 f"('{teacher_elsewhere[1]}') on {day} at {time_slot}.")

        if warnings:
            if not messagebox.askyesno(
                "Scheduling Conflict",
                "The following conflict(s) were found:\n\n" + "\n".join(f"• {w}" for w in warnings) +
                "\n\nSave this assignment anyway?"):
                return

        db.run("INSERT INTO timetable (class_name, day_name, time_slot, subject_name, teacher_name) VALUES (?, ?, ?, ?, ?)",
               (cls, day, time_slot, sub, tch), commit=True)
        log_activity(self.current_user, f"Added timetable assignment: Class {cls}, {day} {time_slot}, {sub}, {tch}")
        self.load_timetable_table()
        messagebox.showinfo("Success", "Timetable Schedule Saved!")

    def load_timetable_table(self):
        self.tree_tt.delete(*self.tree_tt.get_children())
        rows = db.run("SELECT * FROM timetable", fetchall=True)
        for r in rows:
            self.tree_tt.insert("", tk.END, values=r)

    # ------------------------------------------
    # TAB 6: ACCOUNTING (NEW)
    # ------------------------------------------
    def build_accounting_tab(self):
        dash = tk.LabelFrame(self.tab_accounting, text="Financial Dashboard", font=("Segoe UI", 10, "bold"), padx=10, pady=10)
        dash.pack(fill=tk.X, padx=10, pady=10)

        self.lbl_dash = tk.Label(dash, text="", font=("Segoe UI", 11, "bold"), justify="left")
        self.lbl_dash.pack(side=tk.LEFT, padx=10)

        btns = tk.Frame(dash)
        btns.pack(side=tk.RIGHT)
        tk.Button(btns, text="🔄 Refresh", command=self.refresh_accounting_dashboard, bg="#0284c7", fg="white", font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=5)
        if HAS_MATPLOTLIB:
            tk.Button(btns, text="📊 Revenue vs Expense Chart", command=self.show_finance_chart, bg="#7c3aed", fg="white", font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=5)

        entry_frame = tk.LabelFrame(self.tab_accounting, text="Record Revenue / Expense", font=("Segoe UI", 10, "bold"), padx=10, pady=10)
        entry_frame.pack(fill=tk.X, padx=10, pady=5)

        tk.Label(entry_frame, text="Type:").grid(row=0, column=0, padx=5)
        self.combo_acc_type = ttk.Combobox(entry_frame, values=["Revenue", "Expense"], width=10, state="readonly")
        self.combo_acc_type.current(0)
        self.combo_acc_type.grid(row=0, column=1, padx=5)

        tk.Label(entry_frame, text="Category:").grid(row=0, column=2, padx=5)
        self.ent_acc_category = tk.Entry(entry_frame, width=15)
        self.ent_acc_category.insert(0, "Other")
        self.ent_acc_category.grid(row=0, column=3, padx=5)

        tk.Label(entry_frame, text="Amount (Rs):").grid(row=0, column=4, padx=5)
        self.ent_acc_amount = tk.Entry(entry_frame, width=10)
        self.ent_acc_amount.grid(row=0, column=5, padx=5)

        tk.Label(entry_frame, text="Description:").grid(row=0, column=6, padx=5)
        self.ent_acc_desc = tk.Entry(entry_frame, width=20)
        self.ent_acc_desc.grid(row=0, column=7, padx=5)

        tk.Button(entry_frame, text="Save Entry", command=self.save_accounting_entry, bg="#16a34a", fg="white", font=("Segoe UI", 9, "bold")).grid(row=0, column=8, padx=10)

        self.tree_accounting = ttk.Treeview(self.tab_accounting, columns=("id", "type", "category", "amount", "date", "desc"), show="headings")
        for col, h in [("id", "ID"), ("type", "Type"), ("category", "Category/Source"), ("amount", "Amount"), ("date", "Date"), ("desc", "Description")]:
            self.tree_accounting.heading(col, text=h)
            self.tree_accounting.column(col, anchor="center")
        self.tree_accounting.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.refresh_accounting_dashboard()
        self.load_accounting_table()

    def refresh_accounting_dashboard(self):
        totals = accounting.dashboard_totals(self.user_role)
        self.lbl_dash.config(text=(
            f"Total Revenue: Rs. {totals['total_revenue']:.2f}     "
            f"Total Expenses: Rs. {totals['total_expense']:.2f}     "
            f"Net Income: Rs. {totals['net_income']:.2f}\n"
            f"Today — Revenue: Rs. {totals['today_revenue']:.2f}  Expenses: Rs. {totals['today_expense']:.2f}     "
            f"This Month — Revenue: Rs. {totals['month_revenue']:.2f}  Expenses: Rs. {totals['month_expense']:.2f}"
        ))

    def save_accounting_entry(self):
        kind = self.combo_acc_type.get()
        category = self.ent_acc_category.get().strip() or "Other"
        try:
            amount = float(self.ent_acc_amount.get().strip())
        except ValueError:
            messagebox.showerror("Error", "Enter a valid amount.")
            return
        desc = self.ent_acc_desc.get().strip()

        if kind == "Revenue":
            accounting.add_revenue(self.user_role, category, amount, desc, "", "Cash", self.current_user)
        else:
            accounting.add_expense(self.user_role, category, amount, desc, "", "Cash", "", self.current_user)

        log_activity(self.current_user, f"Recorded {kind} entry: {category} Rs.{amount}")
        self.ent_acc_amount.delete(0, tk.END)
        self.ent_acc_desc.delete(0, tk.END)
        self.refresh_accounting_dashboard()
        self.load_accounting_table()

    def load_accounting_table(self):
        self.tree_accounting.delete(*self.tree_accounting.get_children())
        for r in accounting.list_revenue(self.user_role):
            r_id, source, student_id, amount, date, desc, method = r
            self.tree_accounting.insert("", tk.END, values=(r_id, "Revenue", source, f"{amount:.2f}", date, desc or ""))
        for r in accounting.list_expense(self.user_role):
            e_id, category, amount, date, desc, vendor, method = r
            self.tree_accounting.insert("", tk.END, values=(e_id, "Expense", category, f"{amount:.2f}", date, desc or ""))

    def show_finance_chart(self):
        totals = accounting.dashboard_totals(self.user_role)
        plt.figure(figsize=(5, 4))
        plt.bar(["Revenue", "Expenses", "Net Income"],
                [totals["total_revenue"], totals["total_expense"], totals["net_income"]],
                color=["#16a34a", "#dc2626", "#0284c7"])
        plt.ylabel("Rs.")
        plt.title("Revenue vs Expenses vs Net Income")
        plt.tight_layout()
        plt.show()

    # ------------------------------------------
    # TAB 7: SETTINGS — System Settings & Automation
    # ------------------------------------------
    def build_settings_tab(self):
        """Render Settings into self.tab_settings via the dedicated module.

        Appearance / Dark Mode UI has been removed; the app stays on the
        default Light theme. Fee-automation and student-directory refresh
        are passed as callbacks so the module stays decoupled from App.
        """
        build_settings_panel(
            self.tab_settings,
            self.user_role,
            self.current_user,
            log_activity=log_activity,
            run_fee_automation=lambda force=False: self._run_fee_automation_safely(force=force),
            refresh_students=lambda: getattr(self, "load_student_table", lambda: None)(),
        )

    # ------------------------------------------
    # TAB 8: AUDIT LOGS (ADMIN ONLY)
    # ------------------------------------------
    def build_logs_tab(self):
        tk.Label(self.tab_logs, text="SYSTEM SECURITY & AUDIT TRAIL LOGS", font=("Segoe UI", 11, "bold")).pack(anchor="w", padx=10, pady=5)
        self.tree_logs = ttk.Treeview(self.tab_logs, columns=("id", "user", "action", "timestamp"), show="headings")
        for col, h in [("id", "Log ID"), ("user", "Username"), ("action", "Action Performed"), ("timestamp", "Timestamp")]:
            self.tree_logs.heading(col, text=h)
            self.tree_logs.column(col, anchor="center")
        self.tree_logs.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.load_audit_logs()

    def load_audit_logs(self):
        self.tree_logs.delete(*self.tree_logs.get_children())
        for r in db.run("SELECT * FROM audit_logs ORDER BY id DESC LIMIT 1000", fetchall=True):
            self.tree_logs.insert("", tk.END, values=r)

    # ------------------------------------------
    # BACKUP
    # ------------------------------------------
    def make_backup(self):
        dest_folder = filedialog.askdirectory(title="Select USB or Backup Folder")
        if not dest_folder:
            return
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(dest_folder, f"school_backup_{timestamp}.db")
        # Close the connection first so WAL-mode changes are checkpointed
        # into the main file before we copy it — otherwise a raw file copy
        # could grab an inconsistent snapshot while writes are still
        # sitting in the separate -wal file.
        db.close_conn()
        try:
            shutil.copy(db.DB_PATH, backup_path)
        finally:
            db.get_conn()  # reconnect immediately so the rest of the app keeps working
        log_activity(self.current_user, f"Created system backup at {backup_path}")
        messagebox.showinfo("Backup Complete", f"Backup Saved Successfully:\n{backup_path}")

    def restore_backup(self):
        if not rbac.can(self.user_role, "backup.run"):
            messagebox.showerror("Permission Denied", "You are not allowed to restore backups.")
            return
        src_path = filedialog.askopenfilename(title="Select Backup File to Restore", filetypes=[("SQLite Database", "*.db")])
        if not src_path:
            return

        try:
            with open(src_path, "rb") as f:
                header = f.read(16)
        except Exception as e:
            messagebox.showerror("Error", f"Could not read the selected file:\n{e}")
            return
        if header[:16] != b"SQLite format 3\x00":
            messagebox.showerror("Invalid File", "That file doesn't look like a valid SQLite database backup — restore cancelled, nothing was changed.")
            return

        if not messagebox.askyesno(
            "Confirm Restore",
            "Restoring will REPLACE the current live database with the selected backup.\n\n"
            "A safety copy of the CURRENT database will be made automatically first, so this "
            "can be undone by restoring that safety copy again if something goes wrong.\n\n"
            "You will need to restart the application after this completes. Continue?"):
            return

        safety_path = None
        db.close_conn()
        try:
            safety_dir = os.path.join(os.path.dirname(db.DB_PATH), "pre_restore_safety_backups")
            os.makedirs(safety_dir, exist_ok=True)
            safety_path = os.path.join(safety_dir, f"pre_restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db")
            shutil.copy(db.DB_PATH, safety_path)   # never overwrite the live DB before this safety copy exists

            shutil.copy(src_path, db.DB_PATH)
            # Any stale WAL/SHM sidecar files next to the OLD live DB are
            # meaningless against the newly-restored file — remove them so
            # SQLite doesn't try to replay an unrelated WAL against it.
            for ext in ("-wal", "-shm"):
                side = db.DB_PATH + ext
                if os.path.exists(side):
                    os.remove(side)
        except Exception as e:
            messagebox.showerror("Restore Failed", f"Could not restore the backup:\n{e}\n\n"
                                  f"{'A safety copy of your previous database was saved to: ' + safety_path if safety_path else 'No changes were made to your live database.'}")
            db.get_conn()
            return
        finally:
            db.get_conn()

        log_activity(self.current_user, f"Restored database from backup '{src_path}' (safety copy saved at '{safety_path}')")
        messagebox.showinfo("Restore Complete",
                             f"Database restored from:\n{src_path}\n\n"
                             f"A safety copy of your previous database was saved to:\n{safety_path}\n\n"
                             "Please restart the application now so every screen reloads with the restored data.")
        self.logout()


def main():
    db.init_db()
    login_root = tk.Tk()
    theme.apply_ttk_style()
    LoginWindow(login_root)
    login_root.mainloop()


if __name__ == "__main__":
    main()