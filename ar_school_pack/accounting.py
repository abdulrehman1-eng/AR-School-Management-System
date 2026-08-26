"""
accounting.py — Revenue, expenses, and the fee/salary integration.

Key integration rules (per spec sections 19-24):
- A student fee payment (an increase in `paid_fee`) automatically creates a
  matching accounting_revenue row — no need to enter it twice.
- A salary payslip run automatically creates a matching accounting_expense
  row categorized as 'Salary'.
- Every write here goes through rbac.require() so a Teacher/Reception
  account cannot record or read accounting entries even by calling these
  functions directly.
"""

from datetime import datetime
import os
import db
import rbac


def record_fee_revenue(role, student_id, amount, recorded_by, description="Fee payment", payment_method="Cash"):
    if amount <= 0:
        return
    rbac.require(role, "student.fee.edit")
    db.run(
        """INSERT INTO accounting_revenue
           (source_type, student_id, amount, date, description, reference, payment_method, recorded_by)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        ("Student Fee", student_id, amount, datetime.now().strftime("%Y-%m-%d"),
         description, f"STU-{student_id}", payment_method, recorded_by),
        commit=True,
    )


def record_admission_fee_revenue(role, student_id, amount, recorded_by,
                                 description="One-time admission fee", payment_method="Cash"):
    if amount <= 0:
        return
    try:
        rbac.require(role, "student.fee.edit")
    except Exception:
        rbac.require(role, "accounting.revenue.add")
    db.run(
        """INSERT INTO accounting_revenue
           (source_type, student_id, amount, date, description, reference, payment_method, recorded_by)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        ("Admission Fee", student_id, amount, datetime.now().strftime("%Y-%m-%d"),
         description, f"ADM-{student_id}", payment_method, recorded_by),
        commit=True,
    )


def add_revenue(role, source_type, amount, description, reference, payment_method, recorded_by, student_id=None):
    rbac.require(role, "accounting.revenue.add")
    db.run(
        """INSERT INTO accounting_revenue
           (source_type, student_id, amount, date, description, reference, payment_method, recorded_by)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        (source_type, student_id, amount, datetime.now().strftime("%Y-%m-%d"),
         description, reference, payment_method, recorded_by),
        commit=True,
    )


def record_salary_expense(role, teacher_id, teacher_name, amount, month, recorded_by, reference=""):
    rbac.require(role, "teacher.salary.pay")
    db.run(
        """INSERT INTO accounting_expense
           (category, amount, date, description, vendor_or_person, payment_method, reference, recorded_by)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        ("Salary", amount, datetime.now().strftime("%Y-%m-%d"),
         f"Salary payment for {month}", f"{teacher_name} ({teacher_id})", "Bank/Cash", reference, recorded_by),
        commit=True,
    )


def add_expense(role, category, amount, description, vendor, payment_method, reference, recorded_by):
    rbac.require(role, "accounting.expense.add")
    db.run(
        """INSERT INTO accounting_expense
           (category, amount, date, description, vendor_or_person, payment_method, reference, recorded_by)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        (category, amount, datetime.now().strftime("%Y-%m-%d"), description, vendor, payment_method, reference, recorded_by),
        commit=True,
    )


def list_revenue(role, start_date=None, end_date=None):
    rbac.require(role, "accounting.revenue.view")
    if start_date and end_date:
        return db.run(
            "SELECT id, source_type, student_id, amount, date, description, payment_method FROM accounting_revenue "
            "WHERE date BETWEEN ? AND ? ORDER BY id DESC", (start_date, end_date), fetchall=True,
        )
    return db.run(
        "SELECT id, source_type, student_id, amount, date, description, payment_method FROM accounting_revenue ORDER BY id DESC",
        fetchall=True,
    )


def list_expense(role, start_date=None, end_date=None):
    rbac.require(role, "accounting.expense.view")
    if start_date and end_date:
        return db.run(
            "SELECT id, category, amount, date, description, vendor_or_person, payment_method FROM accounting_expense "
            "WHERE date BETWEEN ? AND ? ORDER BY id DESC", (start_date, end_date), fetchall=True,
        )
    return db.run(
        "SELECT id, category, amount, date, description, vendor_or_person, payment_method FROM accounting_expense ORDER BY id DESC",
        fetchall=True,
    )


def dashboard_totals(role, start_date=None, end_date=None):
    """All-time totals PLUS always-current calendar month & today.

    month_revenue / month_expense / month_net_income are ALWAYS the current
    calendar month — used by the main Dashboard cards.
    """
    rbac.require(role, "accounting.dashboard")
    if start_date and end_date:
        rev = db.run("SELECT COALESCE(SUM(amount),0) FROM accounting_revenue WHERE date BETWEEN ? AND ?",
                      (start_date, end_date), fetchone=True)[0]
        exp = db.run("SELECT COALESCE(SUM(amount),0) FROM accounting_expense WHERE date BETWEEN ? AND ?",
                      (start_date, end_date), fetchone=True)[0]
    else:
        rev = db.run("SELECT COALESCE(SUM(amount),0) FROM accounting_revenue", fetchone=True)[0]
        exp = db.run("SELECT COALESCE(SUM(amount),0) FROM accounting_expense", fetchone=True)[0]

    today = datetime.now().strftime("%Y-%m-%d")
    month = datetime.now().strftime("%Y-%m")
    today_rev = db.run("SELECT COALESCE(SUM(amount),0) FROM accounting_revenue WHERE date=?", (today,), fetchone=True)[0]
    today_exp = db.run("SELECT COALESCE(SUM(amount),0) FROM accounting_expense WHERE date=?", (today,), fetchone=True)[0]
    month_rev = db.run("SELECT COALESCE(SUM(amount),0) FROM accounting_revenue WHERE date LIKE ?", (f"{month}%",), fetchone=True)[0]
    month_exp = db.run("SELECT COALESCE(SUM(amount),0) FROM accounting_expense WHERE date LIKE ?", (f"{month}%",), fetchone=True)[0]
    month_net = float(month_rev or 0) - float(month_exp or 0)

    return {
        "total_revenue": float(rev or 0),
        "total_expense": float(exp or 0),
        "net_income": float(rev or 0) - float(exp or 0),
        "today_revenue": float(today_rev or 0),
        "today_expense": float(today_exp or 0),
        "month_revenue": float(month_rev or 0),
        "month_expense": float(month_exp or 0),
        "month_net_income": month_net,
    }


def _period_date_range(period_key, custom_start=None, custom_end=None):
    today = datetime.now().date()
    end = today
    if period_key == "custom" and custom_start and custom_end:
        try:
            start = datetime.strptime(custom_start.strip()[:10], "%Y-%m-%d").date()
            end = datetime.strptime(custom_end.strip()[:10], "%Y-%m-%d").date()
        except ValueError:
            start = today.replace(day=1)
            end = today
        label = f"Custom ({start.strftime('%Y-%m-%d')} → {end.strftime('%Y-%m-%d')})"
    elif period_key in ("1m", "this_month"):
        start = today.replace(day=1)
        label = f"This Month ({start.strftime('%b %Y')})"
    elif period_key == "3m":
        y, m = today.year, today.month
        m -= 2
        while m <= 0:
            m += 12
            y -= 1
        start = datetime(y, m, 1).date()
        label = f"Last 3 Months ({start.strftime('%b %Y')} – {today.strftime('%b %Y')})"
    elif period_key == "6m":
        y, m = today.year, today.month
        m -= 5
        while m <= 0:
            m += 12
            y -= 1
        start = datetime(y, m, 1).date()
        label = f"Last 6 Months ({start.strftime('%b %Y')} – {today.strftime('%b %Y')})"
    elif period_key in ("1y", "this_year"):
        start = today.replace(month=1, day=1)
        label = f"This Year ({today.year})"
    else:
        start = today.replace(day=1)
        label = f"This Month ({start.strftime('%b %Y')})"
    if start > end:
        start, end = end, start
    return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d"), label


def period_totals(role, period_key="1m", custom_start=None, custom_end=None):
    rbac.require(role, "accounting.dashboard")
    start, end, label = _period_date_range(period_key, custom_start, custom_end)
    rev = db.run(
        "SELECT COALESCE(SUM(amount),0) FROM accounting_revenue WHERE date BETWEEN ? AND ?",
        (start, end), fetchone=True,
    )[0]
    exp = db.run(
        "SELECT COALESCE(SUM(amount),0) FROM accounting_expense WHERE date BETWEEN ? AND ?",
        (start, end), fetchone=True,
    )[0]
    return {
        "start_date": start, "end_date": end, "label": label,
        "revenue": float(rev or 0), "expense": float(exp or 0),
        "net_income": float(rev or 0) - float(exp or 0),
    }


def is_teacher_salary_paid_this_month(teacher_id, month=None):
    """Return True if a Salary expense already exists for this teacher this month.

    Matches the same convention used by generate_salary_payslip:
    category='Salary' AND vendor_or_person LIKE '%(TCH-xxx)%' AND date LIKE 'YYYY-MM%'.
    """
    if not teacher_id:
        return False
    ym = month or datetime.now().strftime("%Y-%m")
    row = db.run(
        "SELECT id FROM accounting_expense WHERE category='Salary' AND date LIKE ? AND vendor_or_person LIKE ? LIMIT 1",
        (f"{ym}%", f"%({teacher_id})%"),
        fetchone=True,
    )
    return bool(row)


def teacher_salary_status_map(month=None):
    """Return dict teacher_id -> {'paid': bool, 'amount': float|None, 'date': str|None} for the month."""
    ym = month or datetime.now().strftime("%Y-%m")
    rows = db.run(
        "SELECT vendor_or_person, amount, date FROM accounting_expense "
        "WHERE category='Salary' AND date LIKE ?",
        (f"{ym}%",),
        fetchall=True,
    ) or []
    out = {}
    for vendor, amount, date in rows:
        # vendor format: "Name (TCH-001)"
        tid = None
        if vendor and "(" in vendor and vendor.endswith(")"):
            tid = vendor[vendor.rfind("(") + 1 : -1].strip()
        if tid:
            out[tid] = {"paid": True, "amount": float(amount or 0), "date": date}
    return out


def export_finance_excel(role, period_key, out_path, recorded_by="",
                         custom_start=None, custom_end=None):
    rbac.require(role, "accounting.dashboard")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for Excel export. Install: pip install openpyxl") from exc

    totals = period_totals(role, period_key, custom_start, custom_end)
    start, end = totals["start_date"], totals["end_date"]
    revenues = list_revenue(role, start, end) or []
    expenses = list_expense(role, start, end) or []

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1e3a5f")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    title_font = Font(bold=True, name="Calibri", size=14, color="1e3a5f")
    money_font = Font(name="Calibri", size=11)
    thin = Border(
        left=Side(style="thin", color="cbd5e1"), right=Side(style="thin", color="cbd5e1"),
        top=Side(style="thin", color="cbd5e1"), bottom=Side(style="thin", color="cbd5e1"),
    )
    green_fill = PatternFill("solid", fgColor="dcfce7")
    red_fill = PatternFill("solid", fgColor="fee2e2")
    blue_fill = PatternFill("solid", fgColor="e0f2fe")

    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "AR School Management System — Finance Report"
    ws["A1"].font = title_font
    ws.merge_cells("A1:D1")
    ws["A2"] = f"Period: {totals['label']}"
    ws["A3"] = f"Date range: {start}  →  {end}"
    ws["A4"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  by  {recorded_by or '—'}"
    ws["A6"] = "Metric"
    ws["B6"] = "Amount (Rs.)"
    for col in ("A", "B"):
        ws[f"{col}6"].fill = header_fill
        ws[f"{col}6"].font = header_font
        ws[f"{col}6"].alignment = Alignment(horizontal="center")
    for i, (label, amount, fill) in enumerate([
        ("Total Revenue", totals["revenue"], green_fill),
        ("Total Expenses / Spends", totals["expense"], red_fill),
        ("Net Income", totals["net_income"], blue_fill if totals["net_income"] >= 0 else red_fill),
    ], start=7):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = amount
        ws[f"B{i}"].number_format = "#,##0.00"
        ws[f"A{i}"].font = money_font
        ws[f"B{i}"].font = Font(bold=True, name="Calibri", size=11)
        ws[f"A{i}"].fill = fill
        ws[f"B{i}"].fill = fill
        ws[f"A{i}"].border = thin
        ws[f"B{i}"].border = thin
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    ws_r = wb.create_sheet("Revenue")
    for col_idx, h in enumerate(["ID", "Source / Type", "Student ID", "Amount (Rs.)", "Date", "Description", "Payment Method"], 1):
        cell = ws_r.cell(1, col_idx, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin
    for row_idx, r in enumerate(revenues, start=2):
        for col_idx, val in enumerate(r, start=1):
            cell = ws_r.cell(row_idx, col_idx, val if val is not None else "")
            cell.border = thin
            if col_idx == 4:
                cell.number_format = "#,##0.00"
    for col, w in zip("ABCDEFG", (8, 18, 14, 14, 12, 36, 14)):
        ws_r.column_dimensions[col].width = w
    if not revenues:
        ws_r.cell(2, 1, "No revenue entries in this period.")

    ws_e = wb.create_sheet("Expenses")
    for col_idx, h in enumerate(["ID", "Category", "Amount (Rs.)", "Date", "Description", "Vendor / Person", "Payment Method"], 1):
        cell = ws_e.cell(1, col_idx, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin
    for row_idx, r in enumerate(expenses, start=2):
        for col_idx, val in enumerate(r, start=1):
            cell = ws_e.cell(row_idx, col_idx, val if val is not None else "")
            cell.border = thin
            if col_idx == 3:
                cell.number_format = "#,##0.00"
    for col, w in zip("ABCDEFG", (8, 16, 14, 12, 36, 22, 14)):
        ws_e.column_dimensions[col].width = w
    if not expenses:
        ws_e.cell(2, 1, "No expense entries in this period.")

    parent = os.path.dirname(out_path)
    if parent:
        os.makedirs(parent, exist_ok=True)
    wb.save(out_path)
    return out_path
