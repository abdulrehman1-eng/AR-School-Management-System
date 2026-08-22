import os
import re
from datetime import datetime
import openpyxl  # Python library for Excel export
import db
import rbac
import branding


def _safe_org_slug(name):
    """Turn organization name into a filesystem-safe filename fragment."""
    text = (name or "").strip() or "School"
    text = re.sub(r"[^\w\s\-]", "", text, flags=re.UNICODE)
    text = re.sub(r"[\s\-]+", "_", text).strip("_")
    return (text[:40] or "School")


def _table_exists(table_name):
    row = db.run(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
        (table_name,),
        fetchone=True,
    )
    return bool(row)


def _table_columns(table_name):
    rows = db.run(f"PRAGMA table_info({table_name})", fetchall=True) or []
    # PRAGMA returns: cid, name, type, notnull, dflt_value, pk
    return {r[1].lower(): r[1] for r in rows}


def _pick(cols, *candidates):
    """Return actual column name for the first matching candidate (case-insensitive)."""
    for c in candidates:
        if c.lower() in cols:
            return cols[c.lower()]
    return None


def _fetch_from_fee_cycles(month, year):
    """
    Preferred source: monthly fee_cycles.
    Only Pending / Partial / Overdue are exported — Paid is excluded.
    """
    if not _table_exists("fee_cycles"):
        return None

    cols = _table_columns("fee_cycles")
    c_student = _pick(cols, "student_id", "student")
    # Prefer the real fee_cycles schema (billing_month / billing_year) used by
    # fee_cycles.py; keep older aliases so any legacy table still works.
    c_month = _pick(cols, "billing_month", "month", "fee_month", "cycle_month")
    c_year = _pick(cols, "billing_year", "year", "fee_year", "cycle_year")
    c_status = _pick(cols, "status", "fee_status", "cycle_status")
    c_total = _pick(cols, "amount_due", "fee_amount", "total_fee", "total_amount", "amount", "due_amount")
    c_paid = _pick(cols, "amount_paid", "paid_fee", "paid_amount", "paid")
    c_balance = _pick(cols, "balance", "remaining_fee", "remaining", "due_balance")

    if not (c_student and c_month and c_year):
        return None

    # Build remaining expression — prefer amount_due - amount_paid (ledger truth)
    if c_balance:
        remaining_expr = f"COALESCE(c.{c_balance}, 0)"
    elif c_total and c_paid:
        remaining_expr = f"(COALESCE(c.{c_total}, 0) - COALESCE(c.{c_paid}, 0))"
    elif c_total:
        remaining_expr = f"COALESCE(c.{c_total}, 0)"
    else:
        remaining_expr = "0"

    total_expr = f"COALESCE(c.{c_total}, 0)" if c_total else remaining_expr
    paid_expr = f"COALESCE(c.{c_paid}, 0)" if c_paid else f"({total_expr} - ({remaining_expr}))"

    status_filter = ""
    if c_status:
        # Explicitly keep only unpaid statuses; never include Paid / Cleared / Complete
        status_filter = f"""
          AND LOWER(COALESCE(c.{c_status}, '')) IN (
              'pending', 'partial', 'overdue', 'unpaid', 'due'
          )
        """

    query = f"""
        SELECT
            s.student_id,
            s.name,
            s.father_name,
            s.class_sec,
            s.phone,
            {total_expr} AS total_fee,
            {paid_expr} AS paid_fee,
            {remaining_expr} AS remaining_fee,
            {"c." + c_status if c_status else "'Pending'"} AS fee_status
        FROM fee_cycles c
        JOIN students s ON s.student_id = c.{c_student}
        WHERE c.{c_month} = ?
          AND c.{c_year} = ?
          AND COALESCE(s.status, 'Active') = 'Active'
          AND ({remaining_expr}) > 0
          {status_filter}
        ORDER BY s.class_sec, s.student_id
    """
    return db.run(query, (month, year), fetchall=True)


def _fetch_from_students_fallback():
    """
    Fallback when fee_cycles table is missing:
    students.total_fee - students.paid_fee > 0 only.
    """
    query = """
        SELECT
            s.student_id,
            s.name,
            s.father_name,
            s.class_sec,
            s.phone,
            COALESCE(s.total_fee, 0) AS total_fee,
            COALESCE(s.paid_fee, 0) AS paid_fee,
            (COALESCE(s.total_fee, 0) - COALESCE(s.paid_fee, 0)) AS remaining_fee,
            CASE
                WHEN COALESCE(s.paid_fee, 0) <= 0 THEN 'Pending'
                WHEN COALESCE(s.paid_fee, 0) < COALESCE(s.total_fee, 0) THEN 'Partial'
                ELSE 'Paid'
            END AS fee_status
        FROM students s
        WHERE COALESCE(s.status, 'Active') = 'Active'
          AND (COALESCE(s.total_fee, 0) - COALESCE(s.paid_fee, 0)) > 0
        ORDER BY s.class_sec, s.student_id
    """
    return db.run(query, fetchall=True)


def export_remaining_fees_to_excel(user_role, month, year, output_dir=None, actor="System"):
    """
    Export only students with PENDING fee for the selected month/year.

    Priority:
      1) fee_cycles for that month/year — statuses Pending / Partial / Overdue
         (Paid cycles are never included)
      2) Fallback: students table where total_fee > paid_fee

    Fully paid students are excluded in both paths.
    """
    if not rbac.can(user_role, "fee.reports.view"):
        raise PermissionError("Access Denied: Is action ki permission nahi hai.")

    records = _fetch_from_fee_cycles(month, year)
    source = "fee_cycles"
    if records is None:
        records = _fetch_from_students_fallback()
        source = "students"

    if not records:
        return {
            "success": False,
            "count": 0,
            "message": (
                f"Koi remaining fee nahi mili ({month:02d}/{year}). "
                "Jo students ki fee fully paid hai unko list se hata diya gaya."
            ),
        }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Remaining Fees {month}-{year}"

    headers = [
        "Student ID", "Student Name", "Father Name",
        "Class/Sec", "Phone Number", "Total Fee",
        "Paid Fee", "Remaining Fee", "Fee Status",
    ]
    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_num).font = openpyxl.styles.Font(bold=True)

    count = 0
    total_remaining_sum = 0.0

    for r in records:
        s_id, name, fname, cls, phone, total_f, paid_f, remaining, fee_status = r
        total_f = float(total_f or 0)
        paid_f = float(paid_f or 0)
        remaining = float(remaining or 0)
        status_txt = (fee_status or "Pending").strip().title()

        # Hard safety: never export paid / zero-balance rows
        if remaining <= 0:
            continue
        if status_txt.lower() in ("paid", "cleared", "complete", "completed", "settled"):
            continue

        ws.append([
            s_id,
            name,
            fname,
            cls,
            phone or "N/A",
            total_f,
            paid_f,
            remaining,
            status_txt,
        ])
        count += 1
        total_remaining_sum += remaining

    if count == 0:
        return {
            "success": False,
            "count": 0,
            "message": (
                f"Koi remaining fee nahi mili ({month:02d}/{year}). "
                "Fully paid students exclude ho chuke hain."
            ),
        }

    ws.append([])
    ws.append(["TOTAL", "", "", "", "", "", "", total_remaining_sum, ""])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    if not output_dir:
        output_dir = os.getcwd()

    org = _safe_org_slug(branding.get_branding().get("org_name"))
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{org}_Remaining_Fees_{month:02d}_{year}_{stamp}.xlsx"
    file_path = os.path.join(output_dir, filename)
    wb.save(file_path)

    return {
        "success": True,
        "count": count,
        "total_amount": total_remaining_sum,
        "path": file_path,
        "source": source,
    }
