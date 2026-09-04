"""
smart_attendance.py — Smart barcode/manual Attendance window with
duplicate-scan protection, live counters, and an automatic end-of-day
Absent process.

STANDALONE MODULE. The existing Attendance tab in app.py (and its
`attendance` table) is untouched — this window reads/writes the SAME
`attendance` table, so a record created here is identical to one created
by the original tab. The `attendance` table's existing
UNIQUE(student_id, date) constraint is what this module leans on for
duplicate-scan protection, rather than re-implementing that check by hand.

Two small, additive, self-contained tables are created here (never
touching db.py):
  - attendance_settings(id=1, start_time, closing_time, late_threshold_minutes)
    — configurable timing, per spec section 8, instead of hardcoding it.
  - attendance_auto_absent_log(date PK, run_at) — a one-row-per-day marker
    so the automatic end-of-day Absent sweep is provably safe to run more
    than once and will never double-insert.
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime

import db
import rbac
import theme

STATUS_VALUES = ["Present", "Absent", "Leave", "Late"]


def _pick_date(parent, initial=None, title="Select Date"):
    """Simple pure-tkinter calendar popup. Returns YYYY-MM-DD or None."""
    try:
        base = datetime.strptime(initial, "%Y-%m-%d") if initial else datetime.now()
    except Exception:
        base = datetime.now()

    result = {"date": None}
    top = tk.Toplevel(parent)
    top.title(title)
    top.geometry("280x300")
    top.config(bg=theme.WHITE)
    top.transient(parent)
    top.grab_set()

    y_var = tk.IntVar(value=base.year)
    m_var = tk.IntVar(value=base.month)
    grid_host = tk.Frame(top, bg=theme.WHITE)
    grid_host.pack(fill=tk.BOTH, expand=True, padx=8, pady=4)

    def draw(_ev=None):
        for w in grid_host.winfo_children():
            w.destroy()
        y, m = y_var.get(), m_var.get()
        tk.Label(
            grid_host, text=f"{MONTH_NAMES_FULL[m - 1]} {y}",
            font=theme.FONT_BODY_BOLD, bg=theme.WHITE,
        ).grid(row=0, column=0, columnspan=7, pady=(4, 6))
        for i, dn in enumerate(["Mo", "Tu", "We", "Th", "Fr", "Sa", "Su"]):
            tk.Label(grid_host, text=dn, font=theme.FONT_SMALL, bg=theme.WHITE,
                     fg=theme.TEXT_MUTED).grid(row=1, column=i, padx=2, pady=2)
        import calendar as _cal
        for week_i, week in enumerate(_cal.monthcalendar(y, m), start=2):
            for col, day in enumerate(week):
                if day == 0:
                    tk.Label(grid_host, text="", width=3, bg=theme.WHITE).grid(row=week_i, column=col)
                    continue

                def _sel(d=day, yy=y, mm=m):
                    result["date"] = f"{yy:04d}-{mm:02d}-{d:02d}"
                    top.destroy()

                btn = tk.Button(
                    grid_host, text=str(day), width=3, relief=tk.FLAT,
                    bg=theme.WHITE, activebackground=theme.BRAND_BLUE,
                    command=_sel,
                )
                if day == base.day and m == base.month and y == base.year:
                    btn.config(bg="#dbeafe")
                btn.grid(row=week_i, column=col, padx=1, pady=1)

    nav = tk.Frame(top, bg=theme.WHITE)
    nav.pack(fill=tk.X, pady=4)

    def prev_m():
        m, y = m_var.get() - 1, y_var.get()
        if m < 1:
            m, y = 12, y - 1
        m_var.set(m)
        y_var.set(y)
        draw()

    def next_m():
        m, y = m_var.get() + 1, y_var.get()
        if m > 12:
            m, y = 1, y + 1
        m_var.set(m)
        y_var.set(y)
        draw()

    theme.primary_button(nav, "◀", prev_m, bg=theme.SLATE).pack(side=tk.LEFT, padx=8)
    theme.primary_button(nav, "▶", next_m, bg=theme.SLATE).pack(side=tk.RIGHT, padx=8)
    draw()
    top.wait_window()
    return result["date"]



MONTH_NAMES_FULL = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]


# ---------------------------------------------------------------------------
# Export helpers — Excel (openpyxl) + PDF (reportlab)
# ---------------------------------------------------------------------------

def _ask_save_path(parent, default_name, kind="xlsx"):
    if kind == "xlsx":
        path = filedialog.asksaveasfilename(
            title="Save Excel Report",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Workbook", "*.xlsx"), ("All Files", "*.*")],
            parent=parent,
        )
        if path and not path.lower().endswith(".xlsx"):
            path += ".xlsx"
    else:
        path = filedialog.asksaveasfilename(
            title="Save PDF Report",
            defaultextension=".pdf",
            initialfile=default_name,
            filetypes=[("PDF Document", "*.pdf"), ("All Files", "*.*")],
            parent=parent,
        )
        if path and not path.lower().endswith(".pdf"):
            path += ".pdf"
    return path or None


def export_personal_attendance_excel(
    student_id, name, class_sec, year, month, month_counts, year_counts, day_rows, monthly_rows, path,
):
    """Write personal attendance Excel: summary + monthly breakdown + day history."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Personal Attendance"

    header_fill = PatternFill("solid", fgColor="0F172A")
    sub_fill = PatternFill("solid", fgColor="0284C7")
    thin = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    white_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=12)
    body_font = Font(name="Segoe UI", size=10)

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "Personal Attendance Report — AR School Management System"
    c.font = white_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:G2")
    ws["A2"].value = (
        f"Student: {name} ({student_id})  |  Class: {class_sec or '-'}  |  "
        f"Period: {MONTH_NAMES_FULL[month - 1]} {year}  |  "
        f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    ws["A2"].font = Font(name="Segoe UI", size=9, italic=True, color="64748B")

    ws["A4"] = "SUMMARY"
    ws["A4"].font = Font(name="Segoe UI", bold=True, size=11, color="0F172A")
    headers_sum = ["Scope", "Present", "Absent", "Leave", "Late", "Total", "Rate %"]
    for col, h in enumerate(headers_sum, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = Font(name="Segoe UI", bold=True, color="FFFFFF")
        cell.fill = sub_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    for row_i, (label, counts) in enumerate(
        [
            (f"{MONTH_NAMES_FULL[month - 1]} {year}", month_counts),
            (f"Year {year}", year_counts),
        ],
        start=6,
    ):
        vals = [
            label,
            counts.get("Present", 0),
            counts.get("Absent", 0),
            counts.get("Leave", 0),
            counts.get("Late", 0),
            counts.get("Total", 0),
            f"{counts.get('Rate', 0):.1f}%",
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_i, column=col, value=v)
            cell.font = body_font
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")

    start = 9
    ws.cell(row=start, column=1, value="MONTHLY BREAKDOWN").font = Font(
        name="Segoe UI", bold=True, size=11, color="0F172A"
    )
    for col, h in enumerate(
        ["Month", "Present", "Absent", "Leave", "Late", "Total", "Rate %"], 1
    ):
        cell = ws.cell(row=start + 1, column=col, value=h)
        cell.font = Font(name="Segoe UI", bold=True, color="FFFFFF")
        cell.fill = sub_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin
    for i, row in enumerate(monthly_rows or [], start=start + 2):
        for col, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.font = body_font
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")

    ws2 = wb.create_sheet("Day History")
    ws2.merge_cells("A1:D1")
    ws2["A1"].value = f"Day-wise Attendance — {name} ({student_id}) — {MONTH_NAMES_FULL[month - 1]} {year}"
    ws2["A1"].font = white_font
    ws2["A1"].fill = header_fill
    for col, h in enumerate(["Date", "Status", "Method", "In Time"], 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = Font(name="Segoe UI", bold=True, color="FFFFFF")
        cell.fill = sub_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin
    for i, (d, st, method, in_time) in enumerate(day_rows or [], start=4):
        for col, v in enumerate(
            [d, st, method or "—", (in_time or "").strip() or "—"], 1
        ):
            cell = ws2.cell(row=i, column=col, value=v)
            cell.font = body_font
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")

    for ws_x in (ws, ws2):
        for col in range(1, 8):
            ws_x.column_dimensions[get_column_letter(col)].width = 14

    wb.save(path)
    return path


def export_personal_attendance_pdf(
    student_id, name, class_sec, year, month, month_counts, year_counts, day_rows, path,
):
    """Write personal attendance PDF using reportlab."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    doc = SimpleDocTemplate(
        path, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=16 * mm, bottomMargin=16 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleAR", parent=styles["Heading1"], fontSize=14, textColor=colors.HexColor("#0F172A"), spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        "SubAR", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#64748B"), spaceAfter=8,
    )
    h2 = ParagraphStyle(
        "H2AR", parent=styles["Heading2"], fontSize=11, textColor=colors.HexColor("#0284C7"), spaceBefore=10, spaceAfter=6,
    )

    story = []
    story.append(Paragraph("Personal Attendance Report", title_style))
    story.append(Paragraph("AR School Management System", sub_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#E2E8F0")))
    story.append(Spacer(1, 6))
    story.append(Paragraph(
        f"<b>Student:</b> {name} &nbsp;&nbsp; <b>ID:</b> {student_id} &nbsp;&nbsp; "
        f"<b>Class:</b> {class_sec or '-'}<br/>"
        f"<b>Period:</b> {MONTH_NAMES_FULL[month - 1]} {year} &nbsp;&nbsp; "
        f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        styles["Normal"],
    ))
    story.append(Spacer(1, 10))

    story.append(Paragraph("Summary", h2))
    sum_data = [
        ["Scope", "Present", "Absent", "Leave", "Late", "Total", "Rate %"],
        [
            f"{MONTH_NAMES_FULL[month - 1]} {year}",
            month_counts.get("Present", 0),
            month_counts.get("Absent", 0),
            month_counts.get("Leave", 0),
            month_counts.get("Late", 0),
            month_counts.get("Total", 0),
            f"{month_counts.get('Rate', 0):.1f}%",
        ],
        [
            f"Year {year}",
            year_counts.get("Present", 0),
            year_counts.get("Absent", 0),
            year_counts.get("Leave", 0),
            year_counts.get("Late", 0),
            year_counts.get("Total", 0),
            f"{year_counts.get('Rate', 0):.1f}%",
        ],
    ]
    t = Table(sum_data, colWidths=[90, 55, 55, 50, 50, 50, 55])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0284C7")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(t)

    story.append(Paragraph(f"Day-wise History — {MONTH_NAMES_FULL[month - 1]} {year}", h2))
    day_data = [["Date", "Status", "Method", "In Time"]]
    for d, st, method, in_time in (day_rows or [])[:120]:
        day_data.append([d, st, method or "—", (in_time or "").strip() or "—"])
    if len(day_data) == 1:
        day_data.append(["No records", "—", "—", "—"])
    t2 = Table(day_data, colWidths=[90, 70, 90, 70])
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(t2)
    story.append(Spacer(1, 12))
    story.append(Paragraph(
        "This report is system-generated from live attendance records.",
        sub_style,
    ))
    doc.build(story)
    return path


def export_class_attendance_excel(
    class_name, year, month, month_agg, year_agg, student_rows, path,
):
    """Write class-wise attendance Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Class Attendance"

    header_fill = PatternFill("solid", fgColor="0F172A")
    sub_fill = PatternFill("solid", fgColor="0284C7")
    thin = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    white_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=12)
    body_font = Font(name="Segoe UI", size=10)

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "Class-wise Attendance Report — AR School Management System"
    c.font = white_font
    c.fill = header_fill
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:H2")
    ws["A2"].value = (
        f"Class: {class_name}  |  Period: {MONTH_NAMES_FULL[month - 1]} {year}  |  "
        f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    ws["A2"].font = Font(name="Segoe UI", size=9, italic=True, color="64748B")

    ws["A4"] = "CLASS SUMMARY"
    ws["A4"].font = Font(name="Segoe UI", bold=True, size=11)
    for col, h in enumerate(
        ["Scope", "Present", "Absent", "Leave", "Late", "Total", "Rate %"], 1
    ):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = Font(name="Segoe UI", bold=True, color="FFFFFF")
        cell.fill = sub_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    mp, ma, ml, mlt, mt, mr = month_agg
    yp, ya, yl, ylt, yt, yr = year_agg
    for row_i, vals in enumerate(
        [
            [f"{MONTH_NAMES_FULL[month - 1]} {year}", mp, ma, ml, mlt, mt, f"{mr:.1f}%"],
            [f"Year {year}", yp, ya, yl, ylt, yt, f"{yr:.1f}%"],
        ],
        start=6,
    ):
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row_i, column=col, value=v)
            cell.font = body_font
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")

    ws.cell(row=9, column=1, value="PER-STUDENT BREAKDOWN").font = Font(
        name="Segoe UI", bold=True, size=11
    )
    headers = [
        "Student ID", "Name", "Present", "Absent", "Leave", "Late", "Total", "Rate %",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col, value=h)
        cell.font = Font(name="Segoe UI", bold=True, color="FFFFFF")
        cell.fill = sub_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    for i, row in enumerate(student_rows or [], start=11):
        for col, v in enumerate(row, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.font = body_font
            cell.border = thin
            cell.alignment = Alignment(horizontal="center" if col != 2 else "left")

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.column_dimensions["B"].width = 20

    wb.save(path)
    return path


def export_class_attendance_pdf(
    class_name, year, month, month_agg, year_agg, student_rows, path,
):
    """Write class-wise attendance PDF."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    doc = SimpleDocTemplate(
        path, pagesize=landscape(A4),
        leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleAR", parent=styles["Heading1"], fontSize=13, textColor=colors.HexColor("#0F172A"), spaceAfter=2,
    )
    sub_style = ParagraphStyle(
        "SubAR", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#64748B"), spaceAfter=6,
    )
    h2 = ParagraphStyle(
        "H2AR", parent=styles["Heading2"], fontSize=10, textColor=colors.HexColor("#0284C7"), spaceBefore=8, spaceAfter=4,
    )

    story = []
    story.append(Paragraph("Class-wise Attendance Report", title_style))
    story.append(Paragraph("AR School Management System", sub_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#E2E8F0")))
    story.append(Paragraph(
        f"<b>Class:</b> {class_name} &nbsp;&nbsp; "
        f"<b>Period:</b> {MONTH_NAMES_FULL[month - 1]} {year} &nbsp;&nbsp; "
        f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        styles["Normal"],
    ))
    story.append(Spacer(1, 8))

    mp, ma, ml, mlt, mt, mr = month_agg
    yp, ya, yl, ylt, yt, yr = year_agg
    story.append(Paragraph("Class Summary", h2))
    sum_data = [
        ["Scope", "Present", "Absent", "Leave", "Late", "Total", "Rate %"],
        [f"{MONTH_NAMES_FULL[month - 1]} {year}", mp, ma, ml, mlt, mt, f"{mr:.1f}%"],
        [f"Year {year}", yp, ya, yl, ylt, yt, f"{yr:.1f}%"],
    ]
    t = Table(sum_data, colWidths=[100, 60, 60, 55, 55, 55, 60])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0284C7")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(t)

    story.append(Paragraph("Per-Student Breakdown", h2))
    stu_data = [["Student ID", "Name", "Present", "Absent", "Leave", "Late", "Total", "Rate %"]]
    for row in (student_rows or []):
        stu_data.append(list(row))
    if len(stu_data) == 1:
        stu_data.append(["—", "No students", 0, 0, 0, 0, 0, "0%"])
    t2 = Table(stu_data, colWidths=[85, 120, 55, 55, 50, 50, 50, 55])
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (1, 1), (1, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(t2)
    story.append(Spacer(1, 10))
    story.append(Paragraph(
        "This report is system-generated from live attendance records.",
        sub_style,
    ))
    doc.build(story)
    return path




def _add_column_if_missing(table, column, coltype):
    cols = [r[1] for r in db.run(f"PRAGMA table_info({table})", fetchall=True)]
    if column not in cols:
        db.run(f"ALTER TABLE {table} ADD COLUMN {column} {coltype}", commit=True)


def _ensure_tables():
    # attendance.in_time doesn't exist on older/original schemas — added
    # here additively (never touches db.py) so a scan's clock time can be
    # recorded, matching the same non-destructive pattern db.py itself uses.
    _add_column_if_missing("attendance", "in_time", "TEXT")
    # Legacy table kept for backward compatibility; authoritative timings
    # now live in system_settings (db.get_setting / set_setting).
    db.run("""CREATE TABLE IF NOT EXISTS attendance_settings (
        id INTEGER PRIMARY KEY CHECK (id=1),
        start_time TEXT DEFAULT '08:00',
        closing_time TEXT DEFAULT '14:00',
        late_threshold_minutes INTEGER DEFAULT 15
    )""", commit=True)
    db.run("INSERT OR IGNORE INTO attendance_settings (id) VALUES (1)", commit=True)
    db.run("""CREATE TABLE IF NOT EXISTS attendance_auto_absent_log (
        date TEXT PRIMARY KEY,
        run_at TEXT
    )""", commit=True)


def get_settings():
    """Attendance timings from centralized system_settings.

    late_threshold_time is an absolute clock time (e.g. 08:15). For UI
    compatibility we also expose late_threshold_minutes derived from the
    gap between start and late cutoff when possible.
    """
    start = db.get_setting("school_start_time", "08:00") or "08:00"
    late = db.get_setting("late_threshold_time", "08:15") or "08:15"
    closing = db.get_setting("school_closing_time", "14:00") or "14:00"
    auto_enabled = db.get_setting("auto_absent_enabled", "1")
    minutes = 15
    try:
        s = datetime.strptime(start, "%H:%M")
        l = datetime.strptime(late, "%H:%M")
        minutes = max(0, int((l - s).total_seconds() // 60))
    except Exception:
        pass
    return {
        "start_time": start,
        "closing_time": closing,
        "late_threshold_time": late,
        "late_threshold_minutes": minutes,
        "auto_absent_enabled": str(auto_enabled) in ("1", "true", "True", "yes"),
    }


def set_settings(start_time, closing_time, late_threshold_minutes=None, late_threshold_time=None):
    """Persist attendance timings into system_settings (and mirror legacy table)."""
    from datetime import timedelta
    start_time = (start_time or "08:00").strip()
    closing_time = (closing_time or "14:00").strip()
    if late_threshold_time:
        late_time = str(late_threshold_time).strip()
    else:
        try:
            base = datetime.strptime(start_time, "%H:%M")
            mins = int(late_threshold_minutes if late_threshold_minutes is not None else 15)
            late_time = (base + timedelta(minutes=mins)).strftime("%H:%M")
        except Exception:
            late_time = "08:15"
    db.set_settings_bulk({
        "school_start_time": start_time,
        "school_closing_time": closing_time,
        "late_threshold_time": late_time,
    })
    try:
        mins = 15
        try:
            s = datetime.strptime(start_time, "%H:%M")
            l = datetime.strptime(late_time, "%H:%M")
            mins = max(0, int((l - s).total_seconds() // 60))
        except Exception:
            pass
        db.run(
            "UPDATE attendance_settings SET start_time=?, closing_time=?, late_threshold_minutes=? WHERE id=1",
            (start_time, closing_time, mins),
            commit=True,
        )
    except Exception:
        pass


def run_auto_absent(for_date=None, force=False):
    """Mark every Active student with NO attendance row for `for_date` as
    Absent (method='Auto System').

    Safe to call more than once for the same date — students who already have
    ANY attendance row (Present / Late / Leave / Absent) are left untouched.
    Returns the number of students newly marked.

    When force=False (background worker), respects auto_absent_enabled.
    When force=True (manual Admin button), runs regardless of the toggle.
    """
    _ensure_tables()
    d = for_date or datetime.now().strftime("%Y-%m-%d")
    if not force:
        if for_date is None or d == datetime.now().strftime("%Y-%m-%d"):
            if not get_settings().get("auto_absent_enabled", True):
                return 0
    active_ids = [r[0] for r in db.run(
        "SELECT student_id FROM students WHERE COALESCE(status,'Active')='Active'", fetchall=True)]
    already = {r[0] for r in db.run(
        "SELECT student_id FROM attendance WHERE date=?", (d,), fetchall=True)}
    to_mark = [sid for sid in active_ids if sid not in already]
    now_str = datetime.now().strftime("%H:%M")
    marked = 0
    for sid in to_mark:
        try:
            db.run(
                "INSERT INTO attendance (student_id, date, status, method) VALUES (?, ?, 'Absent', 'Auto System')",
                (sid, d), commit=True,
            )
            marked += 1
        except Exception:
            # UNIQUE(student_id, date) already satisfied by a concurrent write
            pass
    db.run(
        "INSERT INTO attendance_auto_absent_log (date, run_at) VALUES (?, ?) "
        "ON CONFLICT(date) DO UPDATE SET run_at=excluded.run_at",
        (d, now_str), commit=True,
    )
    return marked



def auto_absent_already_run_today(d=None):
    d = d or datetime.now().strftime("%Y-%m-%d")
    row = db.run("SELECT 1 FROM attendance_auto_absent_log WHERE date=?", (d,), fetchone=True)
    return row is not None


def is_past_school_closing():
    """True if the wall clock is at or after school_closing_time today."""
    closing_str = db.get_setting("school_closing_time", "14:00") or "14:00"
    try:
        closing = datetime.strptime(closing_str, "%H:%M").time()
    except ValueError:
        return False
    return datetime.now().time() >= closing


def try_auto_absent_now(force=False):
    """One-shot attempt: if enabled (or force) and past closing and not yet
    run today, mark missing students Absent. Returns (ran, count).
    Never raises."""
    try:
        if not force:
            if not get_settings().get("auto_absent_enabled", True):
                return False, 0
            if not is_past_school_closing():
                return False, 0
        if auto_absent_already_run_today() and not force:
            return False, 0
        n = run_auto_absent(force=force)
        return True, n
    except Exception as exc:
        print(f"[Auto-Absent] try_auto_absent_now error: {exc}")
        return False, 0


def start_auto_absent_worker(tk_root, poll_ms=60000, first_delay_ms=2000):
    """Background worker: while the main window is open, periodically check
    whether school closing time has passed and, if so, auto-mark Absents.

    - App opened AFTER closing time → runs within first_delay_ms
    - App opened BEFORE closing time → keeps polling; fires once closing arrives
    - Already ran today → idle checks, no duplicate rows
    - auto_absent_enabled=0 → does nothing

    Safe: every tick is wrapped; a failure never kills the UI loop.
    """
    state = {"done_date": None}

    def tick():
        try:
            today = datetime.now().strftime("%Y-%m-%d")
            # Reset daily so next calendar day can run again
            if state["done_date"] and state["done_date"] != today:
                state["done_date"] = None

            if state["done_date"] == today:
                pass  # already completed for today
            elif not get_settings().get("auto_absent_enabled", True):
                pass
            elif not is_past_school_closing():
                pass
            elif auto_absent_already_run_today():
                state["done_date"] = today
            else:
                ran, n = try_auto_absent_now(force=False)
                if ran:
                    state["done_date"] = today
                    print(
                        f"[Auto-Absent] Worker marked {n} student(s) Absent "
                        f"at {datetime.now().strftime('%H:%M')} "
                        f"(closing={db.get_setting('school_closing_time', '14:00')})."
                    )
        except Exception as exc:
            print(f"[Auto-Absent] Worker tick error (non-fatal): {exc}")
        finally:
            try:
                if tk_root.winfo_exists():
                    tk_root.after(poll_ms, tick)
            except Exception:
                pass

    try:
        tk_root.after(first_delay_ms, tick)
    except Exception as exc:
        print(f"[Auto-Absent] Could not start worker: {exc}")


class AttendanceWindow:
    def __init__(self, parent, user_role, current_user):
        self.parent = parent
        self.user_role = user_role
        self.current_user = current_user

        if not rbac.can(self.user_role, "attendance.mark") and not rbac.can(self.user_role, "attendance.view"):
            messagebox.showerror("Permission Denied",
                                  f"Role '{self.user_role}' cannot access attendance.", parent=parent)
            return

        _ensure_tables()
        self.can_mark = rbac.can(self.user_role, "attendance.mark")

        self.win = tk.Toplevel(parent)
        self.win.title("Smart Attendance")
        self.win.geometry("1100x820")
        self.win.config(bg="#F8FAFC")
        self.win.transient(parent)

        self._build_ui()
        self._maybe_run_auto_absent_on_open()
        self.win.after(150, lambda: self.ent_scan.focus_set() if self.can_mark else None)

    # ------------------------------------------------------------------
    def _build_ui(self):
        # ---- Dark navy header (matches reference) ----
        header = tk.Frame(self.win, bg="#0f172a", padx=18, pady=12)
        header.pack(fill=tk.X)
        title_row = tk.Frame(header, bg="#0f172a")
        title_row.pack(fill=tk.X)
        tk.Label(
            title_row, text="📇  SMART ATTENDANCE",
            font=("Segoe UI", 14, "bold"), bg="#0f172a", fg="white",
        ).pack(side=tk.LEFT)
        settings = get_settings()
        tk.Label(
            header,
            text=(
                f"School Time: {settings['start_time']} – {settings['closing_time']}  |  "
                f"Late after {settings.get('late_threshold_time', settings['start_time'])}"
            ),
            font=("Segoe UI", 9), bg="#0f172a", fg="#94a3b8",
        ).pack(anchor="w", pady=(2, 0))

        # ---- Fixed footer nav FIRST so it never clips ----
        actions = tk.Frame(self.win, bg="#F8FAFC", padx=12, pady=10)
        actions.pack(side=tk.BOTTOM, fill=tk.X)

        def _nav_btn(parent, text, cmd, bg, fg="white"):
            b = tk.Button(
                parent, text=text, command=cmd, bg=bg, fg=fg,
                font=("Segoe UI", 9, "bold"), bd=0, padx=12, pady=7,
                cursor="hand2", activeforeground=fg,
            )
            b.pack(side=tk.LEFT, padx=3)
            return b

        _nav_btn(actions, "👤  Personal", self._open_personal_attendance, "#2563eb")
        _nav_btn(actions, "◇  Attendance Review", self._open_attendance_review, "#2563eb")
        _nav_btn(actions, "📊  Reports", self._open_unified_reports, "#2563eb")
        if rbac.can(self.user_role, "settings.branding") or self.user_role == "Admin":
            _nav_btn(actions, "⚙  Timing", self._open_settings_dialog, "#475569")
            # Critical action — visually separated
            tk.Frame(actions, bg="#e2e8f0", width=1, height=28).pack(side=tk.LEFT, padx=8)
            _nav_btn(actions, "🌙  End-of-Day Absent", self._manual_run_auto_absent, "#c2410c")

        # ---- Scrollable body ----
        canvas_host = tk.Frame(self.win, bg="#F8FAFC")
        canvas_host.pack(fill=tk.BOTH, expand=True)
        canvas = tk.Canvas(canvas_host, bg="#F8FAFC", highlightthickness=0)
        vscroll = ttk.Scrollbar(canvas_host, orient=tk.VERTICAL, command=canvas.yview)
        body = tk.Frame(canvas, bg="#F8FAFC", padx=14, pady=12)
        body.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")),
        )
        canvas_win = canvas.create_window((0, 0), window=body, anchor="nw")
        canvas.configure(yscrollcommand=vscroll.set)

        def _on_canvas_cfg(event):
            canvas.itemconfig(canvas_win, width=event.width)

        canvas.bind("<Configure>", _on_canvas_cfg)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vscroll.pack(side=tk.RIGHT, fill=tk.Y)

        def _wheel(event):
            if getattr(event, "delta", 0):
                canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            elif getattr(event, "num", None) == 4:
                canvas.yview_scroll(-1, "units")
            elif getattr(event, "num", None) == 5:
                canvas.yview_scroll(1, "units")

        canvas.bind_all("<MouseWheel>", _wheel)
        canvas.bind_all("<Button-4>", _wheel)
        canvas.bind_all("<Button-5>", _wheel)

        def _unbind_wheel():
            try:
                canvas.unbind_all("<MouseWheel>")
                canvas.unbind_all("<Button-4>")
                canvas.unbind_all("<Button-5>")
            except Exception:
                pass

        self.win.protocol("WM_DELETE_WINDOW", lambda: (_unbind_wheel(), self.win.destroy()))

        # ---- Scanner card ----
        if self.can_mark:
            scan_card = tk.Frame(
                body, bg=theme.WHITE,
                highlightbackground="#e2e8f0", highlightthickness=1,
            )
            scan_card.pack(fill=tk.X, pady=(0, 10))
            scan_inner = tk.Frame(scan_card, bg=theme.WHITE, padx=16, pady=12)
            scan_inner.pack(fill=tk.X)

            tk.Label(
                scan_inner, text="Scanner — Ready",
                font=("Segoe UI", 12, "bold"), bg=theme.WHITE, fg="#0f172a",
            ).pack(anchor="w")

            row = tk.Frame(scan_inner, bg=theme.WHITE)
            row.pack(fill=tk.X, pady=(10, 0))
            tk.Label(
                row, text="Scan / Enter Student ID:",
                font=("Segoe UI", 10, "bold"), bg=theme.WHITE, fg="#334155",
            ).pack(side=tk.LEFT)
            self.ent_scan = tk.Entry(
                row, font=("Segoe UI", 12, "bold"), width=24,
                bg="#eff6ff", fg="#0f172a", insertbackground="#2563eb",
                relief="solid", bd=1,
                highlightthickness=2, highlightbackground="#93c5fd",
                highlightcolor="#2563eb",
            )
            self.ent_scan.pack(side=tk.LEFT, padx=10, ipady=6)
            self.ent_scan.bind("<Return>", lambda e: self.process_scan())
            tk.Button(
                row, text="Mark Present", command=self.process_scan,
                bg="#2563eb", fg="white", font=("Segoe UI", 10, "bold"),
                bd=0, padx=14, pady=7, cursor="hand2",
                activebackground="#1d4ed8", activeforeground="white",
            ).pack(side=tk.LEFT, padx=4)

            self.lbl_scan_result = tk.Label(
                scan_inner, text="Ready for next scan...",
                font=("Segoe UI", 10), bg=theme.WHITE, fg="#64748b",
            )
            self.lbl_scan_result.pack(anchor="w", pady=(10, 0))
        else:
            self.ent_scan = None
            self.lbl_scan_result = None

        # ---- Live summary cards (Present / Absent / Leave / Late / Total) ----
        self.summary_frame = tk.Frame(body, bg="#F8FAFC")
        self.summary_frame.pack(fill=tk.X, pady=(0, 10))
        self._render_summary()

        # ---- Manual Attendance ----
        if self.can_mark:
            man_card = tk.Frame(
                body, bg=theme.WHITE,
                highlightbackground="#e2e8f0", highlightthickness=1,
            )
            man_card.pack(fill=tk.X, pady=(0, 10))
            man_inner = tk.Frame(man_card, bg=theme.WHITE, padx=14, pady=12)
            man_inner.pack(fill=tk.BOTH, expand=True)

            tk.Label(
                man_inner, text="Manual Attendance",
                font=("Segoe UI", 12, "bold"), bg=theme.WHITE, fg="#0f172a",
            ).pack(anchor="w")

            search_row = tk.Frame(man_inner, bg=theme.WHITE)
            search_row.pack(fill=tk.X, pady=(10, 8))
            tk.Label(
                search_row, text="Search (ID / Name / Class):",
                bg=theme.WHITE, font=("Segoe UI", 9), fg="#64748b",
            ).pack(side=tk.LEFT)
            self.ent_manual_search = tk.Entry(
                search_row, font=("Segoe UI", 10), width=26,
                relief="solid", bd=1,
            )
            self.ent_manual_search.pack(side=tk.LEFT, padx=8, ipady=5)
            self.ent_manual_search.bind("<Return>", lambda e: self._manual_search())
            tk.Button(
                search_row, text="Search", command=self._manual_search,
                bg="#2563eb", fg="white", font=("Segoe UI", 9, "bold"),
                bd=0, padx=12, pady=6, cursor="hand2",
                activebackground="#1d4ed8", activeforeground="white",
            ).pack(side=tk.LEFT)

            # Table with dark header styling
            style = ttk.Style()
            try:
                style.configure(
                    "AttManual.Treeview",
                    font=("Segoe UI", 10),
                    rowheight=28,
                    background=theme.WHITE,
                    fieldbackground=theme.WHITE,
                    foreground="#0f172a",
                )
                style.configure(
                    "AttManual.Treeview.Heading",
                    font=("Segoe UI", 9, "bold"),
                    background="#0f172a",
                    foreground="white",
                )
                style.map("AttManual.Treeview", background=[("selected", "#dbeafe")])
            except Exception:
                pass

            table_wrap = tk.Frame(man_inner, bg=theme.WHITE)
            table_wrap.pack(fill=tk.BOTH, expand=True)
            self.tree_manual = ttk.Treeview(
                table_wrap,
                columns=("id", "name", "class", "today_status"),
                show="headings",
                height=6,
                style="AttManual.Treeview",
            )
            for c, h, w in [
                ("id", "Student ID", 120),
                ("name", "Name", 180),
                ("class", "Class", 100),
                ("today_status", "Today's Status", 120),
            ]:
                self.tree_manual.heading(c, text=h)
                self.tree_manual.column(c, width=w, anchor="center")
            man_scroll = ttk.Scrollbar(table_wrap, orient=tk.VERTICAL, command=self.tree_manual.yview)
            self.tree_manual.configure(yscrollcommand=man_scroll.set)
            man_scroll.pack(side=tk.RIGHT, fill=tk.Y)
            self.tree_manual.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, pady=(0, 8))
            try:
                self.tree_manual.tag_configure("Present", foreground="#16a34a")
                self.tree_manual.tag_configure("Absent", foreground="#dc2626")
                self.tree_manual.tag_configure("Leave", foreground="#2563eb")
                self.tree_manual.tag_configure("Late", foreground="#ea580c")
                self.tree_manual.tag_configure("even", background="#ffffff")
                self.tree_manual.tag_configure("odd", background="#f8fafc")
            except Exception:
                pass

            # Status action buttons — reference colours
            btnrow = tk.Frame(man_inner, bg=theme.WHITE)
            btnrow.pack(fill=tk.X, pady=(0, 2))
            status_colors = {
                "Present": "#16a34a",
                "Absent": "#dc2626",
                "Leave": "#2563eb",
                "Late": "#ea580c",
            }
            for status in STATUS_VALUES:
                tk.Button(
                    btnrow, text=status,
                    command=lambda st=status: self._manual_mark(st),
                    bg=status_colors.get(status, "#64748b"), fg="white",
                    font=("Segoe UI", 9, "bold"), bd=0, padx=14, pady=7,
                    cursor="hand2", activeforeground="white",
                ).pack(side=tk.LEFT, padx=(0, 6))
        else:
            self.ent_manual_search = None
            self.tree_manual = None

        # ---- Recent Attendance ----
        recent_card = tk.Frame(
            body, bg=theme.WHITE,
            highlightbackground="#e2e8f0", highlightthickness=1,
        )
        recent_card.pack(fill=tk.BOTH, expand=True, pady=(0, 4))
        recent_inner = tk.Frame(recent_card, bg=theme.WHITE, padx=14, pady=12)
        recent_inner.pack(fill=tk.BOTH, expand=True)

        tk.Label(
            recent_inner, text="Recent Attendance",
            font=("Segoe UI", 12, "bold"), bg=theme.WHITE, fg="#0f172a",
        ).pack(anchor="w", pady=(0, 8))

        try:
            style = ttk.Style()
            style.configure(
                "AttRecent.Treeview",
                font=("Segoe UI", 10),
                rowheight=28,
                background=theme.WHITE,
                fieldbackground=theme.WHITE,
                foreground="#0f172a",
            )
            style.configure(
                "AttRecent.Treeview.Heading",
                font=("Segoe UI", 9, "bold"),
                background="#0f172a",
                foreground="white",
            )
            style.map("AttRecent.Treeview", background=[("selected", "#dbeafe")])
        except Exception:
            pass

        recent_wrap = tk.Frame(recent_inner, bg=theme.WHITE)
        recent_wrap.pack(fill=tk.BOTH, expand=True)
        self.tree_recent = ttk.Treeview(
            recent_wrap,
            columns=("time", "id", "name", "class", "method", "status"),
            show="headings",
            height=8,
            style="AttRecent.Treeview",
        )
        col_cfg = [
            ("time", "Time", 70),
            ("id", "Student ID", 120),
            ("name", "Name", 160),
            ("class", "Class", 90),
            ("method", "Method", 100),
            ("status", "Status", 90),
        ]
        for c, h, w in col_cfg:
            self.tree_recent.heading(c, text=h)
            self.tree_recent.column(c, width=w, anchor="center")
        for st, color in [
            ("Present", "#16a34a"),
            ("Absent", "#dc2626"),
            ("Leave", "#2563eb"),
            ("Late", "#ea580c"),
        ]:
            self.tree_recent.tag_configure(st, foreground=color)
        try:
            self.tree_recent.tag_configure("even", background="#ffffff")
            self.tree_recent.tag_configure("odd", background="#f8fafc")
        except Exception:
            pass
        recent_scroll = ttk.Scrollbar(recent_wrap, orient=tk.VERTICAL, command=self.tree_recent.yview)
        self.tree_recent.configure(yscrollcommand=recent_scroll.set)
        recent_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree_recent.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self._refresh_recent()

    @staticmethod
    def _status_color(status):
        return {
            "Present": "#16a34a",
            "Absent": "#dc2626",
            "Leave": "#2563eb",
            "Late": "#ea580c",
        }.get(status, "#64748b")

    # ------------------------------------------------------------------
    def _render_summary(self):
        """Modern summary cards: Present | Absent | Leave | Late | Total."""
        for w in self.summary_frame.winfo_children():
            w.destroy()
        today = datetime.now().strftime("%Y-%m-%d")
        counts = {}
        for st in STATUS_VALUES:
            try:
                counts[st] = db.run(
                    "SELECT COUNT(*) FROM attendance WHERE date=? AND status=?",
                    (today, st), fetchone=True,
                )[0]
            except Exception:
                counts[st] = 0
        total = sum(counts.values())

        row = tk.Frame(self.summary_frame, bg="#F8FAFC")
        row.pack(fill=tk.X)

        # label, value, accent, icon glyph, soft icon bg
        items = [
            ("Present", counts.get("Present", 0), "#16a34a", "✓", "#dcfce7"),
            ("Absent", counts.get("Absent", 0), "#dc2626", "✕", "#fee2e2"),
            ("Leave", counts.get("Leave", 0), "#2563eb", "◷", "#dbeafe"),
            ("Late", counts.get("Late", 0), "#ea580c", "⏰", "#ffedd5"),
            ("Total", total, "#0f172a", "☰", "#e2e8f0"),
        ]
        for i, (label, val, accent, icon, icon_bg) in enumerate(items):
            card = tk.Frame(
                row, bg="#ffffff",
                highlightbackground="#e2e8f0", highlightthickness=1,
            )
            card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0 if i == 0 else 8, 0))
            stripe = tk.Frame(card, bg=accent, width=4)
            stripe.pack(side=tk.LEFT, fill=tk.Y)
            inner = tk.Frame(card, bg="#ffffff", padx=12, pady=12)
            inner.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

            top = tk.Frame(inner, bg="#ffffff")
            top.pack(fill=tk.X)
            icon_lbl = tk.Label(
                top, text=icon, font=("Segoe UI", 11, "bold"),
                bg=icon_bg, fg=accent, width=3, pady=2,
            )
            icon_lbl.pack(side=tk.LEFT, padx=(0, 8))
            tk.Label(
                top, text=label, font=("Segoe UI", 9),
                bg="#ffffff", fg="#64748b",
            ).pack(side=tk.LEFT, anchor="w")
            tk.Label(
                inner, text=str(val), font=("Segoe UI", 20, "bold"),
                bg="#ffffff", fg="#0f172a",
            ).pack(anchor="w", pady=(6, 0))

    def _show_toast(self, message, kind="info"):
        """Brief non-blocking banner at top of attendance window (scan feedback)."""
        colors = {
            "success": ("#166534", "#dcfce7", "#16a34a"),
            "error": ("#991b1b", "#fee2e2", "#dc2626"),
            "warning": ("#9a3412", "#ffedd5", "#ea580c"),
            "info": ("#1e40af", "#dbeafe", "#2563eb"),
        }
        fg, bg, border = colors.get(kind, colors["info"])
        try:
            if getattr(self, "_toast_frame", None) is not None:
                try:
                    self._toast_frame.destroy()
                except Exception:
                    pass
        except Exception:
            pass
        toast = tk.Frame(
            self.win, bg=bg,
            highlightbackground=border, highlightthickness=1,
        )
        toast.place(relx=0.5, y=8, anchor="n")
        tk.Label(
            toast, text=message, font=("Segoe UI", 10, "bold"),
            bg=bg, fg=fg, padx=16, pady=8,
        ).pack()
        self._toast_frame = toast

        def _hide():
            try:
                toast.destroy()
            except Exception:
                pass
            if getattr(self, "_toast_frame", None) is toast:
                self._toast_frame = None

        try:
            self.win.after(3200, _hide)
        except Exception:
            pass

    def _refresh_recent(self):
        """Reload today's attendance into the Recent Attendance table.

        Shows scan/mark time (in_time), student, class, method, and status.
        Newest records first (by id DESC). Safe if in_time column is missing
        on very old DBs — falls back to '—'.
        """
        self.tree_recent.delete(*self.tree_recent.get_children())
        today = datetime.now().strftime("%Y-%m-%d")
        try:
            rows = db.run(
                "SELECT a.id, a.student_id, s.name, s.class_sec, a.method, a.status, a.in_time "
                "FROM attendance a "
                "LEFT JOIN students s ON s.student_id = a.student_id "
                "WHERE a.date=? ORDER BY a.id DESC LIMIT 50",
                (today,), fetchall=True,
            ) or []
        except Exception:
            # Older schema without in_time
            rows = db.run(
                "SELECT a.id, a.student_id, s.name, s.class_sec, a.method, a.status "
                "FROM attendance a "
                "LEFT JOIN students s ON s.student_id = a.student_id "
                "WHERE a.date=? ORDER BY a.id DESC LIMIT 50",
                (today,), fetchall=True,
            ) or []
            rows = [tuple(list(r) + [None]) for r in rows]

        for row in rows:
            if len(row) >= 7:
                aid, sid, name, cls, method, status, in_time = row[:7]
            else:
                aid, sid, name, cls, method, status = row[:6]
                in_time = None
            time_str = (in_time or "").strip() or "—"
            tags = []
            if status in STATUS_VALUES:
                tags.append(status)
            # zebra after insert count
            tags.append("odd" if len(self.tree_recent.get_children()) % 2 else "even")
            self.tree_recent.insert(
                "", tk.END,
                values=(time_str, sid, name or "?", cls or "-", method or "-", status or "-"),
                tags=tuple(tags),
            )

    # ------------------------------------------------------------------
    # Barcode / scan
    # ------------------------------------------------------------------
    def process_scan(self):
        sid = self.ent_scan.get().strip()
        self.ent_scan.delete(0, tk.END)
        if not sid:
            return
        if not self.can_mark:
            return
        self._mark(sid, "Present", method="Barcode", scan_ui=True)

    # ------------------------------------------------------------------
    # Manual
    # ------------------------------------------------------------------
    def _manual_search(self):
        q = self.ent_manual_search.get().strip()
        self.tree_manual.delete(*self.tree_manual.get_children())
        if not q:
            return
        rows = db.run(
            "SELECT student_id, name, class_sec FROM students WHERE "
            "(student_id LIKE ? OR name LIKE ? OR class_sec LIKE ?) AND COALESCE(status,'Active')='Active'",
            (f"%{q}%", f"%{q}%", f"%{q}%"), fetchall=True)
        today = datetime.now().strftime("%Y-%m-%d")
        for idx, (sid, name, cls) in enumerate(rows):
            existing = db.run("SELECT status FROM attendance WHERE student_id=? AND date=?",
                               (sid, today), fetchone=True)
            st = existing[0] if existing else "—"
            tags = []
            if st in STATUS_VALUES:
                tags.append(st)
            tags.append("odd" if idx % 2 else "even")
            self.tree_manual.insert(
                "", tk.END,
                values=(sid, name, cls or "-", st),
                tags=tuple(tags),
            )

    def _manual_mark(self, status):
        sel = self.tree_manual.focus()
        if not sel:
            messagebox.showinfo("No Selection", "Select a student in the search results first.", parent=self.win)
            return
        sid = self.tree_manual.item(sel, "values")[0]
        self._mark(sid, status, method="Manual", scan_ui=False)
        self._manual_search()

    # ------------------------------------------------------------------
    def _mark(self, sid, status, method, scan_ui):
        if not self.can_mark:
            messagebox.showerror("Permission Denied", "You are not allowed to mark attendance.", parent=self.win)
            return
        student = db.run("SELECT name, class_sec, photo_path, status FROM students WHERE student_id=?",
                          (sid,), fetchone=True)
        if not student:
            if scan_ui:
                msg = f"✖ Student Not Found — '{sid}'"
                if self.lbl_scan_result is not None:
                    self.lbl_scan_result.config(text=msg, fg="#dc2626")
                self._show_toast(msg, "error")
            else:
                messagebox.showerror("Student Not Found", f"'{sid}' does not match any student.", parent=self.win)
            return
        name, cls, photo_path, active_status = student
        if (active_status or "Active") != "Active":
            msg = f"'{name}' is Archived and cannot be marked."
            if scan_ui:
                if self.lbl_scan_result is not None:
                    self.lbl_scan_result.config(text=f"✖ {msg}", fg="#dc2626")
                self._show_toast(f"✖ {msg}", "error")
            else:
                messagebox.showerror("Student Archived", msg, parent=self.win)
            return

        today = datetime.now().strftime("%Y-%m-%d")
        now_time = datetime.now().strftime("%H:%M")
        existing = db.run("SELECT status, in_time FROM attendance WHERE student_id=? AND date=?",
                           (sid, today), fetchone=True)
        if existing:
            ex_status, ex_time = existing
            msg = f"Already Marked — {name}'s attendance is already recorded today.\nFirst: {ex_time or '-'}  Status: {ex_status}"
            if scan_ui:
                short = f"⚠ Already marked — {name} ({ex_status})"
                if self.lbl_scan_result is not None:
                    self.lbl_scan_result.config(text=f"⚠ {msg}", fg="#ea580c")
                self._show_toast(short, "warning")
            else:
                messagebox.showinfo("Already Marked", msg, parent=self.win)
            return

        # Late detection — compare wall clock to late_threshold_time from system_settings
        if status == "Present" and method == "Barcode":
            settings = get_settings()
            try:
                late_cutoff = datetime.strptime(
                    settings.get("late_threshold_time") or "08:15", "%H:%M"
                ).time()
                if datetime.now().time() > late_cutoff:
                    status = "Late"
            except Exception:
                pass

        try:
            db.run("INSERT INTO attendance (student_id, date, status, method, in_time) VALUES (?, ?, ?, ?, ?)",
                   (sid, today, status, method, now_time), commit=True)
        except Exception:
            # UNIQUE(student_id, date) constraint hit — a duplicate slipped
            # in between our SELECT and INSERT (e.g. two scans milliseconds
            # apart). Treat exactly like "already marked", never a crash.
            existing2 = db.run("SELECT status, in_time FROM attendance WHERE student_id=? AND date=?",
                                (sid, today), fetchone=True)
            msg = f"Already Marked — {name}'s attendance was just recorded.\n{existing2}"
            if scan_ui:
                if self.lbl_scan_result is not None:
                    self.lbl_scan_result.config(text=f"⚠ {msg}", fg="#ea580c")
                self._show_toast(f"⚠ Already marked — {name}", "warning")
            else:
                messagebox.showinfo("Already Marked", msg, parent=self.win)
            return

        if scan_ui:
            msg = f"✓  {name}  ·  {sid}  ·  {status.upper()}  ·  {now_time}"
            if self.lbl_scan_result is not None:
                self.lbl_scan_result.config(text=msg, fg="#16a34a")
            self._show_toast(msg, "success")

        self._render_summary()
        self._refresh_recent()

    # ------------------------------------------------------------------
    # Automatic end-of-day absent
    # ------------------------------------------------------------------
    def _maybe_run_auto_absent_on_open(self):
        """When Attendance window opens after closing time, apply auto-absent
        if the background worker has not already done so today."""
        settings = get_settings()
        if not settings.get("auto_absent_enabled", True):
            return
        if not is_past_school_closing():
            return
        if auto_absent_already_run_today():
            return
        ran, n = try_auto_absent_now(force=False)
        if ran and n:
            messagebox.showinfo(
                "End-of-Day Absent Applied",
                f"School closing time ({settings['closing_time']}) has passed.\n"
                f"{n} active student(s) with no Present/Late/Leave record today were "
                f"automatically marked ABSENT (Auto System).",
                parent=self.win,
            )
            self._render_summary()
            self._refresh_recent()

    def _manual_run_auto_absent(self):
        if auto_absent_already_run_today():
            if not messagebox.askyesno("Already Run Today",
                                        "The end-of-day Absent sweep already ran today. Run it again?\n"
                                        "(It only affects students still missing a record — no duplicates "
                                        "will be created.)", parent=self.win):
                return
        n = run_auto_absent(force=True)
        messagebox.showinfo("Done", f"{n} student(s) marked Absent.", parent=self.win)
        self._render_summary()
        self._refresh_recent()

    # ------------------------------------------------------------------
    # Reports
    # ------------------------------------------------------------------
    # ------------------------------------------------------------------
    # Attendance analytics helpers
    # ------------------------------------------------------------------
    @staticmethod
    def _counts_for_student(student_id, year=None, month=None):
        q = "SELECT status, COUNT(*) FROM attendance WHERE student_id=?"
        params = [student_id]
        if year and month:
            q += " AND date LIKE ?"
            params.append(f"{year}-{month:02d}%")
        elif year:
            q += " AND date LIKE ?"
            params.append(f"{year}%")
        q += " GROUP BY status"
        rows = db.run(q, tuple(params), fetchall=True) or []
        counts = {"Present": 0, "Absent": 0, "Leave": 0, "Late": 0}
        for st, c in rows:
            if st in counts:
                counts[st] = int(c or 0)
        total = sum(counts.values())
        rate = ((counts["Present"] + counts["Late"]) / total * 100.0) if total else 0.0
        counts["Total"] = total
        counts["Rate"] = rate
        return counts

    @staticmethod
    def _day_rows_for_student(student_id, year=None, month=None, limit=250):
        try:
            q = (
                "SELECT date, status, COALESCE(method,''), COALESCE(in_time,'') "
                "FROM attendance WHERE student_id=?"
            )
            params = [student_id]
            if year and month:
                q += " AND date LIKE ?"
                params.append(f"{year}-{month:02d}%")
            elif year:
                q += " AND date LIKE ?"
                params.append(f"{year}%")
            q += " ORDER BY date DESC LIMIT ?"
            params.append(limit)
            return db.run(q, tuple(params), fetchall=True) or []
        except Exception:
            q = "SELECT date, status, COALESCE(method,''), '' FROM attendance WHERE student_id=?"
            params = [student_id]
            if year and month:
                q += " AND date LIKE ?"
                params.append(f"{year}-{month:02d}%")
            elif year:
                q += " AND date LIKE ?"
                params.append(f"{year}%")
            q += " ORDER BY date DESC LIMIT ?"
            params.append(limit)
            return db.run(q, tuple(params), fetchall=True) or []

    # ------------------------------------------------------------------
    # Personal Attendance (monthly + yearly) + Excel/PDF export
    # ------------------------------------------------------------------
    def _open_personal_attendance(self):
        if not rbac.can(self.user_role, "attendance.view") and not self.can_mark:
            messagebox.showerror("Permission Denied", "You cannot view attendance reports.", parent=self.win)
            return

        win = tk.Toplevel(self.win)
        win.title("Personal Attendance — Monthly & Yearly")
        win.geometry("940x720")
        win.config(bg=theme.SILVER)
        win.transient(self.win)

        header = tk.Frame(win, bg=theme.NAVY, padx=18, pady=12)
        header.pack(fill=tk.X)
        tk.Label(header, text="👤  PERSONAL ATTENDANCE", font=theme.FONT_H1,
                 bg=theme.NAVY, fg="white").pack(anchor="w")
        tk.Label(header, text="Search student → Monthly & Yearly summary · Export Excel / PDF",
                 font=theme.FONT_SMALL, bg=theme.NAVY, fg="#94a3b8").pack(anchor="w")

        body = tk.Frame(win, bg=theme.SILVER, padx=14, pady=12)
        body.pack(fill=tk.BOTH, expand=True)

        search_row = tk.Frame(body, bg=theme.WHITE, padx=12, pady=10,
                              highlightbackground=theme.SILVER_BORDER, highlightthickness=1)
        search_row.pack(fill=tk.X, pady=(0, 10))
        tk.Label(search_row, text="Student ID / Name:", font=theme.FONT_BODY_BOLD,
                 bg=theme.WHITE).pack(side=tk.LEFT)
        ent_q = tk.Entry(search_row, font=theme.FONT_BODY, width=20)
        ent_q.pack(side=tk.LEFT, padx=8, ipady=3)
        now = datetime.now()
        year_var = tk.StringVar(value=str(now.year))
        month_var = tk.StringVar(value=str(now.month))
        tk.Label(search_row, text="Year:", font=theme.FONT_SMALL, bg=theme.WHITE).pack(side=tk.LEFT, padx=(10, 2))
        cmb_year = ttk.Combobox(search_row, textvariable=year_var,
                                 values=[str(y) for y in range(now.year, now.year - 6, -1)],
                                 width=7, state="readonly")
        cmb_year.pack(side=tk.LEFT)
        tk.Label(search_row, text="Month:", font=theme.FONT_SMALL, bg=theme.WHITE).pack(side=tk.LEFT, padx=(8, 2))
        cmb_month = ttk.Combobox(search_row, textvariable=month_var,
                                  values=[str(i) for i in range(1, 13)], width=5, state="readonly")
        cmb_month.pack(side=tk.LEFT)

        lbl_student = tk.Label(body, text="Search a student to load personal attendance.",
                                font=theme.FONT_H2, bg=theme.SILVER, fg=theme.TEXT_MUTED)
        lbl_student.pack(anchor="w", pady=(0, 6))

        cards = tk.Frame(body, bg=theme.SILVER)
        cards.pack(fill=tk.X, pady=(0, 8))

        export_row = tk.Frame(body, bg=theme.SILVER)
        export_row.pack(fill=tk.X, pady=(0, 8))

        tk.Label(body, text="📅 Monthly Breakdown (selected year)", font=theme.FONT_BODY_BOLD,
                 bg=theme.SILVER).pack(anchor="w")
        month_tree = ttk.Treeview(
            body, columns=("month", "present", "absent", "leave", "late", "total", "rate"),
            show="headings", height=5,
        )
        for c, h, w in [
            ("month", "Month", 100), ("present", "Present", 75), ("absent", "Absent", 75),
            ("leave", "Leave", 65), ("late", "Late", 65), ("total", "Total", 70),
            ("rate", "Rate %", 80),
        ]:
            month_tree.heading(c, text=h)
            month_tree.column(c, width=w, anchor="center")
        month_tree.pack(fill=tk.X, pady=(2, 8))

        tk.Label(body, text="📋 Day-wise History (selected month)", font=theme.FONT_BODY_BOLD,
                 bg=theme.SILVER).pack(anchor="w")
        day_tree = ttk.Treeview(
            body, columns=("date", "status", "method", "time"), show="headings", height=9,
        )
        for c, h, w in [("date", "Date", 110), ("status", "Status", 90),
                        ("method", "Method", 110), ("time", "In Time", 90)]:
            day_tree.heading(c, text=h)
            day_tree.column(c, width=w, anchor="center")
        for st, color in [("Present", theme.SUCCESS), ("Absent", theme.DANGER),
                          ("Leave", theme.INFO), ("Late", theme.WARNING)]:
            day_tree.tag_configure(st, foreground=color)
        day_scroll = ttk.Scrollbar(body, orient=tk.VERTICAL, command=day_tree.yview)
        day_tree.configure(yscrollcommand=day_scroll.set)
        day_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        day_tree.pack(fill=tk.BOTH, expand=True, pady=(2, 0))

        state = {"sid": None, "name": "", "class": "", "month_counts": {}, "year_counts": {},
                 "day_rows": [], "monthly_rows": []}

        def _card(parent, title, counts):
            box = tk.Frame(parent, bg=theme.WHITE, highlightbackground=theme.SILVER_BORDER,
                           highlightthickness=1, padx=12, pady=8)
            box.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 8))
            tk.Label(box, text=title, font=theme.FONT_SMALL, bg=theme.WHITE,
                     fg=theme.TEXT_MUTED).pack(anchor="w")
            rate_fg = theme.SUCCESS if counts["Rate"] >= 80 else (
                theme.WARNING if counts["Rate"] >= 60 else theme.DANGER)
            tk.Label(box, text=f"{counts['Rate']:.1f}%", font=("Segoe UI", 18, "bold"),
                     bg=theme.WHITE, fg=rate_fg).pack(anchor="w")
            tk.Label(
                box,
                text=(f"P {counts['Present']} · A {counts['Absent']} · "
                      f"L {counts['Leave']} · Late {counts['Late']} · Total {counts['Total']}"),
                font=theme.FONT_SMALL, bg=theme.WHITE, fg=theme.TEXT_MUTED,
            ).pack(anchor="w")

        def load_for(sid, name, cls):
            state["sid"] = sid
            state["name"] = name
            state["class"] = cls or ""
            lbl_student.config(
                text=f"{name}  |  ID: {sid}  |  Class: {cls or '-'}",
                fg=theme.TEXT_DARK,
            )
            try:
                y = int(year_var.get())
                m = int(month_var.get())
            except ValueError:
                y, m = now.year, now.month

            for w in cards.winfo_children():
                w.destroy()
            mc = self._counts_for_student(sid, year=y, month=m)
            yc = self._counts_for_student(sid, year=y)
            state["month_counts"] = mc
            state["year_counts"] = yc
            _card(cards, f"This Month — {MONTH_NAMES_FULL[m - 1]} {y}", mc)
            _card(cards, f"Yearly — {y}", yc)

            month_tree.delete(*month_tree.get_children())
            monthly_rows = []
            for mi in range(1, 13):
                c = self._counts_for_student(sid, year=y, month=mi)
                if c["Total"] == 0 and mi > m and y == now.year:
                    continue
                row = (MONTH_NAMES_FULL[mi - 1], c["Present"], c["Absent"],
                       c["Leave"], c["Late"], c["Total"], f"{c['Rate']:.1f}%")
                monthly_rows.append(row)
                month_tree.insert("", tk.END, values=row)
            state["monthly_rows"] = monthly_rows

            day_tree.delete(*day_tree.get_children())
            rows = self._day_rows_for_student(sid, year=y, month=m)
            state["day_rows"] = rows
            if not rows:
                day_tree.insert("", tk.END, values=("No records", "—", "—", "—"))
            else:
                for d, st, method, in_time in rows:
                    tag = st if st in STATUS_VALUES else ""
                    day_tree.insert(
                        "", tk.END,
                        values=(d, st, method or "—", (in_time or "").strip() or "—"),
                        tags=(tag,) if tag else (),
                    )

        def do_search(_ev=None):
            q = ent_q.get().strip()
            if not q:
                messagebox.showinfo("Search", "Enter Student ID or Name.", parent=win)
                return
            rows = db.run(
                "SELECT student_id, name, class_sec FROM students WHERE "
                "(student_id LIKE ? OR name LIKE ?) AND COALESCE(status,'Active')='Active' "
                "ORDER BY name LIMIT 20",
                (f"%{q}%", f"%{q}%"), fetchall=True,
            ) or []
            if not rows:
                messagebox.showinfo("Not Found", f"No active student matched '{q}'.", parent=win)
                return
            if len(rows) == 1:
                load_for(rows[0][0], rows[0][1], rows[0][2])
                return
            pick = tk.Toplevel(win)
            pick.title("Select Student")
            pick.geometry("420x320")
            pick.config(bg=theme.WHITE)
            pick.transient(win)
            tk.Label(pick, text="Multiple matches — select one:", font=theme.FONT_BODY_BOLD,
                     bg=theme.WHITE).pack(anchor="w", padx=12, pady=8)
            tv = ttk.Treeview(pick, columns=("id", "name", "class"), show="headings", height=10)
            for c, h, w in [("id", "ID", 110), ("name", "Name", 160), ("class", "Class", 90)]:
                tv.heading(c, text=h)
                tv.column(c, width=w, anchor="center")
            tv.pack(fill=tk.BOTH, expand=True, padx=12, pady=4)
            for r in rows:
                tv.insert("", tk.END, values=r)

            def choose():
                sel = tv.focus()
                if not sel:
                    return
                vals = tv.item(sel, "values")
                load_for(vals[0], vals[1], vals[2])
                pick.destroy()

            theme.primary_button(pick, "Open", choose).pack(pady=8)
            tv.bind("<Double-1>", lambda e: choose())

        def refresh_current(_ev=None):
            if state["sid"]:
                row = db.run(
                    "SELECT student_id, name, class_sec FROM students WHERE student_id=?",
                    (state["sid"],), fetchone=True,
                )
                if row:
                    load_for(row[0], row[1], row[2])

        def do_export_excel():
            if not state["sid"]:
                messagebox.showinfo("No Data", "Search and load a student first.", parent=win)
                return
            try:
                y = int(year_var.get())
                m = int(month_var.get())
            except ValueError:
                y, m = now.year, now.month
            default = f"Personal_Attendance_{state['sid']}_{y}-{m:02d}.xlsx"
            path = _ask_save_path(win, default, "xlsx")
            if not path:
                return
            try:
                export_personal_attendance_excel(
                    state["sid"], state["name"], state["class"], y, m,
                    state["month_counts"], state["year_counts"],
                    state["day_rows"], state["monthly_rows"], path,
                )
                messagebox.showinfo("Export Complete", f"Excel saved:\n{path}", parent=win)
            except Exception as exc:
                messagebox.showerror("Export Failed", str(exc), parent=win)

        def do_export_pdf():
            if not state["sid"]:
                messagebox.showinfo("No Data", "Search and load a student first.", parent=win)
                return
            try:
                y = int(year_var.get())
                m = int(month_var.get())
            except ValueError:
                y, m = now.year, now.month
            default = f"Personal_Attendance_{state['sid']}_{y}-{m:02d}.pdf"
            path = _ask_save_path(win, default, "pdf")
            if not path:
                return
            try:
                export_personal_attendance_pdf(
                    state["sid"], state["name"], state["class"], y, m,
                    state["month_counts"], state["year_counts"],
                    state["day_rows"], path,
                )
                messagebox.showinfo("Export Complete", f"PDF saved:\n{path}", parent=win)
            except Exception as exc:
                messagebox.showerror("Export Failed", str(exc), parent=win)

        theme.primary_button(search_row, "🔍 Search", do_search).pack(side=tk.LEFT, padx=8)
        theme.primary_button(search_row, "↻ Refresh", refresh_current, bg=theme.SLATE).pack(side=tk.LEFT)
        theme.primary_button(export_row, "📁 Export Excel", do_export_excel, bg=theme.SUCCESS).pack(side=tk.LEFT, padx=(0, 8))
        theme.primary_button(export_row, "📄 Export PDF", do_export_pdf, bg=theme.SLATE).pack(side=tk.LEFT)
        ent_q.bind("<Return>", do_search)
        cmb_year.bind("<<ComboboxSelected>>", refresh_current)
        cmb_month.bind("<<ComboboxSelected>>", refresh_current)
        ent_q.focus_set()

    # ------------------------------------------------------------------
    # Class-wise Attendance (monthly + yearly) + Excel/PDF export
    # ------------------------------------------------------------------

    def _open_class_attendance(self):
        """Merged into Attendance Review — single place for class / status / gender filters."""
        self._open_attendance_review()


    # ------------------------------------------------------------------
    # Attendance Review — Class / Status / Gender / Consecutive days
    # ------------------------------------------------------------------
    def _open_attendance_review(self):
        """Unified attendance review with Class, Status, Gender and consecutive-absent filters."""
        if not rbac.can(self.user_role, "attendance.view") and not self.can_mark:
            messagebox.showerror("Permission Denied", "You cannot view attendance.", parent=self.win)
            return

        win = tk.Toplevel(self.win)
        win.title("Review Attendance")
        win.geometry("1080x700")
        win.config(bg=theme.SILVER)
        win.transient(self.win)
        try:
            win.lift()
            win.focus_force()
        except Exception:
            pass

        header = tk.Frame(win, bg=theme.NAVY, padx=18, pady=12)
        header.pack(fill=tk.X)
        tk.Label(header, text="📋  REVIEW ATTENDANCE", font=theme.FONT_H1,
                 bg=theme.NAVY, fg="white").pack(anchor="w")
        tk.Label(
            header,
            text="Class-wise + filters · Status · Gender · Date · Consecutive Absent · one place to review",
            font=theme.FONT_SMALL, bg=theme.NAVY, fg="#94a3b8",
        ).pack(anchor="w")

        body = tk.Frame(win, bg=theme.SILVER, padx=14, pady=12)
        body.pack(fill=tk.BOTH, expand=True)

        now = datetime.now()
        classes = ["All Classes"] + [
            r[0] for r in (
                db.run(
                    "SELECT DISTINCT class_sec FROM students "
                    "WHERE class_sec IS NOT NULL AND TRIM(class_sec)<>'' "
                    "AND COALESCE(status,'Active')='Active' ORDER BY class_sec",
                    fetchall=True,
                ) or []
            )
        ]

        ctrl = tk.Frame(body, bg=theme.WHITE, padx=12, pady=10,
                        highlightbackground=theme.SILVER_BORDER, highlightthickness=1)
        ctrl.pack(fill=tk.X, pady=(0, 8))

        tk.Label(ctrl, text="Class:", font=theme.FONT_BODY_BOLD, bg=theme.WHITE).pack(side=tk.LEFT)
        class_var = tk.StringVar(value="All Classes")
        cmb_class = ttk.Combobox(ctrl, textvariable=class_var, values=classes, width=12, state="readonly")
        cmb_class.pack(side=tk.LEFT, padx=(3, 8))

        tk.Label(ctrl, text="Status:", font=theme.FONT_BODY_BOLD, bg=theme.WHITE).pack(side=tk.LEFT)
        status_var = tk.StringVar(value="All")
        cmb_status = ttk.Combobox(
            ctrl, textvariable=status_var,
            values=["All", "Present", "Absent", "Leave", "Late"], width=9, state="readonly",
        )
        cmb_status.pack(side=tk.LEFT, padx=(3, 8))

        tk.Label(ctrl, text="Gender:", font=theme.FONT_BODY_BOLD, bg=theme.WHITE).pack(side=tk.LEFT)
        gender_var = tk.StringVar(value="All")
        cmb_gender = ttk.Combobox(
            ctrl, textvariable=gender_var,
            values=["All", "Male", "Female", "Other"], width=8, state="readonly",
        )
        cmb_gender.pack(side=tk.LEFT, padx=(3, 8))

        tk.Label(ctrl, text="Date:", font=theme.FONT_BODY_BOLD, bg=theme.WHITE).pack(side=tk.LEFT)
        date_var = tk.StringVar(value=now.strftime("%Y-%m-%d"))
        ent_date = tk.Entry(ctrl, textvariable=date_var, width=11, font=theme.FONT_BODY)
        ent_date.pack(side=tk.LEFT, padx=(3, 8))

        tk.Label(ctrl, text="Absent ≥ days:", font=theme.FONT_BODY_BOLD, bg=theme.WHITE).pack(side=tk.LEFT)
        min_var = tk.StringVar(value="0")
        cmb_min = ttk.Combobox(
            ctrl, textvariable=min_var,
            values=["0", "2", "3", "4", "5", "7", "10"], width=4, state="readonly",
        )
        cmb_min.pack(side=tk.LEFT, padx=(3, 8))

        theme.primary_button(ctrl, "↻ Review", lambda: refresh()).pack(side=tk.LEFT, padx=4)

        info_lbl = tk.Label(body, text="", font=theme.FONT_BODY, bg=theme.SILVER, fg=theme.TEXT_MUTED)
        info_lbl.pack(anchor="w", pady=(0, 6))

        cols = ("id", "name", "class", "gender", "date", "status", "method", "time", "streak")
        tree = ttk.Treeview(body, columns=cols, show="headings", height=18)
        headers = {
            "id": ("Student ID", 100), "name": ("Name", 140), "class": ("Class", 90),
            "gender": ("Gender", 70), "date": ("Date", 95), "status": ("Status", 80),
            "method": ("Method", 90), "time": ("In Time", 70), "streak": ("Absent Streak", 100),
        }
        for c in cols:
            h, w = headers[c]
            tree.heading(c, text=h)
            tree.column(c, width=w, anchor="center")
        for st, color in [
            ("Present", theme.SUCCESS), ("Absent", theme.DANGER),
            ("Leave", theme.INFO), ("Late", theme.WARNING),
        ]:
            tree.tag_configure(st, foreground=color)
        tree.tag_configure("streak_high", foreground=theme.DANGER)
        scroll = ttk.Scrollbar(body, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        tree.pack(fill=tk.BOTH, expand=True)

        quick = tk.Frame(body, bg=theme.SILVER)
        quick.pack(fill=tk.X, pady=(8, 0))
        theme.primary_button(
            quick, "⚠ 3+ Days Absent",
            lambda: (min_var.set("3"), refresh()),
            bg=theme.WARNING,
        ).pack(side=tk.LEFT, padx=(0, 8))
        theme.primary_button(
            quick, "⚠ 5+ Days Absent",
            lambda: (min_var.set("5"), refresh()),
            bg=theme.DANGER,
        ).pack(side=tk.LEFT, padx=(0, 8))
        theme.primary_button(
            quick, "Today's Attendance",
            lambda: (min_var.set("0"), date_var.set(now.strftime("%Y-%m-%d")), refresh()),
            bg=theme.SLATE,
        ).pack(side=tk.LEFT)

        def _streak(sid):
            rows = db.run(
                "SELECT date, status FROM attendance WHERE student_id=? ORDER BY date DESC LIMIT 45",
                (sid,), fetchall=True,
            ) or []
            n = 0
            for _d, st in rows:
                if st == "Absent":
                    n += 1
                else:
                    break
            return n

        def refresh(_ev=None):
            cls = class_var.get() or "All Classes"
            st_filter = status_var.get() or "All"
            gdr = gender_var.get() or "All"
            day = (date_var.get() or "").strip()
            try:
                min_days = int(min_var.get())
            except ValueError:
                min_days = 0

            if day:
                try:
                    datetime.strptime(day, "%Y-%m-%d")
                except ValueError:
                    messagebox.showerror("Invalid Date", "Date must be YYYY-MM-DD.", parent=win)
                    return

            if min_days >= 2:
                q = (
                    "SELECT s.student_id, s.name, COALESCE(s.class_sec,''), "
                    "COALESCE(e.gender,'') FROM students s "
                    "LEFT JOIN student_admission_extra e ON e.student_id=s.student_id "
                    "WHERE COALESCE(s.status,'Active')='Active'"
                )
                params = []
                if cls != "All Classes":
                    q += " AND s.class_sec=?"
                    params.append(cls)
                if gdr != "All":
                    q += " AND LOWER(COALESCE(e.gender,'')) LIKE ?"
                    params.append(f"%{gdr.lower()}%")
                q += " ORDER BY s.name"
                students = db.run(q, tuple(params), fetchall=True) or []

                tree.delete(*tree.get_children())
                count = 0
                for sid, name, cls_sec, gender in students:
                    streak = _streak(sid)
                    if streak < min_days:
                        continue
                    latest = db.run(
                        "SELECT date, status, COALESCE(method,''), COALESCE(in_time,'') "
                        "FROM attendance WHERE student_id=? ORDER BY date DESC LIMIT 1",
                        (sid,), fetchone=True,
                    )
                    if latest:
                        d, st, method, in_time = latest
                    else:
                        d, st, method, in_time = "—", "Absent", "—", "—"
                    if st_filter != "All" and st != st_filter:
                        continue
                    tag = "streak_high" if streak >= 5 else (st if st in STATUS_VALUES else "")
                    tree.insert(
                        "", tk.END,
                        values=(
                            sid, name, cls_sec or "-", gender or "-",
                            d, st, method or "—", (in_time or "").strip() or "—", f"{streak} days",
                        ),
                        tags=(tag,) if tag else (),
                    )
                    count += 1
                info_lbl.config(
                    text=f"⚠ {count} student(s) with {min_days}+ consecutive Absent days"
                         + (f" · Class: {cls}" if cls != "All Classes" else "")
                         + (f" · Gender: {gdr}" if gdr != "All" else ""),
                    fg=theme.DANGER if count else theme.SUCCESS,
                )
                try:
                    win.title(f"Review Attendance · {count} with {min_days}+ absent streak")
                except Exception:
                    pass
                return

            q = (
                "SELECT a.student_id, s.name, COALESCE(s.class_sec,''), COALESCE(e.gender,''), "
                "a.date, a.status, COALESCE(a.method,''), COALESCE(a.in_time,'') "
                "FROM attendance a "
                "JOIN students s ON s.student_id=a.student_id "
                "LEFT JOIN student_admission_extra e ON e.student_id=s.student_id "
                "WHERE COALESCE(s.status,'Active')='Active'"
            )
            params = []
            if day:
                q += " AND a.date=?"
                params.append(day)
            if cls != "All Classes":
                q += " AND s.class_sec=?"
                params.append(cls)
            if st_filter != "All":
                q += " AND a.status=?"
                params.append(st_filter)
            if gdr != "All":
                q += " AND LOWER(COALESCE(e.gender,'')) LIKE ?"
                params.append(f"%{gdr.lower()}%")
            q += " ORDER BY s.name, a.date DESC LIMIT 500"

            try:
                rows = db.run(q, tuple(params), fetchall=True) or []
            except Exception:
                q2 = q.replace("COALESCE(a.in_time,'')", "''")
                rows = db.run(q2, tuple(params), fetchall=True) or []

            tree.delete(*tree.get_children())
            for sid, name, cls_sec, gender, d, st, method, in_time in rows:
                tag = st if st in STATUS_VALUES else ""
                tree.insert(
                    "", tk.END,
                    values=(
                        sid, name, cls_sec or "-", gender or "-",
                        d, st, method or "—", (in_time or "").strip() or "—", "—",
                    ),
                    tags=(tag,) if tag else (),
                )

            parts = []
            if day:
                parts.append(f"Date {day}")
            if cls != "All Classes":
                parts.append(cls)
            if st_filter != "All":
                parts.append(st_filter)
            if gdr != "All":
                parts.append(gdr)
            filt = " · ".join(parts) if parts else "All"
            info_lbl.config(text=f"Showing {len(rows)} record(s) — {filt}", fg=theme.TEXT_DARK)
            try:
                win.title(f"Review Attendance · {len(rows)} records")
            except Exception:
                pass

        for w in (cmb_class, cmb_status, cmb_gender, cmb_min):
            w.bind("<<ComboboxSelected>>", refresh)
        ent_date.bind("<Return>", refresh)
        refresh()


    def _class_breakdown(self, start_date, end_date):
        rows = db.run(
            "SELECT s.class_sec, a.status, COUNT(*) FROM attendance a "
            "JOIN students s ON s.student_id=a.student_id WHERE a.date BETWEEN ? AND ? "
            "GROUP BY s.class_sec, a.status", (start_date, end_date), fetchall=True)
        return rows


    def _open_report_dialog(self):
        self._open_unified_reports()

    def _open_range_dialog(self):
        self._open_unified_reports()

    def _open_unified_reports(self):
        """Monthly + Custom-range reports in one window, with calendar + PDF/Excel."""
        if not rbac.can(self.user_role, "attendance.view") and not self.can_mark:
            messagebox.showerror("Permission Denied", "You cannot view attendance reports.", parent=self.win)
            return

        win = tk.Toplevel(self.win)
        win.title("Attendance Reports")
        win.geometry("720x560")
        win.config(bg=theme.SILVER)
        win.transient(self.win)
        try:
            win.lift()
            win.focus_force()
        except Exception:
            pass

        header = tk.Frame(win, bg=theme.NAVY, padx=18, pady=12)
        header.pack(fill=tk.X)
        tk.Label(header, text="📊  ATTENDANCE REPORTS", font=theme.FONT_H1,
                 bg=theme.NAVY, fg="white").pack(anchor="w")
        tk.Label(
            header,
            text="Monthly or Custom range · Class filter · Calendar date pick · Excel / PDF",
            font=theme.FONT_SMALL, bg=theme.NAVY, fg="#94a3b8",
        ).pack(anchor="w")

        body = tk.Frame(win, bg=theme.SILVER, padx=16, pady=12)
        body.pack(fill=tk.BOTH, expand=True)

        now = datetime.now()
        classes = ["All Classes"] + [
            r[0] for r in (
                db.run(
                    "SELECT DISTINCT class_sec FROM students "
                    "WHERE class_sec IS NOT NULL AND TRIM(class_sec)<>'' "
                    "AND COALESCE(status,'Active')='Active' ORDER BY class_sec",
                    fetchall=True,
                ) or []
            )
        ]

        form = tk.Frame(body, bg=theme.WHITE, padx=14, pady=12,
                        highlightbackground=theme.SILVER_BORDER, highlightthickness=1)
        form.pack(fill=tk.X, pady=(0, 10))

        tk.Label(form, text="Report Type:", font=theme.FONT_BODY_BOLD, bg=theme.WHITE).grid(
            row=0, column=0, sticky="e", padx=6, pady=6)
        type_var = tk.StringVar(value="Monthly")
        cmb_type = ttk.Combobox(
            form, textvariable=type_var, values=["Monthly", "Custom Range"],
            width=14, state="readonly",
        )
        cmb_type.grid(row=0, column=1, sticky="w", pady=6)

        tk.Label(form, text="Class:", font=theme.FONT_BODY_BOLD, bg=theme.WHITE).grid(
            row=0, column=2, sticky="e", padx=6, pady=6)
        class_var = tk.StringVar(value="All Classes")
        cmb_class = ttk.Combobox(form, textvariable=class_var, values=classes, width=14, state="readonly")
        cmb_class.grid(row=0, column=3, sticky="w", pady=6)

        # Monthly controls
        mon_frame = tk.Frame(form, bg=theme.WHITE)
        mon_frame.grid(row=1, column=0, columnspan=4, sticky="w", pady=4)
        tk.Label(mon_frame, text="Year:", font=theme.FONT_BODY, bg=theme.WHITE).pack(side=tk.LEFT)
        year_var = tk.StringVar(value=str(now.year))
        cmb_year = ttk.Combobox(
            mon_frame, textvariable=year_var,
            values=[str(y) for y in range(now.year, now.year - 6, -1)],
            width=7, state="readonly",
        )
        cmb_year.pack(side=tk.LEFT, padx=(4, 12))
        tk.Label(mon_frame, text="Month:", font=theme.FONT_BODY, bg=theme.WHITE).pack(side=tk.LEFT)
        month_var = tk.StringVar(value=str(now.month))
        cmb_month = ttk.Combobox(
            mon_frame, textvariable=month_var,
            values=[str(i) for i in range(1, 13)], width=5, state="readonly",
        )
        cmb_month.pack(side=tk.LEFT, padx=4)

        # Custom range controls with calendar
        range_frame = tk.Frame(form, bg=theme.WHITE)
        range_frame.grid(row=2, column=0, columnspan=4, sticky="w", pady=4)
        tk.Label(range_frame, text="From:", font=theme.FONT_BODY, bg=theme.WHITE).pack(side=tk.LEFT)
        from_var = tk.StringVar(value=now.strftime("%Y-%m-01"))
        ent_from = tk.Entry(range_frame, textvariable=from_var, width=12, font=theme.FONT_BODY)
        ent_from.pack(side=tk.LEFT, padx=4)

        def pick_from():
            d = _pick_date(win, from_var.get(), "From Date")
            if d:
                from_var.set(d)

        theme.primary_button(range_frame, "📅", pick_from, bg=theme.SLATE).pack(side=tk.LEFT, padx=(0, 12))
        tk.Label(range_frame, text="To:", font=theme.FONT_BODY, bg=theme.WHITE).pack(side=tk.LEFT)
        to_var = tk.StringVar(value=now.strftime("%Y-%m-%d"))
        ent_to = tk.Entry(range_frame, textvariable=to_var, width=12, font=theme.FONT_BODY)
        ent_to.pack(side=tk.LEFT, padx=4)

        def pick_to():
            d = _pick_date(win, to_var.get(), "To Date")
            if d:
                to_var.set(d)

        theme.primary_button(range_frame, "📅", pick_to, bg=theme.SLATE).pack(side=tk.LEFT)

        def sync_type(_ev=None):
            if type_var.get() == "Monthly":
                for w in (cmb_year, cmb_month):
                    w.configure(state="readonly")
                for w in (ent_from, ent_to):
                    w.configure(state="disabled")
            else:
                for w in (cmb_year, cmb_month):
                    w.configure(state="disabled")
                for w in (ent_from, ent_to):
                    w.configure(state="normal")

        cmb_type.bind("<<ComboboxSelected>>", sync_type)
        sync_type()

        result_box = tk.Frame(body, bg=theme.WHITE, padx=14, pady=12,
                              highlightbackground=theme.SILVER_BORDER, highlightthickness=1)
        result_box.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        result_lbl = tk.Label(
            result_box, text="Select options and click Generate.",
            font=theme.FONT_BODY, bg=theme.WHITE, justify="left", anchor="nw",
        )
        result_lbl.pack(fill=tk.BOTH, expand=True)

        state = {"start": None, "end": None, "counts": None, "cls": "All Classes", "label": ""}

        def _resolve_range():
            cls = class_var.get() or "All Classes"
            if type_var.get() == "Monthly":
                try:
                    y = int(year_var.get())
                    m = int(month_var.get())
                except ValueError:
                    messagebox.showerror("Invalid", "Select valid year/month.", parent=win)
                    return None
                start = f"{y:04d}-{m:02d}-01"
                # last day of month
                import calendar as _cal
                last = _cal.monthrange(y, m)[1]
                end = f"{y:04d}-{m:02d}-{last:02d}"
                label = f"{MONTH_NAMES_FULL[m - 1]} {y}"
            else:
                start, end = from_var.get().strip(), to_var.get().strip()
                for d in (start, end):
                    try:
                        datetime.strptime(d, "%Y-%m-%d")
                    except ValueError:
                        messagebox.showerror("Invalid Date", "Use calendar or YYYY-MM-DD.", parent=win)
                        return None
                if start > end:
                    messagebox.showerror("Invalid Range", "From date must be before To date.", parent=win)
                    return None
                label = f"{start} → {end}"
            return start, end, cls, label

        def generate():
            resolved = _resolve_range()
            if not resolved:
                return
            start, end, cls, label = resolved
            q = (
                "SELECT a.status, COUNT(*) FROM attendance a "
                "JOIN students s ON s.student_id=a.student_id "
                "WHERE a.date BETWEEN ? AND ? AND COALESCE(s.status,'Active')='Active'"
            )
            params = [start, end]
            if cls != "All Classes":
                q += " AND s.class_sec=?"
                params.append(cls)
            q += " GROUP BY a.status"
            rows = db.run(q, tuple(params), fetchall=True) or []
            counts = {st: 0 for st in STATUS_VALUES}
            for st, c in rows:
                if st in counts:
                    counts[st] = int(c or 0)
            total = sum(counts.values())
            present_like = counts["Present"] + counts["Late"]
            rate = (present_like / total * 100.0) if total else 0.0
            state.update({"start": start, "end": end, "counts": counts, "cls": cls, "label": label, "total": total, "rate": rate})
            result_lbl.config(text=(
                f"Period: {label}\n"
                f"Class: {cls}\n\n"
                f"Present : {counts['Present']}\n"
                f"Absent  : {counts['Absent']}\n"
                f"Leave   : {counts['Leave']}\n"
                f"Late    : {counts['Late']}\n"
                f"Total   : {total}\n"
                f"Attendance Rate: {rate:.1f}%"
            ))

        def export_excel():
            if not state.get("counts"):
                messagebox.showinfo("No Data", "Generate the report first.", parent=win)
                return
            safe = (state["cls"] or "All").replace(" ", "_").replace("/", "-")
            default = f"Attendance_Report_{safe}_{state['start']}_{state['end']}.xlsx"
            path = _ask_save_path(win, default, "xlsx")
            if not path:
                return
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Border, Side
                wb = Workbook()
                ws = wb.active
                ws.title = "Report"
                header_fill = PatternFill("solid", fgColor="0F172A")
                sub_fill = PatternFill("solid", fgColor="0284C7")
                thin = Border(
                    left=Side(style="thin", color="E2E8F0"),
                    right=Side(style="thin", color="E2E8F0"),
                    top=Side(style="thin", color="E2E8F0"),
                    bottom=Side(style="thin", color="E2E8F0"),
                )
                ws.merge_cells("A1:B1")
                ws["A1"].value = "Attendance Report — AR School Management System"
                ws["A1"].font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=12)
                ws["A1"].fill = header_fill
                ws["A3"] = "Period"
                ws["B3"] = state["label"]
                ws["A4"] = "Class"
                ws["B4"] = state["cls"]
                ws["A5"] = "Generated"
                ws["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M")
                for col, h in enumerate(["Status", "Count"], 1):
                    cell = ws.cell(row=7, column=col, value=h)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = sub_fill
                    cell.border = thin
                for i, st in enumerate(STATUS_VALUES + ["Total"], start=8):
                    val = state["total"] if st == "Total" else state["counts"].get(st, 0)
                    ws.cell(row=i, column=1, value=st).border = thin
                    ws.cell(row=i, column=2, value=val).border = thin
                ws.cell(row=13, column=1, value="Attendance Rate %")
                ws.cell(row=13, column=2, value=round(state["rate"], 1))
                ws.column_dimensions["A"].width = 22
                ws.column_dimensions["B"].width = 28
                wb.save(path)
                messagebox.showinfo("Export Complete", f"Excel saved:\n{path}", parent=win)
            except Exception as exc:
                messagebox.showerror("Export Failed", str(exc), parent=win)

        def export_pdf():
            if not state.get("counts"):
                messagebox.showinfo("No Data", "Generate the report first.", parent=win)
                return
            safe = (state["cls"] or "All").replace(" ", "_").replace("/", "-")
            default = f"Attendance_Report_{safe}_{state['start']}_{state['end']}.pdf"
            path = _ask_save_path(win, default, "pdf")
            if not path:
                return
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.lib.units import mm
                from reportlab.platypus import (
                    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable,
                )
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

                doc = SimpleDocTemplate(
                    path, pagesize=A4,
                    leftMargin=18 * mm, rightMargin=18 * mm,
                    topMargin=16 * mm, bottomMargin=16 * mm,
                )
                styles = getSampleStyleSheet()
                title_style = ParagraphStyle(
                    "T", parent=styles["Heading1"], fontSize=14,
                    textColor=colors.HexColor("#0F172A"), spaceAfter=4,
                )
                sub = ParagraphStyle(
                    "S", parent=styles["Normal"], fontSize=9,
                    textColor=colors.HexColor("#64748B"), spaceAfter=6,
                )
                story = []
                story.append(Paragraph("Attendance Report", title_style))
                story.append(Paragraph("AR School Management System", sub))
                story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#E2E8F0")))
                story.append(Spacer(1, 8))
                story.append(Paragraph(
                    f"<b>Period:</b> {state['label']}<br/>"
                    f"<b>Class:</b> {state['cls']}<br/>"
                    f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                    styles["Normal"],
                ))
                story.append(Spacer(1, 12))
                data = [
                    ["Status", "Count"],
                    ["Present", state["counts"]["Present"]],
                    ["Absent", state["counts"]["Absent"]],
                    ["Leave", state["counts"]["Leave"]],
                    ["Late", state["counts"]["Late"]],
                    ["Total", state["total"]],
                    ["Attendance Rate", f"{state['rate']:.1f}%"],
                ]
                t = Table(data, colWidths=[180, 120])
                t.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0284C7")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                    ("BACKGROUND", (0, 1), (-1, -2), colors.HexColor("#F8FAFC")),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E2E8F0")),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]))
                story.append(t)
                story.append(Spacer(1, 14))
                story.append(Paragraph(
                    "System-generated report from live attendance records.", sub,
                ))
                doc.build(story)
                messagebox.showinfo("Export Complete", f"PDF saved:\n{path}", parent=win)
            except Exception as exc:
                messagebox.showerror("Export Failed", str(exc), parent=win)

        btns = tk.Frame(body, bg=theme.SILVER)
        btns.pack(fill=tk.X)
        theme.primary_button(btns, "↻ Generate", generate).pack(side=tk.LEFT, padx=(0, 8))
        theme.primary_button(btns, "📁 Excel", export_excel, bg=theme.SUCCESS).pack(side=tk.LEFT, padx=(0, 8))
        theme.primary_button(btns, "📄 PDF", export_pdf, bg=theme.SLATE).pack(side=tk.LEFT)

    def _range_report_window(self, title, monthly):
        # Back-compat shim
        self._open_unified_reports()


    def _open_settings_dialog(self):
        settings = get_settings()
        win = tk.Toplevel(self.win)
        win.title("Attendance Timing Settings")
        win.geometry("380x300")
        win.config(bg=theme.WHITE)
        tk.Label(win, text="Attendance Timing", font=theme.FONT_H1, bg=theme.WHITE).pack(pady=(12, 6))
        tk.Label(
            win, text="Values are stored in System Settings and apply app-wide.",
            font=theme.FONT_SMALL, bg=theme.WHITE, fg=theme.TEXT_MUTED,
        ).pack(pady=(0, 8))

        form = tk.Frame(win, bg=theme.WHITE)
        form.pack(pady=4, padx=16, fill=tk.X)
        tk.Label(form, text="School Start (HH:MM):", bg=theme.WHITE, font=theme.FONT_BODY).grid(
            row=0, column=0, sticky="e", padx=6, pady=8)
        ent_start = tk.Entry(form, width=12, font=theme.FONT_BODY)
        ent_start.insert(0, settings["start_time"])
        ent_start.grid(row=0, column=1, pady=8, sticky="w")

        tk.Label(form, text="Late Cutoff (HH:MM):", bg=theme.WHITE, font=theme.FONT_BODY).grid(
            row=1, column=0, sticky="e", padx=6, pady=8)
        ent_late = tk.Entry(form, width=12, font=theme.FONT_BODY)
        ent_late.insert(0, settings.get("late_threshold_time") or "08:15")
        ent_late.grid(row=1, column=1, pady=8, sticky="w")

        tk.Label(form, text="Closing Time (HH:MM):", bg=theme.WHITE, font=theme.FONT_BODY).grid(
            row=2, column=0, sticky="e", padx=6, pady=8)
        ent_close = tk.Entry(form, width=12, font=theme.FONT_BODY)
        ent_close.insert(0, settings["closing_time"])
        ent_close.grid(row=2, column=1, pady=8, sticky="w")

        def save():
            try:
                datetime.strptime(ent_start.get().strip(), "%H:%M")
                datetime.strptime(ent_close.get().strip(), "%H:%M")
                datetime.strptime(ent_late.get().strip(), "%H:%M")
            except ValueError:
                messagebox.showerror(
                    "Invalid Input", "All times must be HH:MM (e.g. 08:15).", parent=win,
                )
                return
            set_settings(
                ent_start.get().strip(),
                ent_close.get().strip(),
                late_threshold_time=ent_late.get().strip(),
            )
            messagebox.showinfo("Saved", "Attendance timing updated in System Settings.", parent=win)
            win.destroy()
            self.win.destroy()

        theme.primary_button(win, "💾 Save", save).pack(pady=14)


def launch_attendance_window(parent, user_role, current_user):
    return AttendanceWindow(parent, user_role, current_user)
